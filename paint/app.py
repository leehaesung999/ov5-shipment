# -*- coding: utf-8 -*-
"""색칠 (창고 레이아웃).

흐름: ERP Item_*.xlsx 원본 업로드 → 자동 편집(고정+Item+non-OV 필터) → 레이아웃 색칠
    → 편집본 xlsx + 색칠본 xlsm 각각 다운로드
"""
import sys
import traceback
from pathlib import Path
from datetime import datetime

import openpyxl
import streamlit as st

HERE = Path(__file__).parent
sys.path.insert(0, str(HERE))
import fill_locations as fl  # noqa: E402

DATA = HERE / "data"
DEFAULT_LAYOUT = DATA / "기본레이아웃.xlsm"
TMP = HERE / "_tmp"
TMP.mkdir(exist_ok=True)


def edit_원본(src_path: Path, out_path: Path) -> dict:
    """ERP Item_*.xlsx 원본에서 색칠용 편집본을 만든다.
    필터:
      - 보관타입 == '고정로케이션'
      - Item code 있음
      - 로케이션이 'OV'로 시작하지 않음
    헤더는 원본 그대로 유지. 통과된 행만 새 워크북에 씀.
    """
    wb_src = openpyxl.load_workbook(src_path, data_only=True)
    ws_src = wb_src.active

    header = [c.value for c in ws_src[1]]

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = ws_src.title or "sheet"
    ws_out.append(header)

    total = kept = 0
    for row in ws_src.iter_rows(min_row=2, values_only=True):
        if not row or not row[2]:
            continue
        total += 1
        loc = str(row[2]).strip()
        보관 = str(row[3] or "").strip()
        item_code = row[6] if len(row) > 6 else None
        if 보관 != "고정로케이션":
            continue
        if item_code is None:
            continue
        if loc.upper().startswith("OV"):
            continue
        ws_out.append(list(row))
        kept += 1

    wb_src.close()
    wb_out.save(out_path)
    wb_out.close()
    return {"총행수": total, "통과행수": kept, "결과파일": str(out_path)}


st.title("🎨 창고 레이아웃 색칠")
st.caption(
    "ERP `Item_*.xlsx` 원본을 업로드하면 자동으로 **편집본**(고정로케이션 · Item code 있음 · OV 제외) "
    "과 **색칠본**을 생성합니다."
)

up_src = st.file_uploader(
    "① 원본 데이터 xlsx (예: `Item_20260723105335.xlsx`)",
    type=["xlsx"],
    help="ERP에서 받은 그대로 업로드하면 자동 필터 후 색칠까지 처리",
)

with st.expander("옵션 — 레이아웃 파일 · 시트 · 디자인", expanded=False):
    up_layout = st.file_uploader(
        "레이아웃 파일 xlsx / xlsm (선택, 미업로드 시 내장 기본 레이아웃 사용)",
        type=["xlsx", "xlsm"],
    )
    sheet_name = st.text_input(
        "작업할 시트 이름 (비우면 '재고조사' 또는 첫 시트)",
        value="",
    )
    do_style = st.checkbox("시각 디자인 적용", value=True)

# 레이아웃 상태
if not DEFAULT_LAYOUT.exists() and not up_layout:
    st.warning("내장 레이아웃이 없고 업로드도 안 됨. 옵션에서 업로드하세요.")
    layout_ready = False
elif up_layout:
    st.caption(f"레이아웃: 업로드 파일 사용 (`{up_layout.name}`)")
    layout_ready = True
else:
    st.caption(f"레이아웃: 내장 기본 레이아웃 사용 (`{DEFAULT_LAYOUT.name}`)")
    layout_ready = True

run_clicked = st.button("▶ 편집 + 색칠 실행", type="primary",
                        disabled=not (up_src and layout_ready),
                        width='stretch')

if not up_src:
    st.info("원본 xlsx 하나만 업로드하면 됩니다. 편집·색칠 자동 진행.")

if run_clicked:
    try:
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        src_path = TMP / f"_src_{up_src.name}"
        src_path.write_bytes(up_src.getvalue())

        # 1) 원본 → 편집본
        with st.spinner("원본 편집 중..."):
            edit_out = TMP / f"편집본_{ts}.xlsx"
            edit_result = edit_원본(src_path, edit_out)
        st.success(
            f"편집 완료: 원본 {edit_result['총행수']:,}행 → "
            f"편집본 {edit_result['통과행수']:,}행"
        )

        # 2) 레이아웃 준비
        if up_layout:
            layout_path = TMP / f"_layout_{up_layout.name}"
            layout_path.write_bytes(up_layout.getvalue())
        else:
            layout_path = DEFAULT_LAYOUT

        # 3) 색칠
        stem = Path(layout_path.name).stem
        suffix = Path(layout_path.name).suffix
        paint_out = TMP / f"{stem}_완성_{ts}{suffix}"

        with st.spinner("레이아웃 색칠 중..."):
            fl.run(
                edit_out, layout_path, paint_out,
                do_style=do_style,
                sheet_name=(sheet_name.strip() or None),
            )
        st.success(f"색칠 완료: {paint_out.name}")

        # 4) 두 파일 다운로드
        col1, col2 = st.columns(2)
        with col1:
            st.download_button(
                "📥 편집본 xlsx 다운로드",
                data=edit_out.read_bytes(),
                file_name=edit_out.name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width='stretch',
            )
        with col2:
            paint_mime = ("application/vnd.ms-excel.sheet.macroEnabled.12" if suffix == ".xlsm"
                          else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            st.download_button(
                "📥 색칠본 다운로드",
                data=paint_out.read_bytes(),
                file_name=paint_out.name,
                mime=paint_mime,
                width='stretch',
            )
    except Exception as e:
        st.error(f"오류: {e}")
        st.code(traceback.format_exc())

st.divider()
with st.expander("도움말"):
    st.markdown("""
- **원본**: ERP에서 받은 `Item_*.xlsx` 그대로. 첫 열이 Inventory code, 3열이 로케이션ID, 4열이 보관타입, 7~8열이 Item code/Item
- **자동 편집 규칙**:
    - `보관타입 == '고정로케이션'`
    - `Item code`가 비어있지 않음
    - 로케이션이 `OV`로 시작하지 않음
- **레이아웃 파일**: 기본은 저장소 내장. 다른 배치도를 쓰려면 옵션에서 업로드
- **시각 디자인 적용 해제**: 값만 채우고 색은 안 넣음
- **시트 이름**: 비우면 `재고조사`를 찾고 없으면 첫 시트
""")
