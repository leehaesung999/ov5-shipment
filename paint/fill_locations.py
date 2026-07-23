"""
색칠.xlsx 레이아웃에 원본 로케이션-Item 매핑을 자동으로 채우고 시각적으로 꾸미는 스크립트.

사용법:
    python fill_locations.py 원본.xlsx 색칠.xlsx
    python fill_locations.py 원본.xlsx 색칠.xlsx -o 결과.xlsx

옵션:
    -o, --output    출력 파일 경로 (기본값: 색칠_완성.xlsx)
    --no-style      시각 디자인 적용 생략 (값만 채움)
"""

from __future__ import annotations

import argparse
import re
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


LOC_PATTERN = re.compile(r"^[A-Z]+\d+-\d+-\d+$")
HEADER_PATTERN = re.compile(r"^[A-Z]+\d+$")
AISLE_LABEL = "통로"


PALETTE = {
    "header_bg":     "1F3864",  # 진한 네이비 — 그룹 헤더
    "header_fg":     "FFFFFF",
    "aisle_bg":      "D9D9D9",  # 연회색 — 통로
    "aisle_fg":      "595959",
    "row_label_bg":  "BDD7EE",  # 연파랑 — '10' 같은 행 라벨
    "loc_top_bg":    "FFF2CC",  # 연노랑 — 위 랙 로케이션 ID
    "loc_bot_bg":    "E2EFDA",  # 연초록 — 아래 랙 로케이션 ID
    "item_filled_bg": "FFE699",  # 노랑 — 매핑된 상품명
    "item_empty_bg":  "F2F2F2",  # 옅은 회색 — 매핑 없음
    "border":        "808080",
}


def build_location_map(src_path: Path) -> dict[str, list[tuple]]:
    """원본 파일에서 로케이션ID → [(item_code, item_name), ...] 매핑 생성."""
    wb = load_workbook(src_path, data_only=True)
    ws = wb.active
    mapping: dict[str, list[tuple]] = defaultdict(list)

    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 8:
            continue
        loc_id = row[2]
        item_code = row[6]
        item_name = row[7]
        if loc_id and item_name:
            mapping[str(loc_id).strip()].append((item_code, str(item_name).strip()))

    return dict(mapping)


def find_aisle_rows(ws: Worksheet) -> set[int]:
    """첫 컬럼이 '통로'인 행 번호 집합."""
    aisles = set()
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and v.strip() == AISLE_LABEL:
            aisles.add(r)
    return aisles


def find_header_rows(ws: Worksheet) -> set[int]:
    """첫 컬럼이 그룹 헤더(A7, A11 등)인 행 번호 집합."""
    headers = set()
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and HEADER_PATTERN.match(v.strip()):
            headers.add(r)
    return headers


def fill_items(
    ws: Worksheet,
    loc_map: dict[str, list[tuple]],
    aisle_rows: set[int],
) -> tuple[int, int, int]:
    """
    로케이션 ID 셀의 인접 행(통로 반대편)에 상품명을 기록한다.
    상품명 자리에 기존 수식/값이 있으면 모두 지운 뒤 새로 채운다.
    반환: (총 로케이션 수, 매핑된 수, 중복(여러 상품) 수)
    """
    # 1단계: 외부 워크북([1]붙여넣기) 참조 수식을 모두 제거 — 어디에 남아있든 정리
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.startswith("=") and "'[" in v:
                cell.value = None

    # 2단계: 상품명 자리(통로의 위/아래 두 번째 행)를 통째로 비운다 — 이전에 직접 입력된 텍스트도 새로 채울 수 있도록
    item_rows = set()
    for ar in aisle_rows:
        item_rows.add(ar - 2)
        item_rows.add(ar + 2)
    max_col = ws.max_column
    for ir in item_rows:
        if ir < 1 or ir > ws.max_row:
            continue
        for c in range(2, max_col + 1):  # col 1은 그룹 헤더(A7 등)이므로 보존
            ws.cell(ir, c).value = None

    total = matched = multi = 0
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if not (isinstance(v, str) and LOC_PATTERN.match(v)):
                continue
            total += 1
            r, c = cell.row, cell.column

            if (r + 1) in aisle_rows:
                target_row = r - 1  # 위 랙
            elif (r - 1) in aisle_rows:
                target_row = r + 1  # 아래 랙
            else:
                continue
            if target_row < 1:
                continue

            target = ws.cell(target_row, c)
            items = loc_map.get(v.strip())
            if items:
                matched += 1
                if len(items) > 1:
                    multi += 1
                    target.value = "\n".join(name for _, name in items)
                else:
                    target.value = items[0][1]
            else:
                target.value = None  # 외부 참조 수식 등 잔여 데이터 정리

    return total, matched, multi


def apply_design(
    ws: Worksheet,
    loc_map: dict[str, list[tuple]],
    aisle_rows: set[int],
    header_rows: set[int],
) -> None:
    """창고 레이아웃이 한눈에 들어오도록 셀 스타일을 적용."""
    # 원본에 칠해져 있던 색(하늘/연두/노랑 등)을 먼저 모두 제거 — 새 디자인으로 통일
    blank_fill = PatternFill(fill_type=None)
    for row in ws.iter_rows():
        for cell in row:
            cell.fill = blank_fill

    side = Side(border_style="thin", color=PALETTE["border"])
    border = Border(left=side, right=side, top=side, bottom=side)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    header_font = Font(name="맑은 고딕", size=14, bold=True, color=PALETTE["header_fg"])
    aisle_font = Font(name="맑은 고딕", size=10, italic=True, color=PALETTE["aisle_fg"])
    row_label_font = Font(name="맑은 고딕", size=10, bold=True)
    loc_font = Font(name="맑은 고딕", size=9)
    item_font = Font(name="맑은 고딕", size=9, bold=True)
    item_empty_font = Font(name="맑은 고딕", size=9, color="9C9C9C")

    fills = {k: PatternFill("solid", fgColor=v) for k, v in PALETTE.items() if k.endswith("_bg")}

    max_col = ws.max_column

    for row in ws.iter_rows():
        for cell in row:
            r, c = cell.row, cell.column
            v = cell.value

            if r in header_rows and c == 1:
                cell.fill = fills["header_bg"]
                cell.font = header_font
                cell.alignment = center
                cell.border = border
                continue

            if r in aisle_rows:
                if v is not None or c == 1:
                    cell.fill = fills["aisle_bg"]
                    cell.font = aisle_font
                    cell.alignment = center
                    cell.border = border
                continue

            if c == 1 and v in (10, "10"):
                cell.fill = fills["row_label_bg"]
                cell.font = row_label_font
                cell.alignment = center
                cell.border = border
                continue

            if isinstance(v, str) and LOC_PATTERN.match(v):
                is_top = (r + 1) in aisle_rows
                cell.fill = fills["loc_top_bg" if is_top else "loc_bot_bg"]
                cell.font = loc_font
                cell.alignment = center
                cell.border = border

                target_row = r - 1 if is_top else r + 1
                if target_row >= 1:
                    item_cell = ws.cell(target_row, c)
                    has_value = item_cell.value not in (None, "")
                    in_src = v.strip() in loc_map
                    if has_value or in_src:
                        item_cell.fill = fills["item_filled_bg" if has_value else "item_empty_bg"]
                        item_cell.font = item_font if has_value else item_empty_font
                        item_cell.alignment = center
                        item_cell.border = border

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 3
    ws.column_dimensions["C"].width = 3
    for c in range(4, max_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 13

    for r in range(1, ws.max_row + 1):
        if r in header_rows:
            ws.row_dimensions[r].height = 28
        elif r in aisle_rows:
            ws.row_dimensions[r].height = 18
        else:
            ws.row_dimensions[r].height = 36

    ws.freeze_panes = "D1"
    ws.sheet_view.zoomScale = 85


DEFAULT_SHEET = "재고조사"


def pick_sheet(wb, preferred: str | None = None):
    """워크북에서 작업 대상 시트를 선택. preferred가 있으면 그걸, 없으면 재고조사, 그것도 없으면 active."""
    if preferred and preferred in wb.sheetnames:
        return wb[preferred]
    if DEFAULT_SHEET in wb.sheetnames:
        return wb[DEFAULT_SHEET]
    return wb.active


def run(
    src: Path,
    layout: Path,
    out: Path,
    *,
    do_style: bool = True,
    sheet_name: str | None = None,
) -> None:
    print(f"[1/4] 원본 매핑 로드: {src.name}")
    loc_map = build_location_map(src)
    print(f"     → 매핑 {len(loc_map):,}건")

    print(f"[2/4] 색칠 레이아웃 로드: {layout.name}")
    is_macro = layout.suffix.lower() in (".xlsm", ".xltm")
    wb = load_workbook(layout, keep_vba=is_macro)
    ws = pick_sheet(wb, sheet_name)
    print(f"     → 시트 '{ws.title}' 사용{'  (매크로 보존)' if is_macro else ''}")
    aisle_rows = find_aisle_rows(ws)
    header_rows = find_header_rows(ws)
    print(f"     → 통로 행 {len(aisle_rows)}개 / 헤더 행 {len(header_rows)}개")

    print("[3/4] 상품명 채우는 중…")
    total, matched, multi = fill_items(ws, loc_map, aisle_rows)
    print(f"     → 로케이션 {total:,}개 중 {matched:,}개 매칭 (중복 {multi}개)")

    if do_style:
        print("[4/4] 디자인 적용 중…")
        apply_design(ws, loc_map, aisle_rows, header_rows)
    else:
        print("[4/4] 디자인 적용 건너뜀")

    # 색칠.xlsx에 남아 있던 외부 워크북 참조([1]붙여넣기) 정리
    if getattr(wb, "_external_links", None):
        wb._external_links = []

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"\n저장 완료: {out}")


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="원본 로케이션 데이터를 색칠.xlsx 레이아웃에 반영합니다.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    p.add_argument("source", type=Path, help="원본 데이터 (예: 0507 기본 로케이션.xlsx)")
    p.add_argument("layout", type=Path, help="레이아웃 파일 (.xlsx 또는 .xlsm)")
    p.add_argument("-o", "--output", type=Path, default=None, help="출력 파일 (기본: <레이아웃이름>_완성.확장자)")
    p.add_argument("--sheet", default=None, help=f"작업할 시트 이름 (기본: '{DEFAULT_SHEET}' 또는 첫 시트)")
    p.add_argument("--no-style", action="store_true", help="디자인 적용 생략")
    return p.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)

    if not args.source.exists():
        print(f"ERROR: 원본 파일을 찾을 수 없습니다 — {args.source}", file=sys.stderr)
        return 2
    if not args.layout.exists():
        print(f"ERROR: 레이아웃 파일을 찾을 수 없습니다 — {args.layout}", file=sys.stderr)
        return 2

    out = args.output or args.layout.with_name(args.layout.stem + "_완성" + args.layout.suffix)
    run(args.source, args.layout, out, do_style=not args.no_style, sheet_name=args.sheet)
    return 0


if __name__ == "__main__":
    sys.exit(main())
