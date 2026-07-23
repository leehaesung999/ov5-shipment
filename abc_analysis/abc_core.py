"""
ABC 재배치 분석 - 핵심 로직 (GUI/CLI 공용)

주 진입점: run_analysis(monthly_dir, master_file, output_dir, log=print) -> out_path
"""

import os
import re
import glob
import sys

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ========== CONFIG ==========
WINDOW_MONTHS          = 24          # 평균 산출 윈도우 (2년)
SUMMER                 = {6, 7, 8, 9}
WINTER                 = {11, 12, 1, 2}
PALLET_2SLOT_THRESHOLD = 0.7         # 팔레트/일 ≥ 0.7 → 2슬롯 필요
NEW_PRODUCT_WINDOW     = 12          # 최근 12개월 내 출시 = NEW
ABC_CUT_A              = 0.70
ABC_CUT_B              = 0.90
# ============================


class AnalysisError(Exception):
    """분석 중 발생한 사용자가 이해 가능한 오류."""


# -------------------------------------------------------------------
# 1. 월별 파일 로딩
# -------------------------------------------------------------------
def load_monthly_files(monthly_dir, log=print):
    files = sorted(glob.glob(os.path.join(monthly_dir, '????-??.xlsx')))
    if not files:
        raise AnalysisError(
            f'월별 데이터가 없습니다.\n경로: {monthly_dir}\n'
            f'YYYY-MM.xlsx 파일을 먼저 등록해 주세요.'
        )

    frames = []
    for f in files:
        m = re.search(r'(\d{4})-(\d{2})\.xlsx$', os.path.basename(f))
        if not m:
            continue
        y, mo = int(m.group(1)), int(m.group(2))

        df = pd.read_excel(f, sheet_name=0)
        df.columns = [str(c).strip() for c in df.columns]

        need = ['품번', '품명', '일평균출고']
        missing = [c for c in need if c not in df.columns]
        if missing:
            log(f'  [경고] {os.path.basename(f)} 필수 컬럼 누락: {missing} → 건너뜀')
            continue

        df = df[need].copy()
        df.columns = ['품번', '품명', '일평균']
        df['품번'] = _normalize_code(df['품번'])
        df['일평균'] = pd.to_numeric(df['일평균'], errors='coerce').fillna(0)
        df['년'] = y
        df['월'] = mo
        frames.append(df)
        log(f'  loaded {os.path.basename(f)}: {len(df)} 품목')

    if not frames:
        raise AnalysisError('유효한 월별 파일이 하나도 없습니다.')

    return pd.concat(frames, ignore_index=True)


# -------------------------------------------------------------------
# 2. 평균 계산 (연 / 하절기 / 동절기)
# -------------------------------------------------------------------
def compute_averages(long_df, ref_year, ref_month):
    long_df = long_df.copy()
    long_df['diff'] = (ref_year - long_df['년']) * 12 + (ref_month - long_df['월'])
    win = long_df[(long_df['diff'] >= 0) & (long_df['diff'] < WINDOW_MONTHS)]

    rows = []
    for 품번, g in win.groupby('품번'):
        annual = g['일평균'].mean()
        summer_g = g[g['월'].isin(SUMMER)]
        winter_g = g[g['월'].isin(WINTER)]
        summer = summer_g['일평균'].mean() if len(summer_g) else np.nan
        winter = winter_g['일평균'].mean() if len(winter_g) else np.nan

        n_months = len(g)
        rows.append({
            '품번': 품번,
            '품명': g['품명'].mode().iloc[0] if len(g['품명'].mode()) else g['품명'].iloc[0],
            '데이터월수': n_months,
            '연평균': round(annual, 2),
            '하절기일평균': round(summer, 2) if pd.notna(summer) else round(annual, 2),
            '동절기일평균': round(winter, 2) if pd.notna(winter) else round(annual, 2),
            '하절기출처': 'actual' if pd.notna(summer) else 'annual_fallback',
            '동절기출처': 'actual' if pd.notna(winter) else 'annual_fallback',
            '기간구분': '최근2년' if n_months >= WINDOW_MONTHS else f'출시후({n_months}M)',
        })
    return pd.DataFrame(rows)


# -------------------------------------------------------------------
# 3. 마스터 조인 + ABC 분류 + 2슬롯 판정
# -------------------------------------------------------------------
def classify_abc(avg_df, master_df, ref_year, ref_month):
    df = avg_df.merge(
        master_df[['품번', '하대(박스/팔레트)', '출시월', '현재로케']],
        on='품번', how='left'
    )

    # 품명 master 우선
    df = df.merge(
        master_df[['품번', '품명']].rename(columns={'품명': '품명_m'}),
        on='품번', how='left'
    )
    df['품명'] = df['품명_m'].fillna(df['품명'])
    df = df.drop(columns=['품명_m'])

    # 하대 없으면 1 (팔레트 환산 불가 → 박스/일 그대로)
    df['하대(박스/팔레트)'] = df['하대(박스/팔레트)'].fillna(1).replace(0, 1)

    # ABC
    df = df.sort_values('연평균', ascending=False).reset_index(drop=True)
    df['순위'] = df.index + 1
    total = df['연평균'].sum()
    df['출고비중'] = df['연평균'] / total if total > 0 else 0
    df['누적비율'] = df['출고비중'].cumsum()

    def grade(cum):
        if cum <= ABC_CUT_A: return 'A'
        if cum <= ABC_CUT_B: return 'B'
        return 'C'
    df['ABC'] = df['누적비율'].apply(grade)

    # 팔레트/일
    df['연팔레트일']   = (df['연평균']       / df['하대(박스/팔레트)']).round(3)
    df['하절기팔레트일'] = (df['하절기일평균'] / df['하대(박스/팔레트)']).round(3)
    df['동절기팔레트일'] = (df['동절기일평균'] / df['하대(박스/팔레트)']).round(3)

    # 2슬롯
    df['2슬롯연']    = np.where(df['연팔레트일']    >= PALLET_2SLOT_THRESHOLD, 'Y', 'N')
    df['2슬롯하절기'] = np.where(df['하절기팔레트일'] >= PALLET_2SLOT_THRESHOLD, 'Y', 'N')
    df['2슬롯동절기'] = np.where(df['동절기팔레트일'] >= PALLET_2SLOT_THRESHOLD, 'Y', 'N')

    # 신제품 (최근 12개월 내 출시). 출시월 포맷: YYMM (예: 2505 = 2025-05)
    def is_new(rm):
        if pd.isna(rm): return 'N'
        try:
            rm = int(rm)
            ry, rmo = 2000 + rm // 100, rm % 100
            diff = (ref_year - ry) * 12 + (ref_month - rmo)
            return 'NEW' if 0 <= diff < NEW_PRODUCT_WINDOW else 'N'
        except Exception:
            return 'N'
    df['신제품'] = df['출시월'].apply(is_new)

    cols = ['순위', '품번', '품명', 'ABC', '신제품', '출시월', '기간구분', '데이터월수',
            '누적비율', '출고비중',
            '연평균', '하절기일평균', '동절기일평균',
            '하대(박스/팔레트)',
            '연팔레트일', '하절기팔레트일', '동절기팔레트일',
            '현재로케', '2슬롯연', '2슬롯하절기', '2슬롯동절기',
            '하절기출처', '동절기출처']
    return df[cols]


# -------------------------------------------------------------------
# 4. 히트맵 시트 작성
# -------------------------------------------------------------------
def _normalize_code(s):
    """품번을 문자열로 정규화 (엑셀이 숫자로 읽어 '.0'이 붙는 것 제거)."""
    return (
        s.astype(str)
         .str.strip()
         .str.replace(r'\.0+$', '', regex=True)
         .str.replace(r'^nan$', '', regex=True)
    )


_LOC_RE = re.compile(r'^([A-Z]\d+)-(\d+)-(\d+)$')

def parse_location(loc):
    if not isinstance(loc, str):
        return None, None
    m = _LOC_RE.match(loc.strip())
    if not m:
        return None, None
    return m.group(1), int(m.group(2))


# 7단계 색상 팔레트 (낮음 → 높음, 팔레트/일 기준)
HEATMAP_TIERS = [
    (0.00, 'FFF2F2F2'),   # 연회색 (비어있거나 거의 0)
    (0.20, 'FFD9E7F5'),   # 아주연한파랑
    (0.50, 'FFBDD7EE'),   # 연파랑
    (1.00, 'FFFFD966'),   # 노랑
    (2.50, 'FFF4B084'),   # 주황
    (5.00, 'FFC00000'),   # 빨강
    (9e9 , 'FF8B0000'),   # 진한빨강 (상한 없음)
]

# 고정 색상
FILL_TITLE   = PatternFill('solid', fgColor='FF305496')  # 구역 제목 파란색
FILL_HEADER  = PatternFill('solid', fgColor='FFF2F2F2')  # 로케/품목 행 회색
FILL_CORR    = PatternFill('solid', fgColor='FF1F1F1F')  # 통로 검정
FONT_WHITE   = Font(color='FFFFFFFF', bold=True, size=11)
FONT_HDR     = Font(color='FF606060', size=9)
FONT_NAME    = Font(color='FF000000', size=9)
FONT_NAME_HI = Font(color='FFFFFFFF', size=9, bold=True)  # 진한빨강 위 흰 글씨


def _tier_fill(val):
    """값에 따라 7단계 색상 중 하나 반환."""
    if val is None or pd.isna(val) or val <= 0:
        return None, False
    for thr, color in HEATMAP_TIERS:
        if val < thr:
            dark_bg = color in ('FFC00000', 'FF8B0000')
            return PatternFill('solid', fgColor=color), dark_bg
    # 마지막 tier
    return PatternFill('solid', fgColor=HEATMAP_TIERS[-1][1]), True


def write_heatmap(ws, abc_df, value_col, title):
    grid = {}  # (zone, pos) -> {val, name, abc, role}
    for _, r in abc_df.iterrows():
        # 메인 자리
        zone, pos = parse_location(r['현재로케'])
        if zone is not None:
            grid[(zone, pos)] = {
                'val':  r[value_col],
                'name': r['품명'],
                'abc':  r['ABC'],
                'role': 'main',
            }
        # 보조 자리
        sub_loc = r.get('보조로케', '') if '보조로케' in abc_df.columns else ''
        if isinstance(sub_loc, str) and sub_loc and sub_loc != '—':
            sz, sp = parse_location(sub_loc)
            if sz is not None:
                grid[(sz, sp)] = {
                    'val':  r[value_col],
                    'name': r['품명'] + ' #',  # 보조 표시
                    'abc':  r['ABC'],
                    'role': 'sub',
                }

    zones = sorted({z for z, _ in grid.keys()},
                   key=lambda z: int(re.sub(r'\D', '', z) or 0))

    # 최상단 타이틀
    ws.cell(row=1, column=1, value=f'히트맵 - 마주보는 랙 구조 ({title})').font = Font(bold=True, size=14)

    row = 2
    for zone in zones:
        # ── 구역 제목 행 (파란색) ────────────────────────
        for col in range(1, 52):
            c = ws.cell(row=row, column=col, value=zone if col == 1 else ('max 팔레트/일' if col == 2 else None))
            c.fill = FILL_TITLE
            c.font = FONT_WHITE if col <= 2 else Font(color='FFFFFFFF')
        row += 1

        # ── 앞 로케 ─────────────────────────────────
        ws.cell(row=row, column=1, value='앞 로케').font = FONT_HDR
        ws.cell(row=row, column=1).fill = FILL_HEADER
        for i, pos in enumerate(range(1, 51), start=2):
            c = ws.cell(row=row, column=i, value=f'{zone}-{pos:02d}-10')
            c.font = FONT_HDR
            c.fill = FILL_HEADER
            c.alignment = Alignment(horizontal='center')
        row += 1

        # ── 앞 품목 (색상) ──────────────────────────────
        ws.cell(row=row, column=1, value='앞 품목').font = FONT_HDR
        ws.cell(row=row, column=1).fill = FILL_HEADER
        for i, pos in enumerate(range(1, 51), start=2):
            cell = grid.get((zone, pos))
            c = ws.cell(row=row, column=i)
            if cell:
                c.value = cell['name']
                fill, dark = _tier_fill(cell['val'])
                if fill:
                    c.fill = fill
                    c.font = FONT_NAME_HI if dark else FONT_NAME
                else:
                    c.font = FONT_NAME
                c.alignment = Alignment(horizontal='center', wrap_text=True)
        row += 1

        # ── 통로 (검정) ─────────────────────────────────
        for col in range(1, 52):
            c = ws.cell(row=row, column=col, value='통로' if col == 1 else None)
            c.fill = FILL_CORR
            c.font = Font(color='FFFFFFFF', size=9)
        row += 1

        # ── 뒤 품목 (색상) ──────────────────────────────
        ws.cell(row=row, column=1, value='뒤 품목').font = FONT_HDR
        ws.cell(row=row, column=1).fill = FILL_HEADER
        for i, pos in enumerate(range(100, 50, -1), start=2):
            cell = grid.get((zone, pos))
            c = ws.cell(row=row, column=i)
            if cell:
                c.value = cell['name']
                fill, dark = _tier_fill(cell['val'])
                if fill:
                    c.fill = fill
                    c.font = FONT_NAME_HI if dark else FONT_NAME
                else:
                    c.font = FONT_NAME
                c.alignment = Alignment(horizontal='center', wrap_text=True)
        row += 1

        # ── 뒤 로케 ─────────────────────────────────
        ws.cell(row=row, column=1, value='뒤 로케').font = FONT_HDR
        ws.cell(row=row, column=1).fill = FILL_HEADER
        for i, pos in enumerate(range(100, 50, -1), start=2):
            c = ws.cell(row=row, column=i, value=f'{zone}-{pos:02d}-10')
            c.font = FONT_HDR
            c.fill = FILL_HEADER
            c.alignment = Alignment(horizontal='center')
        row += 1

        # 구역 간 공백 3행
        row += 3

    # 범례 추가
    _write_legend(ws, start_row=row + 1)

    # 열 너비 & 행 고정
    ws.column_dimensions['A'].width = 10
    for i in range(2, 52):
        ws.column_dimensions[get_column_letter(i)].width = 13
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'B2'


def _write_legend(ws, start_row):
    """색상 범례 출력."""
    ws.cell(row=start_row, column=1, value='범례 (팔레트/일)').font = Font(bold=True, size=11)
    labels = [
        ('0',           'FFF2F2F2'),
        ('0 ~ 0.2',     'FFD9E7F5'),
        ('0.2 ~ 0.5',   'FFBDD7EE'),
        ('0.5 ~ 1.0',   'FFFFD966'),
        ('1.0 ~ 2.5',   'FFF4B084'),
        ('2.5 ~ 5.0',   'FFC00000'),
        ('5.0 이상',     'FF8B0000'),
    ]
    for i, (label, color) in enumerate(labels):
        c = ws.cell(row=start_row + 1, column=i + 1, value=label)
        c.fill = PatternFill('solid', fgColor=color)
        c.alignment = Alignment(horizontal='center')
        dark = color in ('FFC00000', 'FF8B0000')
        c.font = Font(color='FFFFFFFF' if dark else 'FF000000', bold=True, size=9)


# -------------------------------------------------------------------
# 5. 보조 시트 (요약 / 빅무브 / 하절기 재배치 / 변경후 노트)
# -------------------------------------------------------------------
def write_summary(ws, abc_df, ref_year, ref_month, n_months_avail):
    ws.cell(row=1, column=1,
            value=f'ABC + 재배치 요약 ({ref_year}-{ref_month:02d} 기준, 가용월수 {n_months_avail})'
            ).font = Font(bold=True, size=14)

    ws.cell(row=3, column=1, value='-- 전체 --').font = Font(bold=True)
    ws.cell(row=4, column=1, value='전체 품목')
    ws.cell(row=4, column=2, value=len(abc_df))
    ws.cell(row=5, column=1, value='신제품 (최근 12M)')
    ws.cell(row=5, column=2, value=int((abc_df['신제품']=='NEW').sum()))

    ws.cell(row=7, column=1, value='-- ABC (연평균 기준) --').font = Font(bold=True)
    ws.cell(row=8, column=1, value='등급')
    ws.cell(row=8, column=2, value='품목수')
    ws.cell(row=8, column=3, value='신제품포함')
    ws.cell(row=8, column=4, value='출고비중')
    ws.cell(row=8, column=5, value='2슬롯연필요')
    for k, g in enumerate(['A','B','C']):
        sub = abc_df[abc_df['ABC']==g]
        ws.cell(row=9+k, column=1, value=g)
        ws.cell(row=9+k, column=2, value=len(sub))
        ws.cell(row=9+k, column=3, value=int((sub['신제품']=='NEW').sum()))
        ws.cell(row=9+k, column=4, value=f"{sub['출고비중'].sum()*100:.1f}%")
        ws.cell(row=9+k, column=5, value=int((sub['2슬롯연']=='Y').sum()))

    ws.cell(row=13, column=1, value='-- 2슬롯 필요 합계 --').font = Font(bold=True)
    ws.cell(row=14, column=1, value='2슬롯연');    ws.cell(row=14, column=2, value=int((abc_df['2슬롯연']=='Y').sum()))
    ws.cell(row=15, column=1, value='2슬롯하절기'); ws.cell(row=15, column=2, value=int((abc_df['2슬롯하절기']=='Y').sum()))
    ws.cell(row=16, column=1, value='2슬롯동절기'); ws.cell(row=16, column=2, value=int((abc_df['2슬롯동절기']=='Y').sum()))

    ws.cell(row=18, column=1, value='-- 하절기 데이터 출처 --').font = Font(bold=True)
    ws.cell(row=19, column=1, value='actual');          ws.cell(row=19, column=2, value=int((abc_df['하절기출처']=='actual').sum()))
    ws.cell(row=20, column=1, value='annual_fallback'); ws.cell(row=20, column=2, value=int((abc_df['하절기출처']=='annual_fallback').sum()))

    ws.cell(row=22, column=1, value='-- 동절기 데이터 출처 --').font = Font(bold=True)
    ws.cell(row=23, column=1, value='actual');          ws.cell(row=23, column=2, value=int((abc_df['동절기출처']=='actual').sum()))
    ws.cell(row=24, column=1, value='annual_fallback'); ws.cell(row=24, column=2, value=int((abc_df['동절기출처']=='annual_fallback').sum()))

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 14


def write_big_move(ws, abc_df):
    """A 등급인데 메인 A구역(A18/A19) 밖에 있는 품목 → 이동 추천."""
    ws.cell(row=1, column=1, value='빅무브 이동 계획 (A급 → A18/A19 집결)').font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value='목적: A등급을 메인 A구역에 모아 피킹 동선 단축'
            ).font = Font(italic=True, color='555555')

    main = {'A18', 'A19'}
    target = abc_df[abc_df['ABC']=='A'].copy()
    target['zone'] = target['현재로케'].apply(lambda s: parse_location(s)[0])
    target = target[~target['zone'].isin(main) & target['zone'].notna()]
    target = target.sort_values(['zone', '연평균'], ascending=[True, False])

    hdr = ['품번', '품명', 'ABC', '현재로케', '연팔레트일', '하절기팔레트일', '동절기팔레트일']
    for i, h in enumerate(hdr, start=1):
        c = ws.cell(row=4, column=i, value=h); c.font = Font(bold=True)
        c.fill = PatternFill('solid', fgColor='DDDDDD')

    for i, (_, r) in enumerate(target.iterrows(), start=5):
        ws.cell(row=i, column=1, value=r['품번'])
        ws.cell(row=i, column=2, value=r['품명'])
        ws.cell(row=i, column=3, value=r['ABC'])
        ws.cell(row=i, column=4, value=r['현재로케'])
        ws.cell(row=i, column=5, value=r['연팔레트일'])
        ws.cell(row=i, column=6, value=r['하절기팔레트일'])
        ws.cell(row=i, column=7, value=r['동절기팔레트일'])

    widths = [12, 40, 6, 14, 12, 14, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_seasonal_reloc(ws, abc_df, season_col, palet_col, title):
    """
    해당 계절 2슬롯 필요 품목 리스트.
    - 보조로케가 이미 확보됐으면 그대로 표시 (거리 1)
    - 연 기준 비2슬롯이지만 계절만 2슬롯인 품목 = "임시 보조 추가 필요" → 인접 빈자리 추천
    """
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14)
    ws.cell(row=2, column=1,
            value=f'기준: {palet_col} ≥ {PALLET_2SLOT_THRESHOLD} · 보조 자리 확인'
            ).font = Font(italic=True, color='555555')

    # 사용 중 로케 set (메인 + 이미 확보된 보조)
    used = set(abc_df['현재로케'].dropna().tolist())
    if '보조로케' in abc_df.columns:
        used |= set(s for s in abc_df['보조로케'].dropna() if s and s != '—')

    need = abc_df[(abc_df[palet_col] >= PALLET_2SLOT_THRESHOLD)].copy()
    need = need.sort_values(palet_col, ascending=False)

    hdr = ['품번', '품명', 'ABC', '현재로케', palet_col, '연팔레트일',
           '확보된 보조', '추가 인접 빈자리', '거리', '상태']
    for i, h in enumerate(hdr, start=1):
        c = ws.cell(row=4, column=i, value=h); c.font = Font(bold=True)
        c.fill = PatternFill('solid', fgColor='DDDDDD')

    for i, (_, r) in enumerate(need.iterrows(), start=5):
        ws.cell(row=i, column=1, value=r['품번'])
        ws.cell(row=i, column=2, value=r['품명'])
        ws.cell(row=i, column=3, value=r['ABC'])
        ws.cell(row=i, column=4, value=r['현재로케'])
        ws.cell(row=i, column=5, value=r[palet_col])
        ws.cell(row=i, column=6, value=r['연팔레트일'])

        sub = r.get('보조로케', '') if '보조로케' in abc_df.columns else ''
        ws.cell(row=i, column=7, value=sub if sub else '—')

        # 추가 보조가 더 필요한 경우 인접 빈자리 추천
        if sub and sub not in ('', '—'):
            ws.cell(row=i, column=8, value='—')
            ws.cell(row=i, column=10, value='OK (옆자리 확보됨)')
        else:
            rec_loc, dist = _suggest_adjacent(r['현재로케'], used)
            ws.cell(row=i, column=8, value=rec_loc if rec_loc else '—')
            ws.cell(row=i, column=9, value=dist if dist is not None else '')
            ws.cell(row=i, column=10, value='보조 미확보 — 검토 필요')

    widths = [12, 36, 6, 14, 12, 12, 14, 16, 8, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _suggest_adjacent(current_loc, used_set):
    zone, pos = parse_location(current_loc)
    if zone is None:
        return None, None
    for d in range(1, 50):
        for cand_pos in (pos - d, pos + d):
            if 1 <= cand_pos <= 100:
                cand = f'{zone}-{cand_pos:02d}-10'
                if cand not in used_set:
                    return cand, d
    return None, None


# -------------------------------------------------------------------
# 6. 보조 슬롯 최적화 (메인 옆자리 확보)
# -------------------------------------------------------------------
def optimize_secondary_slots(abc_df, log=print):
    """
    2슬롯 필요 품목의 보조 자리를 메인 옆(거리 1)에 배치하도록 최적화.
    필요하면 1슬롯 옆 품목을 다른 빈 자리로 이동하거나, 메인 자체를 옮김.

    Returns: (updated_abc_df, moves)
        updated_abc_df: '현재로케' / '보조로케' 갱신됨
        moves: [{품번, 품명, 이전로케, 신규로케, 사유, 종류}]
    """
    abc_df = abc_df.copy()

    # 점유 상태: (zone, pos) -> {'code': 품번, 'role': 'main' | 'sub'}
    grid = {}
    main_loc = {}  # 품번 -> (zone, pos)
    abc_by_code = {}

    for _, r in abc_df.iterrows():
        abc_by_code[r['품번']] = r
        z, p = parse_location(r['현재로케'])
        if z is None:
            continue
        grid[(z, p)] = {'code': r['품번'], 'role': 'main'}
        main_loc[r['품번']] = (z, p)

    secondary_map = {}  # 품번 -> '보조로케' 문자열
    moves = []

    needs = abc_df[abc_df['2슬롯연'] == 'Y'].sort_values('연팔레트일', ascending=False)

    for _, r in needs.iterrows():
        code = r['품번']
        if code not in main_loc:
            continue
        z, p = main_loc[code]

        adj_pos = [pp for pp in (p - 1, p + 1) if 1 <= pp <= 100]
        chosen_sub = None

        # ── Phase 1: 빈 옆자리
        for ap in adj_pos:
            if (z, ap) not in grid:
                chosen_sub = ap
                grid[(z, ap)] = {'code': code, 'role': 'sub'}
                break

        # ── Phase 2: 옆 1슬롯 품목 swap
        if chosen_sub is None:
            swap_cands = []
            for ap in adj_pos:
                occ = grid.get((z, ap))
                if occ is None or occ['code'] == code:
                    continue
                if occ['role'] == 'sub':
                    continue  # 다른 품목의 보조 자리는 못 건드림
                occ_row = abc_by_code.get(occ['code'])
                if occ_row is None or occ_row['2슬롯연'] == 'Y':
                    continue
                swap_cands.append((ap, occ['code'], occ_row['연팔레트일']))

            if swap_cands:
                swap_cands.sort(key=lambda x: x[2])  # 출고량 적은 순
                ap, occ_code, _ = swap_cands[0]
                new_pos = _find_empty_slot(grid, prefer_zone=z)
                if new_pos:
                    z2, p2 = new_pos
                    old_loc = f'{z}-{ap:02d}-10'
                    new_loc = f'{z2}-{p2:02d}-10'
                    grid[(z2, p2)] = {'code': occ_code, 'role': 'main'}
                    grid[(z, ap)] = {'code': code, 'role': 'sub'}
                    main_loc[occ_code] = (z2, p2)
                    chosen_sub = ap
                    moves.append({
                        '품번': occ_code,
                        '품명': abc_by_code[occ_code]['품명'],
                        '이전로케': old_loc,
                        '신규로케': new_loc,
                        '사유': f'{code}의 보조 자리 확보',
                        '종류': '1슬롯 이동',
                    })

        # ── Phase 3: 메인 자체를 인접 빈 쌍이 있는 곳으로 이동
        if chosen_sub is None:
            pair = _find_empty_pair(grid, prefer_zone=z)
            if pair:
                zp, pp1, pp2 = pair
                old_main = f'{z}-{p:02d}-10'
                # 기존 메인 자리 비움 (해당 품목 것이 맞을 때만)
                if grid.get((z, p), {}).get('code') == code:
                    del grid[(z, p)]
                grid[(zp, pp1)] = {'code': code, 'role': 'main'}
                grid[(zp, pp2)] = {'code': code, 'role': 'sub'}
                main_loc[code] = (zp, pp1)
                chosen_sub = pp2
                # secondary_map에는 zone+pos 저장
                secondary_map[code] = f'{zp}-{pp2:02d}-10'
                moves.append({
                    '품번': code,
                    '품명': r['품명'],
                    '이전로케': old_main,
                    '신규로케': f'{zp}-{pp1:02d}-10',
                    '사유': '인접 빈 쌍 확보 (보조 거리 1)',
                    '종류': '메인 이동',
                })
                continue

        # 보조 위치 기록
        if chosen_sub is not None:
            zsec = z if code not in secondary_map else None  # 이미 phase3에서 처리된 경우 skip
            if code not in secondary_map:
                secondary_map[code] = f'{z}-{chosen_sub:02d}-10'
        else:
            secondary_map[code] = '—'

    # abc_df 업데이트: 메인이 이동했을 수 있음 + 보조로케 컬럼
    new_main = []
    for _, r in abc_df.iterrows():
        code = r['품번']
        if code in main_loc:
            z, p = main_loc[code]
            new_main.append(f'{z}-{p:02d}-10')
        else:
            new_main.append(r['현재로케'])
    abc_df['현재로케'] = new_main
    abc_df['보조로케'] = abc_df['품번'].map(secondary_map).fillna('')

    n_secured = sum(1 for v in secondary_map.values() if v not in ('', '—'))
    n_failed  = sum(1 for v in secondary_map.values() if v == '—')
    n_swap    = sum(1 for m in moves if m['종류'] == '1슬롯 이동')
    n_main    = sum(1 for m in moves if m['종류'] == '메인 이동')

    log(f'  보조 슬롯 확보: {n_secured}건 (실패 {n_failed}건)')
    log(f'  → 1슬롯 이동: {n_swap}건, 메인 이동: {n_main}건')

    return abc_df, moves


def _find_empty_slot(grid, prefer_zone):
    """빈 자리 1개 찾기. prefer_zone을 우선."""
    for p in range(1, 101):
        if (prefer_zone, p) not in grid:
            return (prefer_zone, p)
    other = sorted({z for z, _ in grid.keys() if z != prefer_zone})
    for z in other:
        for p in range(1, 101):
            if (z, p) not in grid:
                return (z, p)
    return None


def _find_empty_pair(grid, prefer_zone):
    """인접한 두 빈 자리 (pos, pos+1) 찾기. prefer_zone 우선."""
    for p in range(1, 100):
        if (prefer_zone, p) not in grid and (prefer_zone, p + 1) not in grid:
            return (prefer_zone, p, p + 1)
    other = sorted({z for z, _ in grid.keys() if z != prefer_zone})
    for z in other:
        for p in range(1, 100):
            if (z, p) not in grid and (z, p + 1) not in grid:
                return (z, p, p + 1)
    return None


def write_optimization_log(ws, moves):
    ws.cell(row=1, column=1,
            value='로케 최적화 이동 내역 (보조를 메인 옆에 두기 위한 자동 재배치)'
            ).font = Font(bold=True, size=14)
    ws.cell(row=2, column=1,
            value='우선순위: 출고량 큰 품목 먼저. Phase1 빈자리 → Phase2 1슬롯 swap → Phase3 메인 이동.'
            ).font = Font(italic=True, color='555555')

    hdr = ['품번', '품명', '이전로케', '신규로케', '사유', '종류']
    for i, h in enumerate(hdr, start=1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill('solid', fgColor='DDDDDD')

    for i, m in enumerate(moves, start=5):
        ws.cell(row=i, column=1, value=m['품번'])
        ws.cell(row=i, column=2, value=m['품명'])
        ws.cell(row=i, column=3, value=m['이전로케'])
        ws.cell(row=i, column=4, value=m['신규로케'])
        ws.cell(row=i, column=5, value=m['사유'])
        ws.cell(row=i, column=6, value=m['종류'])

    widths = [12, 40, 14, 14, 28, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A5'


def write_change_log(ws, abc_df, prev_master_df):
    ws.cell(row=1, column=1, value='변경후 노트 (직전 master 대비 로케 이동)').font = Font(bold=True, size=14)

    if prev_master_df is None or prev_master_df.empty:
        ws.cell(row=3, column=1, value='이전 master 스냅샷이 없어 변경 비교 생략됨.')
        return

    merged = abc_df[['품번', '품명', '현재로케']].merge(
        prev_master_df[['품번', '현재로케']].rename(columns={'현재로케': '이전로케'}),
        on='품번', how='left'
    )
    changed = merged[
        merged['이전로케'].notna() &
        (merged['현재로케'] != merged['이전로케'])
    ]

    hdr = ['품번', '품명', '이전로케', '신규로케']
    for i, h in enumerate(hdr, start=1):
        c = ws.cell(row=3, column=i, value=h); c.font = Font(bold=True)
        c.fill = PatternFill('solid', fgColor='DDDDDD')

    for i, (_, r) in enumerate(changed.iterrows(), start=4):
        ws.cell(row=i, column=1, value=r['품번'])
        ws.cell(row=i, column=2, value=r['품명'])
        ws.cell(row=i, column=3, value=r['이전로케'])
        ws.cell(row=i, column=4, value=r['현재로케'])

    widths = [12, 40, 14, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# -------------------------------------------------------------------
# 6. 메인
# -------------------------------------------------------------------
def run_analysis(monthly_dir, master_file, output_dir, log=print):
    """
    분석 본체. GUI/CLI 공용 진입점.
    성공 시 생성된 output 파일 경로 반환.
    """
    log('== ABC 재배치 분석 시작 ==')

    log('\n[1] 월별 데이터 로드')
    long_df = load_monthly_files(monthly_dir, log=log)

    ref_row = long_df.sort_values(['년', '월']).iloc[-1]
    ref_y, ref_m = int(ref_row['년']), int(ref_row['월'])
    log(f'  기준월: {ref_y}-{ref_m:02d}')

    unique_months = long_df.drop_duplicates(['년', '월'])[['년', '월']]
    unique_months = unique_months[
        ((ref_y - unique_months['년']) * 12 + (ref_m - unique_months['월']))
        .between(0, WINDOW_MONTHS - 1)
    ]
    n_months_avail = len(unique_months)
    log(f'  윈도우 내 월 수: {n_months_avail}/{WINDOW_MONTHS}')

    log('\n[2] 평균 계산 (연 / 하절기 / 동절기)')
    avg_df = compute_averages(long_df, ref_y, ref_m)
    log(f'  품목 수: {len(avg_df)}')

    log('\n[3] 로케 마스터 로드 & ABC 분류')
    if not os.path.exists(master_file):
        raise AnalysisError(f'로케 마스터 파일이 없습니다.\n경로: {master_file}')
    master_df = pd.read_excel(master_file, sheet_name='location_master')
    master_df['품번'] = _normalize_code(master_df['품번'])
    abc_df = classify_abc(avg_df, master_df, ref_y, ref_m)
    log(f"  A={len(abc_df[abc_df['ABC']=='A'])} "
        f"B={len(abc_df[abc_df['ABC']=='B'])} "
        f"C={len(abc_df[abc_df['ABC']=='C'])}")

    prev_master_df = _load_prev_master_snapshot(output_dir, ref_y, ref_m)

    log('\n[4] 보조 슬롯 최적화 (옆자리 자동 배치)')
    abc_df, opt_moves = optimize_secondary_slots(abc_df, log=log)

    log('\n[5] 출력 파일 생성')
    os.makedirs(output_dir, exist_ok=True)
    out_path = os.path.join(output_dir, f'ABC_및_재배치_{ref_y}-{ref_m:02d}.xlsx')

    wb = Workbook()
    ws = wb.active; ws.title = '요약'
    write_summary(ws, abc_df, ref_y, ref_m, n_months_avail)

    ws = wb.create_sheet('ABC_리스트')
    _write_df(ws, abc_df)

    ws = wb.create_sheet('빅무브_이동계획')
    write_big_move(ws, abc_df)

    ws = wb.create_sheet('하절기_재배치_추천')
    write_seasonal_reloc(ws, abc_df, '하절기팔레트일', '하절기팔레트일',
                         '하절기 재배치 추천 (6~9월)')
    ws = wb.create_sheet('동절기_재배치_추천')
    write_seasonal_reloc(ws, abc_df, '동절기팔레트일', '동절기팔레트일',
                         '동절기 재배치 추천 (11~2월)')

    ws = wb.create_sheet('변경후_노트')
    write_change_log(ws, abc_df, prev_master_df)

    ws = wb.create_sheet('로케_최적화_이동내역')
    write_optimization_log(ws, opt_moves)

    write_heatmap(wb.create_sheet('히트맵_연평균'),   abc_df, '연팔레트일',     '연평균 (팔레트/일)')
    write_heatmap(wb.create_sheet('히트맵_하절기'),   abc_df, '하절기팔레트일', '하절기 6~9월 (팔레트/일)')
    write_heatmap(wb.create_sheet('히트맵_동절기'),   abc_df, '동절기팔레트일', '동절기 11~2월 (팔레트/일)')

    # 다음 달 비교용 master 스냅샷 (숨김)
    ws = wb.create_sheet('_master_snapshot')
    _write_df(ws, master_df)
    ws.sheet_state = 'hidden'

    try:
        wb.save(out_path)
    except PermissionError:
        raise AnalysisError(
            f'파일 저장 실패 (권한 없음).\n'
            f'출력 파일이 엑셀에서 열려 있을 수 있습니다. 엑셀을 닫고 다시 시도하세요.\n'
            f'경로: {out_path}'
        )

    log(f'  저장됨 → {out_path}')
    log('\n== 완료 ==')
    return out_path


def _write_df(ws, df):
    """DataFrame을 시트에 기본 테이블로 쓰기."""
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=1, column=j, value=str(col))
        c.font = Font(bold=True)
        c.fill = PatternFill('solid', fgColor='DDDDDD')
    for i, row in enumerate(df.itertuples(index=False), start=2):
        for j, v in enumerate(row, start=1):
            if isinstance(v, (np.floating, np.integer)):
                v = v.item()
            if pd.isna(v):
                v = None
            ws.cell(row=i, column=j, value=v)
    for j, col in enumerate(df.columns, start=1):
        avg_len = df[col].astype(str).str.len().mean()
        if pd.isna(avg_len):
            avg_len = 10
        w = max(10, min(40, int(avg_len * 1.4) + 4))
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = 'A2'


def _load_prev_master_snapshot(output_dir, ref_y, ref_m):
    """직전 달 output 파일의 _master_snapshot 시트 로드 (없으면 None)."""
    prev_m = ref_m - 1 or 12
    prev_y = ref_y if ref_m > 1 else ref_y - 1
    prev_file = os.path.join(output_dir, f'ABC_및_재배치_{prev_y}-{prev_m:02d}.xlsx')
    if not os.path.exists(prev_file):
        return None
    try:
        return pd.read_excel(prev_file, sheet_name='_master_snapshot')
    except Exception:
        return None


# -------------------------------------------------------------------
# 7. CLI 진입점 (개발/디버깅용)
# -------------------------------------------------------------------
if __name__ == '__main__':
    base = os.path.dirname(os.path.abspath(__file__))
    run_analysis(
        monthly_dir=os.path.join(base, 'data', 'monthly_data'),
        master_file=os.path.join(base, 'data', 'location_master.xlsx'),
        output_dir=os.path.join(base, 'output'),
    )
