# -*- coding: utf-8 -*-
"""공용 기준정보 허브 — 한 번 업로드하면 여러 페이지가 공용으로 사용.

두 종류의 공용 데이터를 중앙 저장(Supabase app_settings + 번들 시드):
  1) Item 마스터 (ERP Item_*.xlsx)   → 하대(배면×배단)·품명·입수·유통기한(월)
  2) 고정로케이션 매핑 (편집본_*.xlsx) → Item code → 고정로케이션(홈 위치)

저장: gzip+base64 JSON 블롭을 Supabase에 upsert. Supabase 미설정(로컬)이면
      번들 시드(data/*.json.gz)로 유지 → 로컬 테스트/최초 배포에서도 동작.

다른 페이지에서:
    from master_hub import store
    hadae = store.hadae_map()      # {코드(int): 하대}
    name  = store.name_map()       # {코드(int): 품명}
    loc   = store.loc_map()        # {코드(int): 고정로케이션}
"""
from __future__ import annotations
import base64
import gzip
import io
import json
from datetime import datetime, timezone, timedelta
from pathlib import Path

try:
    import streamlit as st
except Exception:
    st = None

import openpyxl

KST = timezone(timedelta(hours=9))
DATA = Path(__file__).parent / "data"
DATA.mkdir(parents=True, exist_ok=True)

ITEM_KEY, ITEM_META = "hub_item_b64", "hub_item_meta"
LOC_KEY, LOC_META = "hub_loc_b64", "hub_loc_meta"
FLAG_KEY = "hub_itemflag"                  # {코드: {nm,담당자,등록일,flag}} — 조회용(jsonb, 시드없음: 담당자=개인정보)
ITEM_RAW_KEY = "hub_item_raw_b64"          # 원본 ERP Item xlsx 바이트 (coupang/ov5가 그대로 소비)
ITEM_SEED = DATA / "item_seed.json.gz"
LOC_SEED = DATA / "loc_seed.json.gz"
ITEM_RAW_SEED = DATA / "item_raw.xlsx"     # 로컬 원본 캐시(용량 큼 → .gitignore)


# ---------------- Supabase 하부 ----------------
def _secret(n):
    try:
        return st.secrets.get(n) if st is not None else None
    except Exception:
        return None


def use_supabase() -> bool:
    return bool(_secret("SUPABASE_URL")) and bool(_secret("SUPABASE_KEY"))


def _sb():
    from supabase import create_client
    return create_client(st.secrets["SUPABASE_URL"], st.secrets["SUPABASE_KEY"])


def _encode(obj) -> str:
    raw = json.dumps(obj, ensure_ascii=False).encode("utf-8")
    return base64.b64encode(gzip.compress(raw, 6)).decode("ascii")


def _decode(b64: str):
    return json.loads(gzip.decompress(base64.b64decode(b64)).decode("utf-8"))


def _seed_write(path: Path, obj) -> None:
    try:
        with gzip.open(path, "wt", encoding="utf-8") as f:
            json.dump(obj, f, ensure_ascii=False)
    except Exception:
        pass


def _seed_read(path: Path):
    if path.exists():
        try:
            with gzip.open(path, "rt", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return None
    return None


# ---------------- 파서 (업로드 → dict) ----------------
def _to_code(v):
    if v is None:
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        s = str(v).strip()
        return int(s) if s.isdigit() else None


def parse_item_master(file_bytes: bytes) -> dict:
    """ERP Item_*.xlsx → {코드: {'nm','ip','bm','bd','hadae','shelf'}}.

    열: A(0) Item code, B(1) Item, D(3) 입수, G(6) 소비기한(월),
        AD(29) 배면, AE(30) 배단.  하대 = 배면 × 배단.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    out: dict[str, dict] = {}
    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue
        code = _to_code(row[0]) if row else None
        if code is None:
            continue
        bm = row[29] if len(row) > 29 else None
        bd = row[30] if len(row) > 30 else None
        try:
            hadae = int(round(float(bm) * float(bd))) if (bm and bd) else None
        except (TypeError, ValueError):
            hadae = None
        out[str(code)] = {
            "nm": row[1] if len(row) > 1 else "",
            "ip": _numi(row[3]) if len(row) > 3 else None,
            "bm": _numi(bm), "bd": _numi(bd),
            "hadae": hadae,
            "shelf": _numi(row[6]) if len(row) > 6 else None,
        }
    wb.close()
    return out


def parse_fixed_loc(file_bytes: bytes) -> dict:
    """편집본_*.xlsx (ERP 로케이션 export) → {코드: 고정로케이션}.

    열: A(0) Inventory code, C(2) 로케이션ID, D(3) 보관타입, G(6) Item code.
    보관타입에 '고정로케이션' 포함하는 행만. 품목당 최초 1건.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    out: dict[str, str] = {}
    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue
        if len(row) < 7:
            continue
        keep = row[3]
        if keep is None or "고정로케이션" not in str(keep):
            continue
        code = _to_code(row[6])
        loc = row[2]
        if code is None or loc is None:
            continue
        out.setdefault(str(code), str(loc).strip())
    wb.close()
    return out


def _numi(x):
    try:
        return None if x is None else int(round(float(x)))
    except (TypeError, ValueError):
        return None


# ---------------- 저장 ----------------
def _save(key, meta_key, seed_path, data: dict, extra_meta: dict | None = None) -> bool:
    meta = {"품목수": len(data),
            "갱신": datetime.now(KST).strftime("%Y-%m-%d %H:%M")}
    if extra_meta:
        meta.update(extra_meta)
    payload = {"data": data, "meta": meta}
    _seed_write(seed_path, payload)          # 로컬/시드 항상 갱신
    _clear_cache()                           # 업로드 즉시 접근자 캐시 무효화(로컬/클라우드 공통)
    if not use_supabase():
        return False
    try:
        _sb().table("app_settings").upsert({"key": key, "value": _encode(data)}).execute()
        _sb().table("app_settings").upsert({"key": meta_key, "value": meta}).execute()
        return True
    except Exception:
        return False


def save_item(file_bytes: bytes) -> tuple[int, bool]:
    data = parse_item_master(file_bytes)
    ok = _save(ITEM_KEY, ITEM_META, ITEM_SEED, data,
               {"하대보유": sum(1 for v in data.values() if v.get("hadae"))})
    # 원본 xlsx 바이트도 보관 (coupang/ov5가 파일 그대로 소비)
    try:
        ITEM_RAW_SEED.write_bytes(file_bytes)
    except Exception:
        pass
    if use_supabase():
        try:
            _sb().table("app_settings").upsert(
                {"key": ITEM_RAW_KEY, "value": base64.b64encode(file_bytes).decode()}).execute()
        except Exception:
            pass
    return len(data), ok


def item_raw() -> bytes | None:
    """공용 원본 ERP Item xlsx 바이트. Supabase 우선 → 로컬 시드."""
    if use_supabase():
        try:
            r = _sb().table("app_settings").select("value").eq("key", ITEM_RAW_KEY).execute()
            if r.data and r.data[0].get("value"):
                return base64.b64decode(r.data[0]["value"])
        except Exception:
            pass
    if ITEM_RAW_SEED.exists():
        try:
            return ITEM_RAW_SEED.read_bytes()
        except Exception:
            return None
    return None


# ---------------- 물품담당자 (개인정보 → Supabase만, 시드 없음) ----------------
# 실사지/점검·재고출고시점(jaego)이 공유하는 기존 키를 그대로 사용 → 소비자 무변경.
DAMDANGJA_KEY = "inventory_담당자"


def parse_damdangja(file_bytes: bytes) -> dict:
    """물품담당자 xlsx → {코드(str): 담당자}. 헤더에서 '코드'/'담당자' 열 탐색, 없으면 A/B."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    code_i, name_i = None, None
    for i, h in enumerate(header):
        if h is None:
            continue
        ht = str(h).strip()
        if code_i is None and ht == "코드":
            code_i = i
        elif name_i is None and ht == "담당자":
            name_i = i
    if code_i is None or name_i is None:
        code_i, name_i = 0, 1
    m: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or code_i >= len(row) or name_i >= len(row):
            continue
        cval, nval = row[code_i], row[name_i]
        if cval is None:
            continue
        code = str(int(cval)) if isinstance(cval, (int, float)) else str(cval).strip()
        name = str(nval).strip() if nval is not None else None
        if code and name:
            m[code] = name
    wb.close()
    return m


def save_damdangja(file_bytes: bytes) -> tuple[int, bool]:
    """담당자 매핑을 공유 키(inventory_담당자)에 저장. 개인정보라 시드 없음(Supabase 전용)."""
    data = parse_damdangja(file_bytes)
    if not use_supabase():
        return len(data), False
    try:
        _sb().table("app_settings").upsert(
            {"key": DAMDANGJA_KEY, "value": data}, on_conflict="key").execute()
        return len(data), True
    except Exception:
        return len(data), False


def load_damdangja() -> dict:
    """공유 키(inventory_담당자)에서 {코드: 담당자} 로드. Supabase 전용."""
    if not use_supabase():
        return {}
    try:
        r = _sb().table("app_settings").select("value").eq("key", DAMDANGJA_KEY).execute()
        return (r.data[0].get("value") or {}) if r.data else {}
    except Exception:
        return {}


# ---------------- 모니터링 재고표(.xlsb) 한 번에 업데이트 ----------------
# 'Item기준정보' 시트 한 장에 코드·품명·입수·하대·소비기한월·담당자·품목등록일·Flag가 모두 있음.
# 이 파일 하나로 Item마스터 + 담당자 + 등록일/Flag 를 동시에 갱신한다.
def _mon_serial_to_date(x):
    try:
        from datetime import datetime, timedelta
        return (datetime(1899, 12, 30) + timedelta(days=float(x))).strftime("%Y-%m-%d")
    except (TypeError, ValueError):
        return None


def parse_monitoring(file_bytes: bytes) -> tuple[dict, dict, dict]:
    """모니터링 재고표(.xlsb) 'Item기준정보' 시트 → (item, damdangja, flags).

    item     = {코드: {nm, ip, hadae, shelf, bm:None, bd:None}}
    damdangja= {코드: 담당자}
    flags    = {코드: {nm, 담당자, 등록일(YYYY-MM-DD), flag}}
    """
    import pandas as pd
    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name="Item기준정보",
                       engine="pyxlsb", header=0)
    df.columns = [str(c).replace("\n", " ").strip() for c in df.columns]

    def col(*subs, exact=None):
        for c in df.columns:
            if exact is not None and c == exact:
                return c
        for c in df.columns:
            if all(s in c for s in subs):
                return c
        return None

    c_code = col(exact="코드") or col("코드")
    c_nm = col(exact="품목명") or col("품목명")
    c_ip = col(exact="입수") or col("입수")
    c_hadae = col("하대")
    c_shelf = col("소비기한")
    c_dam = col(exact="담당자") or col("담당자")
    c_reg = col("품목등록일")
    c_flag = col("Flag")

    item, dam, flags = {}, {}, {}
    for _, r in df.iterrows():
        cv = r.get(c_code)
        if cv is None or (isinstance(cv, float) and cv != cv):
            continue
        code = str(int(cv)) if isinstance(cv, (int, float)) else str(cv).strip()
        if not code or code in ("nan", "None"):
            continue
        item[code] = {
            "nm": (str(r.get(c_nm)).strip() if c_nm and r.get(c_nm) is not None else ""),
            "ip": _numi(r.get(c_ip)) if c_ip else None,
            "hadae": _numi(r.get(c_hadae)) if c_hadae else None,
            "shelf": _numi(r.get(c_shelf)) if c_shelf else None,
            "bm": None, "bd": None,
        }
        damv = r.get(c_dam) if c_dam else None
        if damv is not None and str(damv).strip():
            dam[code] = str(damv).strip()
        flags[code] = {
            "nm": item[code]["nm"],
            "담당자": dam.get(code, ""),
            "등록일": _mon_serial_to_date(r.get(c_reg)) if c_reg else None,
            "flag": (str(r.get(c_flag)).strip() if c_flag and r.get(c_flag) is not None else ""),
        }
    return item, dam, flags


def _synthesize_item_raw(item: dict) -> bytes:
    """item dict → 전 소비자 로더 호환 xlsx 바이트 (ERP 31열 레이아웃 흉내).

    - 이름 기반(coupang·OV5): 헤더 '배면'/'배단'/'Item code'/'소비기한(월)' 로 찾음.
    - 위치 기반(실사지 inv_core): code=idx0, name=idx1, 배면=idx29, 배단=idx30.
    배면=하대, 배단=1 → 배면×배단=하대. (모니터링 파일엔 하대가 직접 들어있음)
    """
    W = 31
    header = [""] * W
    header[0] = "Item code"; header[1] = "Item"; header[3] = "입수"
    header[6] = "소비기한(월)"; header[29] = "배면"; header[30] = "배단"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(header)
    for code, v in item.items():
        try:
            ic = int(code)
        except ValueError:
            continue
        row = [None] * W
        row[0] = ic; row[1] = v.get("nm", ""); row[3] = v.get("ip")
        row[6] = v.get("shelf"); row[29] = v.get("hadae")
        row[30] = 1 if v.get("hadae") else None
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def save_monitoring(file_bytes: bytes) -> dict:
    """모니터링 재고표 한 번에 반영: Item마스터 + 담당자 + 등록일/Flag.
    반환 통계 {item, 담당자, flag, item_ok, dam_ok, flag_ok}."""
    item, dam, flags = parse_monitoring(file_bytes)
    # 1) Item 마스터 — 병합(모니터링 값이 우선, 기존 ERP 품목은 유지해 손실 방지)
    existing, _ = load_item()
    merged = dict(existing)
    merged.update(item)
    item_ok = _save(ITEM_KEY, ITEM_META, ITEM_SEED, merged,
                    {"하대보유": sum(1 for v in merged.values() if v.get("hadae")),
                     "출처": f"모니터링 재고표(+{len(item)}갱신)"})
    # 2) coupang/OV5/실사지 호환 원본 합성 → item_raw (병합 전체 기준)
    raw = _synthesize_item_raw(merged)
    try:
        ITEM_RAW_SEED.write_bytes(raw)
    except Exception:
        pass
    if use_supabase():
        try:
            _sb().table("app_settings").upsert(
                {"key": ITEM_RAW_KEY, "value": base64.b64encode(raw).decode()}).execute()
        except Exception:
            pass
    # 3) 담당자 (공유 키) — 병합(기존 유지 + 모니터링 갱신)
    dam_ok = False
    if use_supabase() and dam:
        try:
            merged_dam = dict(load_damdangja())
            merged_dam.update(dam)
            _sb().table("app_settings").upsert(
                {"key": DAMDANGJA_KEY, "value": merged_dam}, on_conflict="key").execute()
            dam_ok = True
        except Exception:
            pass
    # 4) 등록일/Flag (조회용)
    flag_ok = False
    if use_supabase():
        try:
            _sb().table("app_settings").upsert(
                {"key": FLAG_KEY, "value": flags}, on_conflict="key").execute()
            flag_ok = True
        except Exception:
            pass
    _clear_cache()
    if st is not None:
        st.session_state.pop("_hub_item_raw_session", None)   # 합성본 재로딩
    return {"item": len(item), "item_total": len(merged),
            "담당자": len(dam), "flag": len(flags),
            "item_ok": item_ok, "dam_ok": dam_ok, "flag_ok": flag_ok}


def load_itemflags() -> dict:
    """{코드: {nm,담당자,등록일,flag}} — 담당자·품목등록일 조회용. Supabase 전용."""
    if not use_supabase():
        return {}
    try:
        r = _sb().table("app_settings").select("value").eq("key", FLAG_KEY).execute()
        return (r.data[0].get("value") or {}) if r.data else {}
    except Exception:
        return {}


def item_raw_session() -> bytes | None:
    """세션당 1회만 원본 Item 바이트를 가져와 재사용(매 rerun Supabase 재요청 방지)."""
    if st is None:
        return item_raw()
    key = "_hub_item_raw_session"
    if key not in st.session_state:
        st.session_state[key] = item_raw()
    return st.session_state[key]


def restore_item_to(dest_path) -> bool:
    """공용 원본 Item xlsx를 dest_path(파일)에 써넣음. 성공 여부 반환.
    coupang/ov5의 기존 파일 기반 로직을 그대로 두고 소스만 허브로 바꾸는 용도."""
    b = item_raw()
    if not b:
        return False
    try:
        Path(dest_path).parent.mkdir(parents=True, exist_ok=True)
        Path(dest_path).write_bytes(b)
        return True
    except Exception:
        return False


def save_loc(file_bytes: bytes) -> tuple[int, bool]:
    data = parse_fixed_loc(file_bytes)
    ok = _save(LOC_KEY, LOC_META, LOC_SEED, data)
    return len(data), ok


# ---------------- 로드 (Supabase 우선 → 시드) ----------------
def _load(key, meta_key, seed_path) -> tuple[dict, dict]:
    if use_supabase():
        try:
            r = _sb().table("app_settings").select("value").eq("key", key).execute()
            m = _sb().table("app_settings").select("value").eq("key", meta_key).execute()
            if r.data and r.data[0].get("value"):
                data = _decode(r.data[0]["value"])
                meta = m.data[0].get("value") if m.data else {}
                return data, meta
        except Exception:
            pass
    p = _seed_read(seed_path)
    if p:
        return p.get("data", {}), p.get("meta", {})
    return {}, {}


def load_item() -> tuple[dict, dict]:
    return _load(ITEM_KEY, ITEM_META, ITEM_SEED)


def load_loc() -> tuple[dict, dict]:
    return _load(LOC_KEY, LOC_META, LOC_SEED)


# ---------------- 편의 접근자 (코드 int 키) ----------------
def _cache(fn):
    if st is not None:
        return st.cache_data(show_spinner=False)(fn)
    return fn


def _clear_cache():
    for f in (_item_data, _loc_data, hadae_map, name_map, ipsu_map, shelf_map, loc_map):
        try:
            f.clear()
        except Exception:
            pass


@_cache
def _item_data():
    return load_item()[0]


@_cache
def _loc_data():
    return load_loc()[0]


@_cache
def hadae_map() -> dict:
    return {int(k): v["hadae"] for k, v in _item_data().items() if v.get("hadae")}


@_cache
def name_map() -> dict:
    return {int(k): (v.get("nm") or "") for k, v in _item_data().items()}


@_cache
def ipsu_map() -> dict:
    return {int(k): v.get("ip") for k, v in _item_data().items() if v.get("ip")}


@_cache
def shelf_map() -> dict:
    return {int(k): v.get("shelf") for k, v in _item_data().items() if v.get("shelf") is not None}


@_cache
def loc_map() -> dict:
    return {int(k): v for k, v in _loc_data().items() if v}
