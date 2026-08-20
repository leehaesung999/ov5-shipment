# -*- coding: utf-8 -*-
"""재고이동계획 (Streamlit 페이지)

매일 하는 일: **재고입력(현재고) 업로드** → 이동_박스·파레트 자동 산출.
기준(안전재고·Min·Max)은 '안전재고 계산기'에서 저장한 값을 사용.
이동 = MIN(요청, 출고가능, 할당) 박스단위. 입고예정 차감·종료 제외·행사 프리쉽 반영.
"""
import io
import sys
from datetime import datetime, timezone, timedelta, date
from pathlib import Path

import openpyxl
import pandas as pd
import streamlit as st

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from transfer import core as T          # noqa: E402
from safety_stock import store          # noqa: E402

KST = timezone(timedelta(hours=9))

st.title("🚚 BNF(비네이버) 재고이동 계획")
baseline, bmeta = store.load_baseline()
st.caption("기준(안전재고·Min·Max)은 '안전재고 계산기'에서 저장한 값 사용. "
           f"현재 기준: {bmeta.get('품목수','?')}품목 · 갱신 {bmeta.get('갱신','—')}")

if not baseline:
    st.warning("안전재고 기준이 없습니다. 먼저 '안전재고 계산기'에서 '재고이동계획에 적용'을 눌러 기준을 저장하세요.")
    st.stop()

lead = int((bmeta.get("설정") or {}).get("lead_time", 3))
plan_date = st.date_input("계획일자 (행사 프리쉽 기준)", value=datetime.now(KST).date())


# ---------- 업로드 파서 ----------
def _rows(file, sheet=None, min_row=2):
    # read_only=False: 일부 시스템 export가 read_only에서 행이 잘려 읽히는 문제 회피(파일 작음)
    wb = openpyxl.load_workbook(io.BytesIO(file.getvalue()), data_only=True)
    ws = wb[sheet] if (sheet and sheet in wb.sheetnames) else wb[wb.sheetnames[0]]
    for r in ws.iter_rows(min_row=min_row, values_only=True):
        yield r
    wb.close()


def _code(v):
    if v is None:
        return None
    if isinstance(v, float):
        return str(int(v))
    return str(v).strip()


def _num(x):
    """숫자로(넘파이/NaN 안전). 아니면 None."""
    try:
        if x is None:
            return None
        f = float(x)
        return None if f != f else f          # NaN 제외
    except (TypeError, ValueError):
        return None


def parse_stock(file):
    """재고입력 — 헤더명 자동인식 + .xls/.xlsx 겸용.
    코드열: 상품코드/품목코드/CJ코드/제품코드/코드
    현재고열: 출고가능량(실가용=재고−등록−대기) 우선 → 없으면 재고수량/현재고/재고EA
    할당열(선택): 할당/할당수량
    → BNF '상품별재고현황' 양식(상품코드·출고가능량) 및 기존 [코드,현재고,할당] 양식 모두 지원."""
    df = pd.read_excel(io.BytesIO(file.getvalue()), header=None, dtype=object)
    hrow = 0
    for i in range(min(6, len(df))):
        vals = [str(x).strip() for x in df.iloc[i].tolist()]
        if any(v in ("상품코드", "품목코드", "CJ코드", "제품코드") for v in vals):
            hrow = i
            break
    hdr = [str(x).strip() for x in df.iloc[hrow].tolist()]

    def find(names, default=None):
        for j, h in enumerate(hdr):
            if h in names:
                return j
        return default
    ci = find(("상품코드", "품목코드", "CJ코드", "제품코드", "코드"), 0)
    # 실가용재고 우선(출고가능량 = 재고 − 출고등록 − 출고대기), 없으면 총재고
    cur_i = find(("출고가능량", "출고가능", "실가용재고", "가용재고"))
    if cur_i is None:
        cur_i = find(("재고수량", "현재고", "현재고(EA)", "재고EA"), 1)
    al_i = find(("할당", "할당수량", "할당수량(EA)"))

    out = {}
    for r in range(hrow + 1, len(df)):
        row = df.iloc[r].tolist()
        c = _code(row[ci]) if ci < len(row) else None
        if not c or not (c.isdigit() or c.startswith("P")):
            continue
        cur = _num(row[cur_i]) if cur_i < len(row) else None
        alloc = _num(row[al_i]) if (al_i is not None and al_i < len(row)) else None
        out[c] = {"cur": cur, "alloc": alloc}
    return out


def parse_avail(file):
    """출고가능재고 WMS export: C(3)=item id, AB(28)=출고가능(Box). 헤더 1행."""
    out = {}
    for r in _rows(file):
        if len(r) < 28:
            continue
        c = _code(r[2])
        box = r[27]
        if c and isinstance(box, (int, float)):
            out[c] = out.get(c, 0) + int(box)
    return out


def parse_incoming(file):
    """입고예정: A=코드, E(5)=낱개. 헤더 2행(그들 양식)."""
    out = {}
    for r in _rows(file, min_row=3):
        c = _code(r[0])
        ea = r[4] if len(r) > 4 and isinstance(r[4], (int, float)) else 0
        if c and ea:
            out[c] = out.get(c, 0) + ea
    return out


def parse_ended(file):
    """종료품목: A=CJ코드. 헤더 1행."""
    s = set()
    for r in _rows(file):
        c = _code(r[0])
        if c and (c.isdigit() or c.startswith("P")):
            s.add(c)
    return s


BNF_CHANNELS_DEFAULT = ["하프클럽", "오늘의집", "버킷플레이스", "카카오", "메이커스",
                        "배민대용량", "알리", "제로샵", "영동군청", "스타일셀러", "떠리몰"]


def _norm_ch(s):
    return str(s or "").replace("[", "").replace("]", "").replace(" ", "").replace("_", "").strip()


def parse_events_new(file, plan_d, reflected, bnf_channels=None, presupply_days=7):
    """행사 파일 → 신규(미반영) BNF 이벤트만 {코드: 수량}.
    - 0818 형식(헤더에 CHANNEL·ITEM_CODE): BNF 채널만 · END_DATE≥계획일자 · SEQ로 중복방지.
      수량 = SALE_TARGET_QUANTITY, 채널 부분일치(대괄호/공백/언더바 무시).
      presupply_days: 시작(START_DATE) N일 전부터만 사전공급 반영(너무 이른 건 제외).
    - 구 행사양식(header_id·customer_name): 하위호환(현행·미지난·header_id 미반영).
    반환: (events{코드:수량}, new_ids{id}, 통계dict)"""
    wb = openpyxl.load_workbook(io.BytesIO(file.getvalue()), data_only=True)
    ws = wb[wb.sheetnames[0]]
    hdr = [str(ws.cell(1, j).value).strip() if ws.cell(1, j).value is not None else ""
           for j in range(1, ws.max_column + 1)]
    kws = [_norm_ch(k) for k in (bnf_channels or BNF_CHANNELS_DEFAULT) if _norm_ch(k)]

    def is_bnf(ch):
        c = _norm_ch(ch)
        return bool(c) and any(k in c for k in kws)

    events, new_ids = {}, set()

    # ---- 0818 형식 (CHANNEL 기반) ----
    if "CHANNEL" in hdr and "ITEM_CODE" in hdr:
        iC = hdr.index("CHANNEL"); iItem = hdr.index("ITEM_CODE")
        iQty = hdr.index("SALE_TARGET_QUANTITY") if "SALE_TARGET_QUANTITY" in hdr else 13
        iEnd = hdr.index("END_DATE") if "END_DATE" in hdr else 7
        iStart = hdr.index("START_DATE") if "START_DATE" in hdr else 6
        iSeq = hdr.index("SEQ") if "SEQ" in hdr else 0
        n_new = n_ref = n_past = n_ch = n_early = 0
        for r in ws.iter_rows(min_row=2, values_only=True):
            code = _code(r[iItem]) if iItem < len(r) else None
            if not code:
                continue
            if not is_bnf(r[iC] if iC < len(r) else None):
                n_ch += 1
                continue
            end = r[iEnd] if iEnd < len(r) else None
            endd = end.date() if hasattr(end, "date") else None
            if isinstance(endd, date) and endd < plan_d:    # 이미 끝난 행사
                n_past += 1
                continue
            start = r[iStart] if iStart < len(r) else None
            startd = start.date() if hasattr(start, "date") else None
            if startd and plan_d < startd - timedelta(days=presupply_days):  # 아직 이른 행사
                n_early += 1
                continue
            sid = str(r[iSeq]) if (iSeq < len(r) and r[iSeq] is not None) else None
            if sid and sid in reflected:                     # 이미 반영
                n_ref += 1
                continue
            q = r[iQty] if iQty < len(r) else None
            if not isinstance(q, (int, float)) or q <= 0:
                continue
            events[code] = events.get(code, 0) + q
            if sid:
                new_ids.add(sid)
            n_new += 1
        wb.close()
        return events, new_ids, {"신규": n_new, "이미반영": n_ref,
                                 "지난행사": n_past, "비BNF채널": n_ch,
                                 "이른행사": n_early, "형식": "0818"}

    # ---- 구 행사양식 형식 (header_id·customer_name) 하위호환 ----
    n_new = n_reflected = n_past = n_flag = 0
    for r in ws.iter_rows(min_row=2, values_only=True):
        if len(r) < 23:
            continue
        hid = r[7]
        if hid is None:
            continue
        hid = str(hid).strip()
        if str(r[9]).strip() != "Y":
            n_flag += 1
            continue
        ld = r[17]
        ldd = (ld.date() if hasattr(ld, "date") else ld) if ld is not None else None
        if isinstance(ldd, date) and ldd < plan_d:
            n_past += 1
            continue
        if hid in reflected:
            n_reflected += 1
            continue
        c = _code(r[18])
        q = r[22]
        if c and isinstance(q, (int, float)) and q:
            events[c] = events.get(c, 0) + q
            new_ids.add(hid)
            n_new += 1
    wb.close()
    return events, new_ids, {"신규": n_new, "이미반영": n_reflected,
                             "지난행사": n_past, "비현행": n_flag, "형식": "행사양식"}


def parse_location(file, sources):
    """로케이션별 재고조회 → {코드: {창고: {avail:출고가능박스, exp:최단소비기한}}}.
    열(0-based): Inventory=0, 제품코드=4, 소비기한=9(YYYYMMDD), 출고가능(Box)=18."""
    out = {}
    for r in _rows(file):
        if len(r) < 19:
            continue
        wh = str(r[0]).strip() if r[0] is not None else ""
        if wh not in sources:
            continue
        c = _code(r[4])
        box = r[18]
        if not c or not isinstance(box, (int, float)) or box <= 0:
            continue
        exp = r[9]
        try:
            exp = int(str(int(exp)) if isinstance(exp, (int, float)) else str(exp).replace("-", "")[:8])
        except (TypeError, ValueError):
            exp = 99999999
        d = out.setdefault(c, {}).setdefault(wh, {"avail": 0.0, "exp": exp})
        d["avail"] += box
        d["exp"] = min(d["exp"], exp)          # 창고 대표 = 최단 유통기한
    return out


def parse_allocation(file, plan_d):
    """카테고리별 할당 모니터링 시트 → BNF 할당 상한. 시트명 코드 기준.
    각 품목시트에서 'Bnf' 채널 행의 할당량(EA)을 합산. 할당 종료일≥계획일자만 활성.
    반환: (alloc{품목코드: BNF할당량EA}, rows[표시용 리스트]).
    """
    import re
    wb = openpyxl.load_workbook(io.BytesIO(file.getvalue()), data_only=True)
    alloc, rows = {}, []
    for sn in wb.sheetnames:
        m = re.match(r"^(\d+)\s", sn)
        if not m:                               # 품목시트(코드로 시작)만
            continue
        code = m.group(1)
        ws = wb[sn]
        # 할당 종료일
        endd = None
        for rr in range(1, 12):
            for cc in range(1, 6):
                v = ws.cell(rr, cc).value
                if isinstance(v, str) and "종료일" in v:
                    ev = ws.cell(rr, cc + 1).value
                    endd = ev.date() if hasattr(ev, "date") else None
        # 채널·할당량 헤더행
        hr = None
        for rr in range(1, 16):
            row = [str(ws.cell(rr, c).value or "") for c in range(1, 11)]
            if "채널" in row and "할당량" in row:
                hr = rr
                break
        if hr is None:
            continue
        hdr = [str(ws.cell(hr, c).value or "").strip() for c in range(1, 11)]
        chc = hdr.index("채널") + 1
        qc = hdr.index("할당량") + 1
        bnf_ea, has = 0, False
        for rr in range(hr + 1, ws.max_row + 1):
            ch = ws.cell(rr, chc).value
            if ch and "bnf" in str(ch).lower():
                q = ws.cell(rr, qc).value
                if isinstance(q, (int, float)):
                    bnf_ea += q
                    has = True
        if not has:
            continue
        active = (endd is None) or (endd >= plan_d)
        end_str = endd.isoformat() if endd else ""
        rows.append({"품목코드": code, "품목명": sn[len(code):].strip(),
                     "BNF할당(EA)": int(bnf_ea), "종료일": str(endd) if endd else "-",
                     "활성": "○" if active else "종료"})
        if active and bnf_ea > 0:
            alloc[code] = {"qty": int(bnf_ea), "end": end_str}
    wb.close()
    return alloc, rows


def _pallet_xlsx(rows, plan_d, only_wh=None):
    """이동_박스>0 품목을 'BNF 파레트 구분기' 입력형식으로.
    only_wh 지정 시 그 배정창고 품목만(창고별 피킹). None이면 전체.
    헤더행에 'Item code' 포함(구분기 find_header_row 대응). 박스=이동_박스, Plt_1차=파레트환산."""
    wb = openpyxl.Workbook()
    ws = wb.active
    title = f"{plan_d} 재고이동 파레트 구분" + (f" [{only_wh}]" if only_wh else "")
    ws.cell(1, 1, title)
    headers = ["Item code", "Item", "입수", "박스", "낱개", "소비기한",
               "plt환산", "Plt_1차", "PLT 번호", "비 고", "순번"]
    for j, h in enumerate(headers, 1):
        ws.cell(2, j, h)
    r = 3
    for row in rows:
        mb = row["★이동_박스"]
        if not isinstance(mb, (int, float)) or mb <= 0:
            continue
        if only_wh is not None and row.get("배정창고") != only_wh:
            continue
        plt = row["하대박스수"]                 # 1파레트당 박스수(구분기 plt환산=박스/파레트 분할기준)
        plt1 = row["파레트환산"]                # 파레트 점유율(=이동박스/하대)
        ws.cell(r, 1, row["품목코드"])
        ws.cell(r, 2, row["품목명"] or row["품목코드"])   # 빈 품명→코드(파레트 구분기 None 방지)
        ws.cell(r, 3, row["입수"])
        ws.cell(r, 4, int(mb))
        ws.cell(r, 5, row["이동_EA"])
        ws.cell(r, 6, row.get("최단유통기한", "") or "")   # 소비기한(배정창고 최단)
        ws.cell(r, 7, plt if plt else 0)       # plt환산 = 하대박스수(파레트당 박스수)
        ws.cell(r, 8, plt1 if plt1 is not None else 0.0)   # Plt_1차 = 파레트 점유율(패킹 기준)
        ws.cell(r, 10, row.get("배정창고", "") or "")       # 비고 = 배정창고
        r += 1
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ---------- 입력 ----------
st.subheader("1️⃣ 입력 (재고 필수, 나머지 선택)")
up_stock = st.file_uploader("재고입력 (BNF 상품별재고현황 .xls 그대로 OK / 또는 코드·현재고·할당)",
                            type=["xlsx", "xls"], key="t_stock")
up_loc = st.file_uploader("🏬 로케이션별 재고조회 (우리 창고 IC930/920/100) — 창고배정 + 이동량 상한",
                          type=["xlsx"], key="t_loc")
cap_by_loc = st.checkbox("창고 재고로 이동량 제한 (있는 만큼만 보냄)", value=True,
                         help="로케이션(우리 창고) 재고 합계를 이동량 상한으로 사용. "
                              "창고에 없는 품목은 이동 0, 부족분은 '창고재고부족'으로 표시.")
up_alloc = st.file_uploader("🎫 할당 파일 (카테고리별 할당 모니터링) — 재고부족 품목 BNF 할당량 제한 · 매주 월요일 갱신",
                            type=["xlsx"], key="t_alloc")
st.caption("품목시트의 'Bnf' 채널 할당량(EA)을 상한으로 사용. 할당 종료일 지난 건 제외. "
           "예: 토장450g BNF 20EA면 그만큼만 이동.")
with st.expander("추가 입력 (출고가능·입고예정·종료)"):
    up_avail = st.file_uploader("출고가능재고 (WMS export)", type=["xlsx"], key="t_avail")
    up_inc = st.file_uploader("입고예정", type=["xlsx"], key="t_inc")
    up_end = st.file_uploader("종료품목 (A:CJ코드)", type=["xlsx"], key="t_end")

st.subheader("2️⃣ 행사 이벤트 (선택)")
up_evt = st.file_uploader("행사 파일 업로드 (0818 이벤트 형식 / 구 행사양식) — 신규(미반영)만 자동 반영",
                          type=["xlsx"], key="t_evt")
_bnf_ch = store.load_bnf_channels() or BNF_CHANNELS_DEFAULT
presupply_days = st.number_input("행사 사전공급 리드(일) — 시작일 이만큼 전부터만 반영",
                                 0, 60, 7, help="예: 7이면 행사 시작 7일 전부터 미리 보냅니다. "
                                 "시작이 한참 뒤인 행사를 너무 일찍 보내지 않도록 제한.")
st.caption(f"**BNF 거래처만** 필터링해 반영합니다 (네이버·토스 등 제외). "
           f"시작 {presupply_days}일 전부터·안 끝난 행사(END≥계획일자)·미반영건만 이동에 더하고, "
           f"'반영완료' 후 다음부터 제외됩니다.")
with st.expander(f"🎯 BNF 거래처 필터 ({len(_bnf_ch)}개) — 편집"):
    _txt = st.text_area("거래처 키워드 (쉼표 구분, 부분일치)", value=", ".join(_bnf_ch), height=80,
                        help="행사 파일의 CHANNEL/거래처명에 이 키워드가 포함되면 BNF로 반영합니다. "
                             "예: 카카오 → [카카오]·[카카오쇼핑]·[카카오메이커스] 모두 포함")
    if st.button("💾 거래처 목록 저장", key="save_bnf_ch"):
        newch = [x.strip() for x in _txt.replace("\n", ",").split(",") if x.strip()]
        ok = store.save_bnf_channels(newch)
        st.success(f"{len(newch)}개 저장" + (" (Supabase)" if ok else " (로컬은 시드파일)"))
        st.rerun()

# ---------- 산출 ----------
if st.button("🚚 이동계획 산출", type="primary", disabled=up_stock is None):
    try:
        stock = parse_stock(up_stock)
        incoming = parse_incoming(up_inc) if up_inc else {}
        ended = parse_ended(up_end) if up_end else set()
        # 할당 상한(기간 총량): 잔여상한 = 할당량 − 그 기간 누적이동. 기존 할당과 더 작은 값.
        alloc_map, alloc_rows = ({}, [])
        if up_alloc:
            alloc_map, alloc_rows = parse_allocation(up_alloc, plan_date)
            used = store.load_alloc_used()
            for code, info in alloc_map.items():
                prev = used.get(code)
                used_ea = prev["used"] if (prev and prev.get("end") == info["end"]) else 0
                eff = max(info["qty"] - used_ea, 0)          # 잔여 할당 상한
                info["used"] = used_ea
                info["eff"] = eff
                if code in stock:
                    ca = stock[code].get("alloc")
                    stock[code]["alloc"] = eff if ca is None else min(ca, eff)
        events, new_ids, estat = {}, set(), {}
        if up_evt:
            reflected = store.load_reflected_events()
            events, new_ids, estat = parse_events_new(
                up_evt, plan_date, reflected,
                store.load_bnf_channels() or BNF_CHANNELS_DEFAULT, int(presupply_days))

        # 이동량 상한(avail): 출고가능재고 파일 ∩ 로케이션(우리 창고) 합계. 둘 다면 더 작은 값.
        loc_inv = parse_location(up_loc, T.WH_SOURCES) if up_loc else None
        sources = []
        if up_avail:
            sources.append(parse_avail(up_avail))
        if loc_inv and cap_by_loc:
            sources.append({c: int(sum((w.get("avail") or 0) for w in whs.values()))
                            for c, whs in loc_inv.items()})
        avail = None
        if sources:
            avail = {}
            for c in set().union(*[set(s) for s in sources]):
                vals = [s[c] for s in sources if c in s]
                avail[c] = min(vals) if vals else 0
        cap_reason = "창고재고부족" if (loc_inv and cap_by_loc) else "출고가능제한"

        rows = T.compute_transfer(baseline, stock, avail, incoming, events, ended,
                                  plan_month=plan_date.month, cap_reason=cap_reason)
        if loc_inv:                              # 창고 배정(로케이션재고)
            T.allocate_warehouse(rows, loc_inv)
        df = pd.DataFrame(rows)
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as xw:
            df.to_excel(xw, index=False, sheet_name="이동계획")
        # 창고별 파레트 파일 (배정창고 있는 경우)
        wh_pallets = {}
        if up_loc:
            for wh in T.WH_SOURCES:
                if any(r.get("배정창고") == wh and isinstance(r.get("★이동_박스"), (int, float))
                       and r["★이동_박스"] > 0 for r in rows):
                    wh_pallets[wh] = _pallet_xlsx(rows, plan_date, only_wh=wh)
        # 미등록 신규품목: 재고파일엔 있는데 안전재고 기준(baseline)에 없는 코드
        unreg = [{"품목코드": c, "현재고": v.get("cur")}
                 for c, v in stock.items()
                 if c not in baseline and (v.get("cur") or 0) > 0]
        # 할당: 이번 계획의 실이동 EA 기록(확정 시 차감) + 표에 누적/잔여/이번이동 표기
        alloc_moves = {}
        if alloc_map:
            mv_by = {r["품목코드"]: r for r in rows}
            for code, info in alloc_map.items():
                r = mv_by.get(code)
                mea = int(r["이동_EA"]) if (r and isinstance(r.get("이동_EA"), (int, float))) else 0
                if mea > 0:
                    alloc_moves[code] = {"ea": mea, "end": info["end"]}
                for ar in alloc_rows:
                    if ar["품목코드"] == code:
                        ar["이미이동"] = info.get("used", 0)
                        ar["잔여상한"] = info.get("eff")
                        ar["이번이동"] = mea
        # 세션에 저장(반영완료 버튼이 결과를 유지하도록)
        st.session_state["t_result"] = {
            "df": df, "summ": T.summarize(rows),
            "xlsx": buf.getvalue(), "pallet": _pallet_xlsx(rows, plan_date),
            "wh_pallets": wh_pallets,
            "estat": estat, "new_ids": sorted(new_ids),
            "unreg": unreg,
            "alloc_rows": alloc_rows, "alloc_moves": alloc_moves,
            "hide_missing": len(stock) < len(baseline),
        }
    except Exception as e:
        st.error(f"산출 실패: {e}")

# ---------- 결과 표시 (세션 유지) ----------
res = st.session_state.get("t_result")
if res:
    summ = res["summ"]; df = res["df"]
    m = st.columns(5)
    m[0].metric("이동 품목", f"{summ['이동품목수']:,}")
    m[1].metric("총 이동_박스", f"{summ['총이동_박스']:,}")
    m[2].metric("총 파레트", f"{summ['총파레트']}")
    m[3].metric("가용부족(창고부족 포함)", f"{summ['가용부족']}")
    m[4].metric("결품/안전재고이하", f"{summ.get('결품',0)}/{summ.get('안전재고이하',0)}")

    # BNF 할당 제한(기간 총량) — 할당량/이미이동/잔여상한/이번이동 표시 + 확정(차감)
    alloc_rows = res.get("alloc_rows") or []
    if alloc_rows:
        active_n = sum(1 for a in alloc_rows if a["활성"] == "○")
        with st.expander(f"🎫 BNF 할당 제한 {len(alloc_rows)}품목 (활성 {active_n}) — 기간 총량(잔여) 상한",
                         expanded=True):
            st.caption("할당량은 **기간 총량**입니다. 잔여상한 = 할당량 − 이미이동. "
                       "이동 실행 후 **아래 '할당 이동 확정'을 눌러야** 차감되어, 다음부터 잔여만큼만 나갑니다. "
                       "종료일 지난 건 자동 해제. 매주 갱신 시 이 표로 매핑도 확인하세요.")
            st.dataframe(pd.DataFrame(alloc_rows), width="stretch", hide_index=True)
            am = res.get("alloc_moves") or {}
            if am:
                tot_ea = sum(v["ea"] for v in am.values())
                if st.button(f"✅ 할당 이동 확정 — {len(am)}품목 {tot_ea}EA 차감", key="alloc_confirm"):
                    used = store.load_alloc_used()
                    for code, mv in am.items():
                        prev = used.get(code)
                        base = prev["used"] if (prev and prev.get("end") == mv["end"]) else 0
                        used[code] = {"used": base + mv["ea"], "end": mv["end"]}
                    okk = store.save_alloc_used(used)
                    st.success(f"{len(am)}품목 차감 완료" + (" (Supabase)" if okk else " (로컬 미저장)")
                               + " — 다음 산출부터 잔여만큼만 이동합니다. (중복 방지: 한 번만 누르세요)")
            else:
                st.caption("이번 계획에서 할당 품목의 실이동이 없어 차감할 것이 없습니다.")

    # 창고재고부족: 요청보다 우리 창고 재고가 모자라 다 못 보낸 품목
    short_rows = [r for r in df.to_dict("records") if r.get("사유") == "창고재고부족"]
    if short_rows:
        with st.expander(f"🏬 창고재고부족 {len(short_rows)}품목 — 요청보다 창고 재고가 모자람(있는 만큼만 이동)",
                         expanded=True):
            st.caption("우리 창고(IC930/920/100)에 있는 만큼만 이동에 반영했습니다. "
                       "부족분(미충족)은 창고 입고 후 다음 계획에 잡힙니다.")
            sdf = pd.DataFrame([{"품목코드": r["품목코드"], "품목명": r["품목명"],
                                 "요청_박스": r["요청_박스"], "이동_박스": r["★이동_박스"],
                                 "미충족_박스": r["미충족_박스"], "창고재고": r.get("창고재고", "")}
                                for r in short_rows])
            st.dataframe(sdf, width="stretch", hide_index=True)

    # 미등록 신규품목 (재고엔 있는데 안전재고 기준 없음 → 계획 누락)
    unreg = res.get("unreg") or []
    if unreg:
        with st.expander(f"🆕 미등록 신규품목 {len(unreg)}개 — 안전재고 기준 없음(계획 누락)",
                         expanded=True):
            st.caption("재고파일엔 있으나 안전재고 기준(baseline)이 없어 이동계획에 안 잡힌 품목입니다. "
                       "'안전재고 산출' 페이지의 🆕신규품목 초기기준(유사품 기반)에 등록 후 재계산·적용하세요.")
            st.dataframe(pd.DataFrame(unreg), width="stretch", hide_index=True)

    # 소진 경고 (행사 초과 등으로 안전재고 밑까지 소진 / 결품)
    burn_rows = [r for r in df.to_dict("records") if r.get("소진경고")]
    if burn_rows:
        with st.expander(f"🟠 소진경고 {len(burn_rows)}품목 — 안전재고 이하/결품 (행사 초과 등 점검)",
                         expanded=True):
            st.caption("현재고가 안전재고 밑으로 내려간 품목입니다. 공급은 매일 보충으로 이어가되, "
                       "행사가 계획보다 많이 나가는지 확인하세요. (자동 추가공급은 하지 않음)")
            bdf = pd.DataFrame([{"품목코드": r["품목코드"], "품목명": r["품목명"],
                                 "현재고": r["현재고"], "Min": r["Min"], "이동_박스": r["★이동_박스"],
                                 "사유": r["사유"], "경고": r["소진경고"]} for r in burn_rows])
            st.dataframe(bdf, width="stretch", hide_index=True)

    # 창고 배정 경고 (분할필요/재고없음)
    warn_rows = [r for r in df.to_dict("records") if r.get("창고경고")]
    if warn_rows:
        with st.expander(f"⚠️ 창고 단독배정 불가 {len(warn_rows)}품목 — 분할/재고 검토 필요", expanded=True):
            wdf = pd.DataFrame([{"품목코드": r["품목코드"], "품목명": r["품목명"],
                                 "이동_박스": r["★이동_박스"], "배정창고": r["배정창고"],
                                 "창고재고": r["창고재고"], "경고": r["창고경고"]} for r in warn_rows])
            st.dataframe(wdf, width="stretch", hide_index=True)

    # 행사 반영 요약 + 반영완료
    est = res.get("estat") or {}
    if est:
        _exj = (f" · 비BNF채널 {est.get('비BNF채널',0)}건 · 이른행사 {est.get('이른행사',0)}건 제외"
                if est.get("형식") == "0818" else f" · 비현행 {est.get('비현행',0)}건")
        st.info(f"행사[{est.get('형식','?')}]: 신규 {est.get('신규',0)}건 반영 · "
                f"이미반영 {est.get('이미반영',0)}건 · 지난행사 {est.get('지난행사',0)}건" + _exj)
        if res.get("new_ids"):
            if st.button(f"✅ 신규 행사 {len(res['new_ids'])}건 반영완료 처리", type="secondary"):
                ok = store.add_reflected_events(res["new_ids"])
                st.success(f"{len(res['new_ids'])}건 반영완료 기록"
                           + (" (Supabase)" if ok else " (로컬)") + " — 다음 업로드부터 제외됩니다")

    show = df[df["사유"] != "미입력"] if res["hide_missing"] else df
    st.dataframe(show, width="stretch", height=460)

    MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    d1, d2 = st.columns(2)
    d1.download_button("📥 이동계획 다운로드 (xlsx)", res["xlsx"],
                       file_name=f"재고이동계획_{plan_date:%y%m%d}.xlsx", mime=MIME, width="stretch")
    d2.download_button("🧱 파레트 구분기용 (전체)", res["pallet"],
                       file_name=f"파레트입력_{plan_date:%y%m%d}.xlsx", mime=MIME, width="stretch")

    # 창고별 파레트 다운로드 (피킹은 창고별로)
    wp = res.get("wh_pallets") or {}
    if wp:
        st.markdown("**🏬 창고별 파레트 구분기용** (창고별 피킹)")
        cols = st.columns(len(wp))
        for i, (wh, data) in enumerate(wp.items()):
            cols[i].download_button(f"🧱 {wh}", data,
                                    file_name=f"파레트입력_{wh}_{plan_date:%y%m%d}.xlsx",
                                    mime=MIME, width="stretch", key=f"wp_{wh}")
    _seasonal = any(r.get("months") for r in baseline.values())
    st.caption(f"📅 **계획일자 {plan_date:%m}월의 계절 Min/Max** 기준으로 산출"
               + ("" if _seasonal else " (구 기준: 연 고정값)") + ". "
               "겨울 성수기엔 높게·비수기엔 낮게 자동 반영됩니다.")
    st.caption("이동=MIN(요청,출고가능,할당) 박스. 행사는 header_id로 신규만 반영(중복 방지). "
               "창고배정=단독가능 창고 중 유통기한 빠른 것(동점 IC930). 분할필요/재고없음은 ⚠️경고. "
               "🧱 창고별 파일을 각 창고 'BNF 파레트 구분기'에 올리면 창고별 피킹 파레트가 나옵니다.")
