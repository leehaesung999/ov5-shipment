"""
통합센터 재고 분석 코어 모듈
- analyze_and_save: 이중적치(보충오류) 분석
- edit_재고지_1단: 1단 재고지(실사지) 편집
"""

import sys
from datetime import datetime
from collections import defaultdict
from pathlib import Path

from copy import copy

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break

# PyInstaller로 exe 빌드된 경우에는 exe 파일 위치를 기준으로 data/ 참조.
# 일반 Python 실행 시에는 이 파일의 폴더 기준.
if getattr(sys, "frozen", False):
    _BASE = Path(sys.executable).parent
else:
    _BASE = Path(__file__).parent

TEMPLATE_1단 = _BASE / "data" / "재고지_1단_템플릿.xlsx"
TEMPLATE_1단_전체 = _BASE / "data" / "재고지_1단_전체_템플릿.xlsx"
TEMPLATE_단별 = _BASE / "data" / "재고지_단별_템플릿.xlsx"
LOC_LIST_2_6단 = _BASE / "data" / "재고지_2_6단_로케이션.xlsx"
TEMPLATE_LOCK_유통기한 = _BASE / "data" / "LOCK_유통기한_템플릿.xlsx"
담당자_PATH = _BASE / "data" / "물품담당자.xlsx"


def parse_ymd(v):
    """20260326(int/str) -> datetime"""
    if v is None or v == "":
        return None
    s = str(int(v)) if isinstance(v, float) else str(v)
    s = s.strip()
    if len(s) != 8:
        return None
    try:
        return datetime.strptime(s, "%Y%m%d")
    except ValueError:
        return None


def load_master(path):
    """기준정보 로드: Item code -> {name, seg1, 배면, 하대}"""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    next(rows)  # header
    m = {}
    for row in rows:
        if not row or row[0] is None:
            continue
        code = str(int(row[0])) if isinstance(row[0], (int, float)) else str(row[0]).strip()
        배면 = row[29] if len(row) > 29 else None  # 기준정보 AD열
        배단 = row[30] if len(row) > 30 else None  # 기준정보 AE열
        try:
            하대 = int(배면) * int(배단) if 배면 is not None and 배단 is not None else None
        except (TypeError, ValueError):
            하대 = None
        m[code] = {
            "name": row[1],
            "seg1": row[11] if len(row) > 11 else None,
            "배면": 배면,
            "하대": 하대,  # 자동계산 = 배면 × 배단
        }
    wb.close()
    return m


def load_stock(path):
    """ERP 재고조회 엑셀 로드 (시트명 'sheet')"""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["sheet"] if "sheet" in wb.sheetnames else wb.active
    rows = ws.iter_rows(values_only=True)
    next(rows)  # header
    records = []
    for row in rows:
        if not row or row[0] is None:
            continue
        code_raw = row[4]
        if code_raw is None:
            continue
        code = str(int(code_raw)) if isinstance(code_raw, (int, float)) else str(code_raw).strip()
        records.append({
            "fixed_loc": row[2],
            "location": row[3],
            "code": code,
            "name": row[5],
            "제조일": parse_ymd(row[8]),
            "유통기한": parse_ymd(row[9]),
            "유통기한_raw": row[9],
            "현재고_box": row[11] or 0,
            "현재고_ea": row[12] or 0,
            "출고가능_box": row[18] or 0,
            "출고가능_ea": row[19] or 0,
            "lock_box": row[24] if len(row) > 24 else None,
            "lock_ea": row[25] if len(row) > 25 else None,
        })
    wb.close()
    return records


def is_locked(r):
    lb, le = r["lock_box"], r["lock_ea"]
    return (lb is not None and lb != 0) or (le is not None and le != 0)


def is_ov_location(r):
    """OV로 시작하는 로케이션(OV1, OV5 등)은 이중적치 대상에서 제외"""
    loc = str(r.get("location", "")).strip().upper()
    return loc.startswith("OV")


def analyze_mixed(records, master=None):
    """혼재 적치 분석.
    제외: Lock 재고 / OV* 로케이션 / 잔량==0 (정파렛트로 딱 떨어지는 정상 보관).
    남은 행에서 같은 Location에 (제품코드, 유통기한) 조합 2개 이상 → 이중적치."""
    master = master or {}
    by_loc = defaultdict(list)
    for r in records:
        if is_locked(r) or is_ov_location(r):
            continue
        if r["location"] is None:
            continue

        m = master.get(r["code"], {})
        하대 = m.get("하대")
        현재고 = r["출고가능_box"] or 0
        try:
            정파렛트 = int(현재고) // int(하대) if 하대 else None
            잔량 = int(현재고) % int(하대) if 하대 else None
        except (TypeError, ValueError):
            정파렛트, 잔량 = None, None

        # 잔량이 0 → 정파렛트 완성 상태이므로 이중적치 판단에서 제외
        if 잔량 == 0:
            continue

        by_loc[r["location"]].append({
            "rec": r,
            "배면": m.get("배면"),
            "하대": 하대,
            "현재고": 현재고,
            "파렛트": 정파렛트,
            "잔량": 잔량,
        })

    detail = []
    for loc, items in sorted(by_loc.items()):
        combos = {(it["rec"]["code"], it["rec"]["유통기한"]) for it in items}
        if len(combos) < 2:
            continue

        items_sorted = sorted(items, key=lambda it: (it["rec"]["유통기한"] or datetime.max))
        for it in items_sorted:
            r = it["rec"]
            detail.append({
                "로케이션": loc,
                "제품코드": r["code"],
                "제품명": r["name"],
                "유통기한": r["유통기한"],
                "배면": it["배면"],
                "하대": it["하대"],
                "현재고": it["현재고"],
                "_빈": None,
                "파렛트": it["파렛트"],
                "잔량": it["잔량"],
                "특이사항": None,
            })
    return detail


def write_sheet(wb, sheet_name, rows, columns, date_cols=()):
    ws = wb.create_sheet(sheet_name)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    for i, col in enumerate(columns, 1):
        cell = ws.cell(row=1, column=i, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for r_idx, row in enumerate(rows, 2):
        for c_idx, col in enumerate(columns, 1):
            v = row.get(col)
            if col in date_cols and isinstance(v, datetime):
                ws.cell(row=r_idx, column=c_idx, value=v.strftime("%Y-%m-%d"))
            else:
                ws.cell(row=r_idx, column=c_idx, value=v)
    for i, col in enumerate(columns, 1):
        max_len = len(str(col))
        for row in rows[:200]:
            v = row.get(col)
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 30)
    ws.freeze_panes = "A2"


def write_보충오류_sheet(wb, rows):
    """원본 xlsm '보충오류' 시트 양식으로 출력.
    A:로케이션 B:제품코드 C:제품명 D:유통기한 E:배면 F:하대 G:현재고
    H:(빈) I:파렛트 J:잔량 K:특이사항"""
    ws = wb.create_sheet("보충오류")
    headers = ["로케이션", "제품코드", "제품명", "유통기한",
               "배면", "하대", "현재고", "", "파렛트", "잔량", "특이사항"]
    keys = ["로케이션", "제품코드", "제품명", "유통기한",
            "배면", "하대", "현재고", "_빈", "파렛트", "잔량", "특이사항"]
    widths = [12, 10, 24, 11, 6, 6, 8, 3, 8, 6, 14]

    # 1행: 날짜/시간 (원본 양식)
    ws.cell(row=1, column=1, value=datetime.today().strftime("%Y-%m-%d"))
    ws.cell(row=1, column=5, value=datetime.now().strftime("%Y-%m-%d %H:%M"))

    # 2행: 헤더
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 3행~ : 데이터
    for r_idx, row in enumerate(rows, 3):
        for c_idx, key in enumerate(keys, 1):
            v = row.get(key)
            if key == "유통기한" and isinstance(v, datetime):
                ws.cell(row=r_idx, column=c_idx, value=v.strftime("%Y-%m-%d"))
            else:
                ws.cell(row=r_idx, column=c_idx, value=v)

    # 컬럼 폭
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A3"


def analyze_and_save(stock_path, master_path, output_path, log=print):
    """ERP 재고 + 기준정보 -> 이중적치 결과 엑셀 저장 (보충오류 양식). 결과 dict 반환."""
    stock_path = Path(stock_path)
    master_path = Path(master_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    log(f"기준정보 로드: {master_path.name}")
    master = load_master(master_path)
    log(f"  제품 {len(master):,}건")

    log(f"재고 파일 로드: {stock_path.name}")
    records = load_stock(stock_path)
    log(f"  재고 {len(records):,}건")

    log("혼재 적치 분석 중...")
    detail = analyze_mixed(records, master)
    locs = {r["로케이션"] for r in detail}
    log(f"  혼재 Location {len(locs):,}개, 상세 {len(detail):,}건")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    write_보충오류_sheet(wb, detail)
    wb.save(output_path)
    log(f"저장 완료: {output_path.name}")

    return {
        "혼재Location수": len(locs),
        "상세건수": len(detail),
        "결과파일": str(output_path),
    }


# ============================================================
# 재고지(1단) 편집
# ============================================================

def load_location_list(path, sheet_name=None):
    """지정 로케이션 xlsx 로드. 로케이션 ID 집합 반환.
    헤더에서 '로케이션ID' / 'Location' 등 후보 컬럼을 찾고, 없으면 A열로 폴백.
    sheet_name 지정 시 해당 시트, 아니면 첫 시트."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    candidates = ("로케이션id", "로케이션 id", "location id", "location", "로케이션")
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    col_idx = 0
    for cand in candidates:
        for i, h in enumerate(header):
            if h is not None and str(h).strip().lower() == cand:
                col_idx = i
                break
        else:
            continue
        break

    s = set()
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or col_idx >= len(r) or r[col_idx] is None:
            continue
        v = str(r[col_idx]).strip()
        if v:
            s.add(v)
    wb.close()
    return s


def load_품목코드_from_제품별리스트(path, inv_code="IC930"):
    """제품별 리스트 xlsx → 출하Inv.가 inv_code인 행의 품목코드 set 반환.
    헤더에서 '출하Inv.' / '품목코드' 컬럼을 자동 탐색, 못 찾으면 A열/E열 폴백."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())

    def find_col(candidates, fallback):
        for cand in candidates:
            for i, h in enumerate(header):
                if h is not None and str(h).strip().lower() == cand:
                    return i
        return fallback

    inv_col = find_col(("출하inv.", "출하inv", "출하 inv.", "출하 inv", "inv.", "inv"), 0)
    code_col = find_col(("품목코드", "제품코드", "item code", "code"), 4)

    target = str(inv_code).strip().upper()
    s = set()
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or inv_col >= len(r) or code_col >= len(r):
            continue
        v_inv = r[inv_col]
        if v_inv is None:
            continue
        if str(v_inv).strip().upper() != target:
            continue
        v_code = r[code_col]
        if v_code is None:
            continue
        if isinstance(v_code, (int, float)):
            code = str(int(v_code))
        else:
            code = str(v_code).strip()
        if code:
            s.add(code)
    wb.close()
    return s


def load_차이수량_from_일일입력(path):
    """일일입력 엑셀 → {품목코드: 차이수량 합계} 매핑.
    - 시트: active
    - 헤더 자동: '품목코드'/'제품코드' (B 폴백), '차이수량' (D 폴백)
    - 같은 품목코드는 합산, int 변환 안 되는 행은 무시
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())

    def find_col(candidates, fallback):
        for cand in candidates:
            for i, h in enumerate(header):
                if h is not None and str(h).strip().lower() == cand:
                    return i
        return fallback

    code_col = find_col(("품목코드", "제품코드", "item code", "code"), 1)
    qty_col = find_col(("차이수량", "차이", "diff"), 3)

    out = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or code_col >= len(r) or qty_col >= len(r):
            continue
        v_code = r[code_col]
        v_qty = r[qty_col]
        if v_code is None or v_qty is None:
            continue
        if isinstance(v_code, (int, float)):
            code = str(int(v_code))
        else:
            code = str(v_code).strip()
        if not code:
            continue
        try:
            qty = int(v_qty)
        except (TypeError, ValueError):
            try:
                qty = int(float(v_qty))
            except (TypeError, ValueError):
                continue
        out[code] = out.get(code, 0) + qty
    wb.close()
    # 합산 0은 실제 차이 없음 → 음영/표시 대상에서 제외
    return {code: qty for code, qty in out.items() if qty != 0}


def _copy_cell_style(dst, src):
    """openpyxl 셀 스타일(폰트/테두리/채우기/정렬/표시형식)을 복제"""
    if src.has_style:
        dst.font = copy(src.font)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)


def write_재고지_1단_from_template(rows, output_path, highlight_set=None,
                                   template_path=None):
    """원본 xlsm 재고조사(1단) 시트를 템플릿으로 복제 → 데이터만 교체.
    양식/테두리/음영/폰트/컬럼폭/행높이가 원본 그대로 보존됨."""
    template_path = Path(template_path or TEMPLATE_1단)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(template_path)
    ws = wb["재고조사(1단)"]

    # 3행(데이터 샘플) 스타일을 셀별로 캐시
    cols = 11  # A:K (J=현차이수량, K=비고)
    row_style = [ws.cell(row=3, column=c) for c in range(1, cols + 1)]

    # 1행 날짜/시간 갱신 (원본은 =TODAY()/=NOW() 수식)
    ws.cell(row=1, column=1, value=datetime.today())
    ws.cell(row=1, column=1).number_format = "yyyy-mm-dd"
    ws.cell(row=1, column=7, value=datetime.now())
    ws.cell(row=1, column=7).number_format = "yyyy-mm-dd hh:mm"

    # 정렬
    rows_sorted = sorted(
        rows,
        key=lambda r: (r["로케이션"] or "", r["제품코드"] or "", r["유통기한"] or datetime.max)
    )

    highlight_fill = PatternFill("solid", fgColor="FFFF00")
    highlight_set = highlight_set or set()

    # 페이지 분할 초기화
    ws.row_breaks.brk = []

    prev_prefix = None
    for r_idx, r in enumerate(rows_sorted, 3):
        loc = r["로케이션"] or ""
        # 값 작성
        values = [
            loc,
            r["제품코드"],
            r["제품명"],
            r["유통기한_원본"],
            r["유통기한"],  # datetime 그대로 → 엑셀 날짜 셀
            None,  # F: 비고 (유통기한 옆)
            r["배면"],
            r["하대"],
            r["현재고"],
            r.get("차이수량"),  # J: 현차이수량
            None,  # K: 비고
        ]
        for c, v in enumerate(values, 1):
            cell = ws.cell(row=r_idx, column=c, value=v)
            _copy_cell_style(cell, row_style[c - 1])

        # E열은 날짜로 포맷
        ws.cell(row=r_idx, column=5).number_format = "yyyy-mm-dd"

        # 강조 로케이션 전체 행 노란색
        if loc in highlight_set:
            for c in range(1, cols + 1):
                ws.cell(row=r_idx, column=c).fill = highlight_fill

        # 차이수량 있는 품목 → B열(제품코드) 셀만 노란색 음영
        if r.get("차이수량") is not None:
            ws.cell(row=r_idx, column=2).fill = highlight_fill

        # 로케이션 접두어("-" 앞) 변경 시 페이지 분할
        cur_prefix = loc.split("-", 1)[0] if "-" in loc else loc
        if prev_prefix is not None and cur_prefix != prev_prefix:
            ws.row_breaks.append(Break(id=r_idx - 1))
        prev_prefix = cur_prefix

    # 인쇄 영역 갱신 (데이터 끝까지)
    last_row = max(2, 2 + len(rows_sorted))
    ws.print_area = f"A1:K{last_row}"

    wb.save(output_path)
    return len(rows_sorted)


def edit_재고지_1단(stock_path, master_path, loc_list_path,
                    output_path, highlight_path=None, code_filter=None,
                    diff_qty=None, log=print):
    """ERP 재고 + 기준정보 + 지정로케이션 → 1단 재고지 엑셀 생성.
    - 단수 = 로케이션 끝 2자리 == '10'
    - 지정로케이션에 있는 로케이션만 포함
    - highlight_path(선택)에 있는 로케이션은 노란색 강조
    - code_filter(선택)에 있는 품목코드만 포함 (None이면 무필터)
    - diff_qty(선택): {품목코드: 차이수량} dict. 해당 품목은 현재고 0이어도 포함,
      B열(제품코드) 음영 + J열(현차이수량)에 값 표시
    - 현재고 = 변환재고 (출고가능_box + lock_box)
    """
    stock_path = Path(stock_path)
    master_path = Path(master_path)
    loc_list_path = Path(loc_list_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    log(f"기준정보 로드: {master_path.name}")
    master = load_master(master_path)
    log(f"  제품 {len(master):,}건")

    log(f"지정 로케이션 로드: {loc_list_path.name}")
    loc_set = load_location_list(loc_list_path)
    log(f"  로케이션 {len(loc_set):,}건")

    highlight_set = set()
    if highlight_path:
        hp = Path(highlight_path)
        if hp.exists():
            highlight_set = load_location_list(hp)
            log(f"강조 로케이션: {len(highlight_set):,}건")

    log(f"재고 파일 로드: {stock_path.name}")
    records = load_stock(stock_path)
    log(f"  재고 {len(records):,}건")

    if code_filter is not None:
        log(f"품목 필터: {len(code_filter):,}건")
    if diff_qty:
        log(f"차이수량 입력 {len(diff_qty):,}건 — 음영/현차이수량 표시")

    log("1단 재고지 편집 중...")
    rows = []
    for r in records:
        loc = r["location"]
        if loc is None:
            continue
        loc = str(loc).strip()
        # 단수 = 끝 2자리 == '10'
        if not loc.endswith("10"):
            continue
        if loc not in loc_set:
            continue
        if code_filter is not None and r["code"] not in code_filter:
            continue

        m = master.get(r["code"], {})
        현재고 = (r["출고가능_box"] or 0) + (r["lock_box"] or 0)
        has_diff = bool(diff_qty) and r["code"] in diff_qty
        if 현재고 == 0 and not has_diff:
            continue  # 현재고 0이면 제외 — 단, 차이수량 있는 품목은 포함
        rows.append({
            "로케이션": loc,
            "제품코드": r["code"],
            "제품명": r["name"],
            "유통기한_원본": r["유통기한_raw"],
            "유통기한": r["유통기한"],
            "배면": m.get("배면"),
            "하대": m.get("하대"),
            "현재고": 현재고,
            "차이수량": diff_qty.get(r["code"]) if diff_qty else None,
        })

    # 같은 품목코드가 여러 행이면, 가장 늦은 유통기한 행에만 차이수량 유지
    if diff_qty:
        by_code = defaultdict(list)
        for i, row in enumerate(rows):
            if row["차이수량"] is not None:
                by_code[row["제품코드"]].append(i)
        for idxs in by_code.values():
            if len(idxs) < 2:
                continue
            best = max(idxs, key=lambda i: rows[i]["유통기한"] or datetime.min)
            for i in idxs:
                if i != best:
                    rows[i]["차이수량"] = None
    log(f"  편집 대상 {len(rows):,}행")

    if not TEMPLATE_1단.exists():
        raise FileNotFoundError(
            f"템플릿 파일이 없습니다: {TEMPLATE_1단}\n"
            f"원본 xlsm의 재고조사(1단) 시트를 추출해서 저장하세요."
        )

    log(f"템플릿 기반 출력: {TEMPLATE_1단.name}")
    n = write_재고지_1단_from_template(rows, output_path, highlight_set)
    log(f"저장 완료: {output_path.name}")

    locs_in_result = {r["로케이션"] for r in rows}
    return {
        "행수": n,
        "로케이션수": len(locs_in_result),
        "결과파일": str(output_path),
    }


# ============================================================
# 재고지(1단) — 전체 버전 (통합센터 재고조사 3.E 양식, 8열, 필터 없음)
# ============================================================

def write_재고지_1단_전체_from_template(rows, output_path, template_path=None):
    """통합센터 3.E 파일의 재고조사(1단) 템플릿 기반 출력 (8열).
    A:로케이션 B:제품코드 C:제품명 D:유통기한(전) E:유통기한 F:배면 G:하대 H:현재고"""
    template_path = Path(template_path or TEMPLATE_1단_전체)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(template_path)
    ws = wb["재고조사(1단)"]

    cols = 8
    row_style = [ws.cell(row=3, column=c) for c in range(1, cols + 1)]

    # 1행 날짜/시간
    ws.cell(row=1, column=1, value=datetime.today())
    ws.cell(row=1, column=1).number_format = "yyyy-mm-dd"
    ws.cell(row=1, column=6, value=datetime.now())
    ws.cell(row=1, column=6).number_format = "yyyy-mm-dd hh:mm"

    rows_sorted = sorted(
        rows,
        key=lambda r: (r["로케이션"] or "", r["제품코드"] or "", r["유통기한"] or datetime.max)
    )

    ws.row_breaks.brk = []
    prev_prefix = None
    for r_idx, r in enumerate(rows_sorted, 3):
        loc = r["로케이션"] or ""
        values = [
            loc,
            r["제품코드"],
            r["제품명"],
            r["유통기한_원본"],
            r["유통기한"],
            r["배면"],
            r["하대"],
            r["현재고"],
        ]
        for c, v in enumerate(values, 1):
            cell = ws.cell(row=r_idx, column=c, value=v)
            _copy_cell_style(cell, row_style[c - 1])
        # E열 날짜 형식
        ws.cell(row=r_idx, column=5).number_format = "yyyy-mm-dd"

        cur_prefix = loc.split("-", 1)[0] if "-" in loc else loc
        if prev_prefix is not None and cur_prefix != prev_prefix:
            ws.row_breaks.append(Break(id=r_idx - 1))
        prev_prefix = cur_prefix

    last_row = max(2, 2 + len(rows_sorted))
    ws.print_area = f"A1:H{last_row}"

    wb.save(output_path)
    return len(rows_sorted)


def edit_재고지_1단_전체(stock_path, master_path, output_path, log=print):
    """ERP 재고 + 기준정보 → 1단 전체 재고지 (지정 로케이션 필터 없음, 8열 양식).
    - 단수 = 로케이션 끝 2자리 == '10'
    - 현재고 = 변환재고 (출고가능_box + lock_box)
    """
    stock_path = Path(stock_path)
    master_path = Path(master_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    log(f"기준정보 로드: {master_path.name}")
    master = load_master(master_path)
    log(f"  제품 {len(master):,}건")

    log(f"재고 파일 로드: {stock_path.name}")
    records = load_stock(stock_path)
    log(f"  재고 {len(records):,}건")

    log("1단 전체 재고지 편집 중...")
    rows = []
    for r in records:
        loc = r["location"]
        if loc is None:
            continue
        loc = str(loc).strip()
        if not loc.endswith("10"):
            continue

        m = master.get(r["code"], {})
        현재고 = (r["출고가능_box"] or 0) + (r["lock_box"] or 0)
        if 현재고 == 0:
            continue  # 재고 없는 행 제외 (원본 xlsm의 G=0 행 삭제와 동일)
        rows.append({
            "로케이션": loc,
            "제품코드": r["code"],
            "제품명": r["name"],
            "유통기한_원본": r["유통기한_raw"],
            "유통기한": r["유통기한"],
            "배면": m.get("배면"),
            "하대": m.get("하대"),
            "현재고": 현재고,
        })
    log(f"  편집 대상 {len(rows):,}행")

    if not TEMPLATE_1단_전체.exists():
        raise FileNotFoundError(
            f"템플릿 파일이 없습니다: {TEMPLATE_1단_전체}"
        )

    log(f"템플릿 기반 출력: {TEMPLATE_1단_전체.name}")
    n = write_재고지_1단_전체_from_template(rows, output_path)
    log(f"저장 완료: {output_path.name}")

    locs_in_result = {r["로케이션"] for r in rows}
    return {
        "행수": n,
        "로케이션수": len(locs_in_result),
        "결과파일": str(output_path),
    }


# ============================================================
# 양식 다운로드 (헤더 + 샘플 행만 추출)
# ============================================================

def create_양식(source_path, output_path, sample_rows=3):
    """원본 xlsx에서 헤더 + 상위 N건 샘플만 추출해 양식 파일로 저장.
    컬럼 폭, 헤더 스타일은 원본 그대로 유지."""
    source_path = Path(source_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    src = openpyxl.load_workbook(source_path, data_only=True)
    ws_src = src.active

    dst = openpyxl.Workbook()
    ws_dst = dst.active
    ws_dst.title = ws_src.title

    max_col = ws_src.max_column
    copy_rows = sample_rows + 1  # header + N sample

    for r_idx in range(1, copy_rows + 1):
        for c_idx in range(1, max_col + 1):
            src_cell = ws_src.cell(row=r_idx, column=c_idx)
            new_cell = ws_dst.cell(row=r_idx, column=c_idx, value=src_cell.value)
            if src_cell.has_style:
                new_cell.font = copy(src_cell.font)
                new_cell.fill = copy(src_cell.fill)
                new_cell.border = copy(src_cell.border)
                new_cell.alignment = copy(src_cell.alignment)
                new_cell.number_format = src_cell.number_format

    # 컬럼 폭/숨김 복사
    for col_letter, dim in ws_src.column_dimensions.items():
        if dim.width:
            ws_dst.column_dimensions[col_letter].width = dim.width
        if dim.hidden:
            ws_dst.column_dimensions[col_letter].hidden = dim.hidden

    # 안내 행 추가
    note_row = copy_rows + 2
    note = ws_dst.cell(
        row=note_row, column=1,
        value="※ 위 형식에 맞춰 데이터를 채우고 [업데이트] 버튼으로 교체하세요. "
              "2행부터가 실제 데이터입니다."
    )
    note.font = Font(italic=True, color="888888", size=9)

    src.close()
    dst.save(output_path)
    return copy_rows


# ============================================================
# 재고지 (2~6단 통합) — 260325 재고실사 양식, 10열, 전산/실물/비고 분할
# ============================================================

def _단_suffix(loc):
    """로케이션 끝 2자리 -> 단 번호 (int) 또는 None"""
    if not loc or len(loc) < 2:
        return None
    suf = loc[-2:]
    if not suf.isdigit():
        return None
    n = int(suf)
    if n in (20, 30, 40, 50, 60):
        return n // 10
    return None


def write_재고지_단별_from_template(rows_by_단, output_path, template_path=None):
    """260325 재고실사 템플릿 기반. 2~6단 시트에 각각 데이터 채움.
    10열 양식: A로케이션 B제품코드 C제품명 D유통기한(전,숨김) E유통기한
              F배면 G하대 H재고(전산) I재고(실물,빈) J비고(빈)
    데이터는 4행부터 시작 (1:날짜, 2:헤더, 3:전산/실물/비고)"""
    template_path = Path(template_path or TEMPLATE_단별)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(template_path)
    total_rows = 0

    for 단 in (2, 3, 4, 5, 6):
        sheet_name = f"{단}단"
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]

        # 4행(데이터 샘플) 스타일 참고가 없으니, 템플릿 데이터가 지워졌더라도
        # 3행까지의 스타일을 참고. 대신 실제 채울 데이터 셀은 스타일 없이 기본 적용.
        # → 원본 템플릿의 4행이 비어있고 openpyxl이 스타일을 가지고 있지 않을 수 있으므로
        #   헤더(2,3행)과 비슷한 테두리를 수동 적용하는 게 안전.

        # 1행 날짜/시간
        ws.cell(row=1, column=1, value=datetime.today())
        ws.cell(row=1, column=1).number_format = "yyyy-mm-dd"
        ws.cell(row=1, column=6, value=datetime.now())
        ws.cell(row=1, column=6).number_format = "yyyy-mm-dd hh:mm"

        rows = rows_by_단.get(단, [])
        rows_sorted = sorted(
            rows,
            key=lambda r: (r["로케이션"] or "", r["제품코드"] or "", r["유통기한"] or datetime.max)
        )

        # 4행의 스타일을 참고할 수 있으면 사용
        sample_cells = [ws.cell(row=4, column=c) for c in range(1, 11)]
        has_sample_style = any(c.has_style for c in sample_cells)

        # 페이지 분할 초기화
        ws.row_breaks.brk = []

        prev_prefix = None
        for r_idx, r in enumerate(rows_sorted, 4):
            loc = r["로케이션"] or ""
            values = [
                loc,
                r["제품코드"],
                r["제품명"],
                r["유통기한_원본"],
                r["유통기한"],
                r["배면"],
                r["하대"],
                r["현재고"],  # H: 전산
                None,          # I: 실물(빈)
                None,          # J: 비고(빈)
            ]
            for c, v in enumerate(values, 1):
                cell = ws.cell(row=r_idx, column=c, value=v)
                if has_sample_style:
                    _copy_cell_style(cell, sample_cells[c - 1])
                else:
                    # 기본 테두리/폰트 (원본 스타일 없을 때)
                    cell.font = Font(name="맑은 고딕", size=11)
                    cell.border = Border(
                        left=Side(style="thin"), right=Side(style="thin"),
                        top=Side(style="thin"), bottom=Side(style="thin"),
                    )
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            # E열 날짜 포맷
            ws.cell(row=r_idx, column=5).number_format = "yyyy-mm-dd"
            # C열은 좌측 정렬이 자연스러움
            ws.cell(row=r_idx, column=3).alignment = Alignment(horizontal="left", vertical="center")

            cur_prefix = loc.split("-", 1)[0] if "-" in loc else loc
            if prev_prefix is not None and cur_prefix != prev_prefix:
                ws.row_breaks.append(Break(id=r_idx - 1))
            prev_prefix = cur_prefix

        last_row = max(3, 3 + len(rows_sorted))
        ws.print_area = f"A1:J{last_row}"
        total_rows += len(rows_sorted)

    wb.save(output_path)
    return total_rows


def edit_재고지_2_6단(stock_path, master_path, output_path, log=print):
    """ERP 재고 + 기준정보 → 2~6단 통합 재고지 엑셀 생성 (한 파일 5개 시트).
    - 단 = 로케이션 끝 2자리 (20/30/40/50/60) → 2/3/4/5/6단
    - 3/4/5단: 재고 있는 것만 (현재고 > 0)
    - 2단/6단: 로케이션 마스터 기준, 재고 없는 빈 로케이션도 포함
      ('재고지_2_6단_로케이션.xlsx'의 '2단','6단' 시트에서 로드)
    """
    stock_path = Path(stock_path)
    master_path = Path(master_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    log(f"기준정보 로드: {master_path.name}")
    master = load_master(master_path)
    log(f"  제품 {len(master):,}건")

    # 2단/6단 로케이션 마스터 로드 (있으면)
    full_loc_by_단 = {}
    if LOC_LIST_2_6단.exists():
        try:
            full_loc_by_단[2] = load_location_list(LOC_LIST_2_6단, sheet_name="2단")
            full_loc_by_단[6] = load_location_list(LOC_LIST_2_6단, sheet_name="6단")
            log(f"2단/6단 로케이션 마스터: 2단 {len(full_loc_by_단[2]):,}건, "
                f"6단 {len(full_loc_by_단[6]):,}건")
        except Exception as e:
            log(f"※ 로케이션 마스터 로드 실패({e}) — 재고 있는 것만 출력")
            full_loc_by_단 = {}
    else:
        log(f"※ {LOC_LIST_2_6단.name} 없음 — 재고 있는 것만 출력")

    log(f"재고 파일 로드: {stock_path.name}")
    records = load_stock(stock_path)
    log(f"  재고 {len(records):,}건")

    log("2~6단 재고지 편집 중...")

    # 단별 로케이션별 레코드 모음
    recs_by_loc = {단: defaultdict(list) for 단 in (2, 3, 4, 5, 6)}
    for r in records:
        loc = r["location"]
        if loc is None:
            continue
        loc = str(loc).strip()
        단 = _단_suffix(loc)
        if 단 is None:
            continue

        m = master.get(r["code"], {})
        현재고 = (r["출고가능_box"] or 0) + (r["lock_box"] or 0)
        if 현재고 == 0:
            continue
        recs_by_loc[단][loc].append({
            "로케이션": loc,
            "제품코드": r["code"],
            "제품명": r["name"],
            "유통기한_원본": r["유통기한_raw"],
            "유통기한": r["유통기한"],
            "배면": m.get("배면"),
            "하대": m.get("하대"),
            "현재고": 현재고,
        })

    # 단별 rows 구성
    rows_by_단 = {단: [] for 단 in (2, 3, 4, 5, 6)}
    stats = {}
    for 단 in (2, 3, 4, 5, 6):
        if 단 in full_loc_by_단:
            # 2단/6단: 로케이션 마스터 기준 (빈 로케이션 포함)
            full_set = full_loc_by_단[단] | set(recs_by_loc[단].keys())
            filled = 0
            empty = 0
            for loc in sorted(full_set):
                if loc in recs_by_loc[단]:
                    rows_by_단[단].extend(recs_by_loc[단][loc])
                    filled += len(recs_by_loc[단][loc])
                else:
                    rows_by_단[단].append({
                        "로케이션": loc,
                        "제품코드": None,
                        "제품명": None,
                        "유통기한_원본": None,
                        "유통기한": None,
                        "배면": None,
                        "하대": None,
                        "현재고": None,
                    })
                    empty += 1
            stats[f"{단}단"] = (filled, empty)
            log(f"  {단}단: 재고있음 {filled:,}행 + 빈 로케이션 {empty:,}행 = {len(rows_by_단[단]):,}행")
        else:
            # 3/4/5단: 재고 있는 것만
            for loc, recs in recs_by_loc[단].items():
                rows_by_단[단].extend(recs)
            stats[f"{단}단"] = (len(rows_by_단[단]), 0)
            log(f"  {단}단: {len(rows_by_단[단]):,}행")

    if not TEMPLATE_단별.exists():
        raise FileNotFoundError(
            f"템플릿 파일이 없습니다: {TEMPLATE_단별}"
        )

    log(f"템플릿 기반 출력: {TEMPLATE_단별.name}")
    total = write_재고지_단별_from_template(rows_by_단, output_path)
    log(f"저장 완료: {output_path.name}")

    return {
        "총_행수": total,
        "단별_행수": {f"{k}단": len(v) for k, v in rows_by_단.items()},
        "2단_6단_빈로케이션": {k: stats[k][1] for k in ("2단", "6단") if k in stats},
        "결과파일": str(output_path),
    }


# ============================================================
# LOCK 재고 유통기한 체크 (잔존율 <= 50%)
# ============================================================

def load_담당자(path=None):
    """물품담당자 xlsx 로드 (A: 코드, B: 담당자). 코드 -> 담당자(trim) dict 반환.
    파일 없으면 빈 dict."""
    path = Path(path or 담당자_PATH)
    if not path.exists():
        return {}
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    m = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        code = str(int(row[0])) if isinstance(row[0], (int, float)) else str(row[0]).strip()
        name = str(row[1]).strip() if row[1] is not None else None
        if code and name:
            m[code] = name
    wb.close()
    return m


def _write_공유_sheet(out_rows, output_path):
    """공유 시트 양식(10열) + K열 '잔존율 40% 도달일'."""
    if not TEMPLATE_LOCK_유통기한.exists():
        raise FileNotFoundError(f"템플릿 파일이 없습니다: {TEMPLATE_LOCK_유통기한}")

    wb = openpyxl.load_workbook(TEMPLATE_LOCK_유통기한)
    ws = wb.active

    # K열 헤더 추가 (J1 스타일 복제)
    k_header = ws.cell(row=1, column=11, value="잔존율 40% 도달일")
    _copy_cell_style(k_header, ws.cell(row=1, column=10))
    ws.column_dimensions["K"].width = 16

    keys = ["품목코드", "품목명", "유통기한_원본", "수량", "유통기한",
            "공유여부", "처리방안", "잔존일수", "잔존개월", "잔존율"]

    for r_idx, row in enumerate(out_rows, 2):
        values = [row[k] for k in keys]
        for c, v in enumerate(values, 1):
            cell = ws.cell(row=r_idx, column=c, value=v)
            cell.font = Font(name="맑은 고딕", size=10)
            cell.border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"),
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=r_idx, column=2).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=r_idx, column=5).number_format = "yyyy-mm-dd"
        ws.cell(row=r_idx, column=9).number_format = "0.00"
        ws.cell(row=r_idx, column=10).number_format = "0.00%"

        # K열: 잔존율 40% 도달일
        k_val = row.get("잔존율40_도달일")
        k_cell = ws.cell(row=r_idx, column=11, value=k_val)
        k_cell.font = Font(name="맑은 고딕", size=10)
        k_cell.border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        k_cell.alignment = Alignment(horizontal="center", vertical="center")
        k_cell.number_format = "yyyy-mm-dd"

    last_row = max(1, 1 + len(out_rows))
    ws.print_area = f"A1:K{last_row}"
    ws.freeze_panes = "A2"
    wb.save(output_path)


def _build_expiry_row(r, today, threshold):
    """한 레코드에 대해 잔존율 계산 후 통과하면 dict 반환, 아니면 None.
    수량은 총재고 박스(없으면 EA). K열용 잔존율 40% 도달일 포함."""
    from datetime import timedelta
    mfg = r["제조일"]
    exp = r["유통기한"]
    if mfg is None or exp is None:
        return None, "date"
    total_days = (exp - mfg).days
    remain_days = (exp - today).days
    if total_days <= 0:
        return None, "date"
    잔존율 = remain_days / total_days
    if 잔존율 > threshold:
        return None, "over"

    # 수량: 현재고_box 우선
    qty_box = r["현재고_box"] or 0
    qty_ea = r["현재고_ea"] or 0
    qty = qty_box if qty_box > 0 else qty_ea

    date_40 = mfg + timedelta(days=int(total_days * 0.6))

    return {
        "품목코드": r["code"],
        "품목명": r["name"],
        "유통기한_원본": r["유통기한_raw"],
        "수량": qty,
        "유통기한": exp,
        "공유여부": None,
        "처리방안": None,
        "잔존일수": remain_days,
        "잔존개월": round(remain_days / 30, 2),
        "잔존율": round(잔존율, 4),
        "잔존율40_도달일": date_40,
    }, "ok"


def _has_any_stock(r):
    """LOCK/출고가능/현재고 중 하나라도 있으면 True"""
    for k in ("lock_box", "lock_ea", "출고가능_box", "출고가능_ea", "현재고_box", "현재고_ea"):
        v = r.get(k) or 0
        if v:
            return True
    return False


def _is_locked(r):
    """LOCK 재고 있음"""
    return (r.get("lock_box") or 0) > 0 or (r.get("lock_ea") or 0) > 0


def analyze_ov5_expiry(stock_path, master_path, output_path,
                       threshold=0.5, today=None, log=print):
    """OV5 로케이션 재고 전체(LOCK 여부 무관) 중 잔존율 ≤ threshold.
    결과 0건이면 파일 저장하지 않음."""
    stock_path = Path(stock_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    today = today or datetime.today()

    log(f"재고 파일 로드: {stock_path.name}")
    records = load_stock(stock_path)
    log(f"  재고 {len(records):,}건")

    담당자 = load_담당자()
    if 담당자:
        log(f"담당자 마스터: {len(담당자):,}건")

    log(f"OV5 + 잔존율 ≤ {int(threshold*100)}% 추출 중...")
    n_ov5 = n_stock = 0
    out_rows = []
    for r in records:
        loc = str(r.get("location") or "").strip().upper()
        if loc != "OV5":
            continue
        n_ov5 += 1
        if not _has_any_stock(r):
            continue
        n_stock += 1
        row, _ = _build_expiry_row(r, today, threshold)
        if row:
            row["공유여부"] = 담당자.get(row["품목코드"])
            out_rows.append(row)

    out_rows.sort(key=lambda x: (x["잔존율"], x["잔존일수"]))
    log(f"  OV5 {n_ov5}건 → 재고있음 {n_stock}건 → 잔존율≤{int(threshold*100)}% {len(out_rows)}건")

    if not out_rows:
        log("→ 대상 없음. 파일 저장을 생략합니다.")
        return {
            "대상건수": 0,
            "OV5_총": n_ov5,
            "OV5_재고있음": n_stock,
            "결과파일": None,
        }

    _write_공유_sheet(out_rows, output_path)
    log(f"저장 완료: {output_path.name}")
    return {
        "대상건수": len(out_rows),
        "OV5_총": n_ov5,
        "OV5_재고있음": n_stock,
        "결과파일": str(output_path),
    }


def analyze_ov6_expiry(stock_path, master_path, output_path,
                       threshold=0.5, today=None, log=print):
    """OV6 로케이션 재고 전체(LOCK 여부 무관) 중 잔존율 ≤ threshold.
    결과 0건이면 파일 저장하지 않음."""
    stock_path = Path(stock_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    today = today or datetime.today()

    log(f"재고 파일 로드: {stock_path.name}")
    records = load_stock(stock_path)
    log(f"  재고 {len(records):,}건")

    담당자 = load_담당자()
    if 담당자:
        log(f"담당자 마스터: {len(담당자):,}건")

    log(f"OV6 + 잔존율 ≤ {int(threshold*100)}% 추출 중...")
    n_ov6 = n_stock = 0
    out_rows = []
    for r in records:
        loc = str(r.get("location") or "").strip().upper()
        if loc != "OV6":
            continue
        n_ov6 += 1
        if not _has_any_stock(r):
            continue
        n_stock += 1
        row, _ = _build_expiry_row(r, today, threshold)
        if row:
            row["공유여부"] = 담당자.get(row["품목코드"])
            out_rows.append(row)

    out_rows.sort(key=lambda x: (x["잔존율"], x["잔존일수"]))
    log(f"  OV6 {n_ov6}건 → 재고있음 {n_stock}건 → 잔존율≤{int(threshold*100)}% {len(out_rows)}건")

    if not out_rows:
        log("→ 대상 없음. 파일 저장을 생략합니다.")
        return {
            "대상건수": 0,
            "OV6_총": n_ov6,
            "OV6_재고있음": n_stock,
            "결과파일": None,
        }

    _write_공유_sheet(out_rows, output_path)
    log(f"저장 완료: {output_path.name}")
    return {
        "대상건수": len(out_rows),
        "OV6_총": n_ov6,
        "OV6_재고있음": n_stock,
        "결과파일": str(output_path),
    }


def analyze_nonlock_expiry(stock_path, master_path, output_path,
                           threshold=0.5, today=None, log=print):
    """LOCK 제외 + 출고가능 재고가 있는 것 중 잔존율 ≤ threshold.
    같은 (제품코드, 유통기한) 조합은 출고가능 수량을 합산하여 한 행으로 표시.
    결과 0건이면 파일 저장하지 않음."""
    stock_path = Path(stock_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    today = today or datetime.today()

    log(f"재고 파일 로드: {stock_path.name}")
    records = load_stock(stock_path)
    log(f"  재고 {len(records):,}건")

    담당자 = load_담당자()
    if 담당자:
        log(f"담당자 마스터: {len(담당자):,}건")

    log(f"LOCK 제외 + 출고가능 있음 + 잔존율 ≤ {int(threshold*100)}% 추출 (수량 총합)...")
    n_nonlock = n_avail = 0
    agg = {}                    # key=(code, 유통기한_raw) -> row dict
    qty_box_sum = defaultdict(float)
    qty_ea_sum = defaultdict(float)
    raw_rows = 0

    for r in records:
        if _is_locked(r):
            continue
        n_nonlock += 1

        # 출고가능 재고가 있는 것만
        ab = r["출고가능_box"] or 0
        ae = r["출고가능_ea"] or 0
        if ab == 0 and ae == 0:
            continue
        n_avail += 1

        mfg = r["제조일"]
        exp = r["유통기한"]
        if mfg is None or exp is None:
            continue
        total_days = (exp - mfg).days
        remain_days = (exp - today).days
        if total_days <= 0:
            continue
        잔존율 = remain_days / total_days
        if 잔존율 > threshold:
            continue

        raw_rows += 1
        key = (r["code"], r["유통기한_raw"])
        if key not in agg:
            # 잔존율 40% 도달일 = 제조일 + 총일수 * 70%
            from datetime import timedelta
            date_40 = mfg + timedelta(days=int(total_days * 0.6))
            agg[key] = {
                "품목코드": r["code"],
                "품목명": r["name"],
                "유통기한_원본": r["유통기한_raw"],
                "수량": 0,
                "유통기한": exp,
                "공유여부": None,
                "처리방안": None,
                "잔존일수": remain_days,
                "잔존개월": round(remain_days / 30, 2),
                "잔존율": round(잔존율, 4),
                "잔존율40_도달일": date_40,
            }
        qty_box_sum[key] += ab
        qty_ea_sum[key] += ae

    # 수량 결정 + 담당자 매핑
    out_rows = []
    for key, row in agg.items():
        qb = qty_box_sum[key]
        qe = qty_ea_sum[key]
        row["수량"] = int(qb) if qb > 0 else int(qe)
        row["공유여부"] = 담당자.get(row["품목코드"])
        out_rows.append(row)

    out_rows.sort(key=lambda x: (x["잔존율"], x["잔존일수"]))
    log(f"  LOCK 제외 {n_nonlock:,}건 → 출고가능 있음 {n_avail:,}건 "
        f"→ 잔존율≤{int(threshold*100)}% {raw_rows:,}행 "
        f"→ (제품,유통기한) 합산 {len(out_rows):,}건")

    if not out_rows:
        log("→ 대상 없음. 파일 저장을 생략합니다.")
        return {
            "대상건수": 0,
            "비락_총": n_nonlock,
            "출고가능_있음": n_avail,
            "합산전_행수": raw_rows,
            "결과파일": None,
        }

    _write_공유_sheet(out_rows, output_path)
    log(f"저장 완료: {output_path.name}")
    return {
        "대상건수": len(out_rows),
        "비락_총": n_nonlock,
        "출고가능_있음": n_avail,
        "합산전_행수": raw_rows,
        "결과파일": str(output_path),
    }

