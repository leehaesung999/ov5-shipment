# -*- coding: utf-8 -*-
"""잔존율 출고 판정기

매일 하는 일: **로케이션별 재고조회 1개만 업로드** → 판정 결과 확인/다운로드.

내장 기준정보 (업로드 불필요)
  · 거래처 잔존율   data/customers.json  — 통합 마스터 기반(고정). 화면에서 개별 조정 가능
  · 품목 기준정보   data/items.json      — 품명·소비기한(월). 월 1회 정도만 갱신

잔존율(일 기준) = (소비기한 − 기준일) ÷ (소비기한 − 제조일)
  · 50% 기준(하프도달)은 전 거래처 공통 정책이라 판정 대상에서 항상 제외한다.

판정
  · 현행유지 : 가장 빠른(FIFO) 로트가 이미 기준 충족 → 그대로 출고
  · 변경필요 : 최선 로트가 미달 → 기준을 만족하는 '가장 빠른' 소비기한으로 변경
  · 출고불가 : 기준을 만족하는 로트가 하나도 없음
"""
import io
import json
import re
from collections import defaultdict
from datetime import date, datetime
# --- KST (Streamlit Cloud는 UTC 기본) ---
try:
    KST  # noqa: F821
except NameError:
    from datetime import datetime as _dt, timedelta as _td, timezone as _tz
    KST = _tz(_td(hours=9))
    def _now_kst():
        return _dt.now(KST)
    def _now_kst_naive():
        return _dt.now(KST).replace(tzinfo=None)
    def _today_kst():
        return _dt.now(KST).date()

# --- KST (Streamlit Cloud는 UTC 기본) ---
try:
    KST  # noqa: F821
except NameError:
    from datetime import timedelta as _td, timezone as _tz
    KST = _tz(_td(hours=9))
    def _now_kst():
        return datetime.now(KST)

from pathlib import Path

import openpyxl
import pandas as pd
import streamlit as st

try:
    st.set_page_config(page_title='잔존율 출고 판정기', page_icon='📅', layout='wide')
except Exception:
    pass

APP_DIR = Path(__file__).parent
import sys  # noqa: E402
sys.path.insert(0, str(APP_DIR.parent))      # 레포 루트 (공용 master_hub)
from master_hub import store as hub  # noqa: E402
DATA_DIR = APP_DIR / 'data'
DATA_DIR.mkdir(exist_ok=True)
CUST_FILE = DATA_DIR / 'customers.json'      # 거래처 잔존율 (고정 + 조정분 반영)
ITEM_FILE = DATA_DIR / 'items.json'          # 품목 기준정보 (월 1회 갱신)

HALF = 0.5          # 하프도달 — 항상 판정 제외

# 출고 기준 인벤토리 — 이 인벤토리 재고만 판정 대상(기본 출고분)으로 본다.
SHIP_INVENTORY = 'IC930'

# 채널 코드 → 정식 명칭 (화면 표기는 '풀네임(코드)')
CHANNEL_NAMES = {
    'CVS': '편의점', 'DPS': '백화점', 'GT': '대리점', 'KAT': '대형마트',
    'MS': '식자재·급식', 'NC': '이커머스', 'SSM': '기업형슈퍼',
    'NH': '농협', 'SPS': '특판',
}


def ch_disp(ch):
    """'GT' → '대리점(GT)'. 매핑 없으면 코드 그대로."""
    full = CHANNEL_NAMES.get(str(ch).strip())
    return f'{full}({ch})' if full else str(ch)


# 지정 출고처 표기 시 개별 거래처 대신 채널명 하나로 묶을 채널.
# 채널 안에서 잔존율 기준이 '통일'된 곳만 안전하게 묶는다(항상 같은 로트로 나감).
# GT(대리점 75%), DPS(백화점 70%)만 통일. 나머지는 기준 혼재라 개별 표기.
COLLAPSE_CHANNELS = {'GT', 'DPS'}


def base_cust_name(s):
    """거래처명 뒤의 괄호 구분자를 떼어 대표명으로. 예:
    '신세계백화점(특정)'·'신세계백화점(햅스토어)' → '신세계백화점',
    '롯데마트(창고형)'·'롯데마트(할인점)' → '롯데마트'.
    맨 앞이 '(주)' 처럼 괄호로 시작하는 건 건드리지 않는다(끝의 괄호만)."""
    return re.sub(r'\s*\(.*\)\s*$', '', str(s)).strip() or str(s).strip()


# ─────────────────────────── 공통 유틸 ───────────────────────────
def to_date(v):
    """20280722 / '2028-07-22' / datetime → date. 실패 시 None."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip().replace('-', '').replace('/', '').replace('.', '')
    if s.endswith('0') and len(s) == 10 and str(v).strip().endswith('.0'):
        s = s[:-1]
    if len(s) == 8 and s.isdigit():
        try:
            return date(int(s[:4]), int(s[4:6]), int(s[6:]))
        except ValueError:
            return None
    return None


def to_float(v):
    if v is None:
        return 0.0
    try:
        return float(str(v).replace(',', '').strip())
    except ValueError:
        return 0.0


def code_key(v):
    """품목코드 정규화: 1010122.0 → '1010122'."""
    if v is None or v == '':
        return None
    try:
        return str(int(v))
    except (TypeError, ValueError):
        return str(v).strip()


def parse_rate(v):
    """'잔존율 75% 이상' / '0.7' / 75 → 0.75. 못 읽으면 None."""
    if v is None:
        return None
    t = str(v)
    m = re.search(r'(\d{1,3})\s*%', t)
    if m:
        n = int(m.group(1))
        return n / 100 if 0 < n <= 100 else None
    try:
        f = float(t.strip())
    except ValueError:
        return None
    if 0 < f <= 1:
        return f
    if 1 < f <= 100:
        return f / 100
    return None


# ─────────────────────────── 내장 기준정보 ───────────────────────────
def load_customers():
    """거래처별 잔존율 (내장). 없으면 빈 DF."""
    if not CUST_FILE.exists():
        return pd.DataFrame(columns=['채널', '거래처코드', '거래처', '기준원문', '잔존율']), None
    d = json.loads(CUST_FILE.read_text(encoding='utf-8'))
    return pd.DataFrame(d.get('rows', [])), d.get('updated')


def save_customers(df, source=None):
    d = {'updated': _now_kst().strftime('%Y-%m-%d %H:%M'),
         'source': source or '화면 조정',
         'rows': df.to_dict('records')}
    CUST_FILE.write_text(json.dumps(d, ensure_ascii=False, indent=1), encoding='utf-8')
    return d['updated']


def load_items():
    """{코드: (품명, 소비기한월)} (내장)."""
    if not ITEM_FILE.exists():
        return {}, {}, None
    d = json.loads(ITEM_FILE.read_text(encoding='utf-8'))
    items = d.get('items', {})
    names = {k: v.get('n', '') for k, v in items.items()}
    months = {k: v['m'] for k, v in items.items() if v.get('m')}
    return names, months, d.get('updated')


def rebuild_items(file_bytes, source_name):
    """Item_*.xlsx 업로드 → 내장 기준정보 갱신. A=코드, B=품명, G=소비기한(월)."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    items = {}
    for r in range(2, ws.max_row + 1):
        k = code_key(ws.cell(r, 1).value)
        if not k:
            continue
        nm = ws.cell(r, 2).value
        mv = to_float(ws.cell(r, 7).value)
        items[k] = {'n': str(nm).strip() if nm else '', 'm': mv if mv > 0 else None}
    wb.close()
    d = {'updated': _now_kst().strftime('%Y-%m-%d %H:%M'), 'source': source_name, 'items': items}
    ITEM_FILE.write_text(json.dumps(d, ensure_ascii=False), encoding='utf-8')
    return len(items), d['updated']


def rebuild_customers(file_bytes, source_name):
    """통합 마스터 업로드 → 거래처 잔존율 재생성 (필요 시에만)."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = None
    for s in wb.sheetnames:
        if s.replace(' ', '') == '물류통합마스터전체':
            ws = wb[s]
            break
    if ws is None:
        ws = wb[wb.sheetnames[0]]
    rows = []
    for r in range(3, ws.max_row + 1):
        ch, code = ws.cell(r, 1).value, ws.cell(r, 2).value
        hq, std = ws.cell(r, 3).value, ws.cell(r, 15).value
        if all(v in (None, '') for v in (ch, code, hq, std)):
            continue
        rows.append({
            '채널': str(ch or '').strip(),
            '거래처코드': str(code).strip() if code not in (None, '') else '',
            '거래처': str(hq or '').strip().replace('\n', ' '),
            '기준원문': str(std).replace('\n', ' / ') if std is not None else '',
            '잔존율': parse_rate(std),
        })
    wb.close()
    df = pd.DataFrame(rows)
    up = save_customers(df, source_name)
    return len(df), up


# ─────────────────────────── 재고 (매일 업로드) ───────────────────────────
S_CODE, S_NAME, S_MFG, S_EXP, S_AVAIL_BOX, S_INV = 5, 6, 9, 10, 19, 1


@st.cache_data(show_spinner=False)
def load_stock(file_bytes, base_day: date, item_months: dict):
    """로트(품목×소비기한)별 출고가능 재고 + 잔존율.
    반환: (판정대상, 기한경과, 데이터오류, 결측행수)"""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    agg = defaultdict(lambda: {'qty': 0.0, 'mfg': None, 'name': '', 'invs': set()})
    skipped = 0
    for r in range(2, ws.max_row + 1):
        code = code_key(ws.cell(r, S_CODE).value)
        if not code:
            continue
        inv = ws.cell(r, S_INV).value                       # 기본 출고분 = IC930만
        if SHIP_INVENTORY and str(inv).strip() != SHIP_INVENTORY:
            continue
        qty = to_float(ws.cell(r, S_AVAIL_BOX).value)
        if qty <= 0:
            continue
        exp, mfg = to_date(ws.cell(r, S_EXP).value), to_date(ws.cell(r, S_MFG).value)
        if exp is None or mfg is None:
            skipped += 1
            continue
        a = agg[(code, exp)]
        a['qty'] += qty
        a['mfg'] = mfg if a['mfg'] is None else min(a['mfg'], mfg)
        if not a['name']:
            nm = ws.cell(r, S_NAME).value
            a['name'] = str(nm).strip() if nm else ''
        if inv:
            a['invs'].add(str(inv).strip())
    wb.close()

    rows, expired_rows, bad_rows = [], [], []
    for (code, exp), a in agg.items():
        left = (exp - base_day).days
        total = (exp - a['mfg']).days
        basis = '제조일'
        if total <= 0:                       # 제조일 오류 → 기준정보 소비기한(월)로 보정
            mon = item_months.get(code)
            if mon:
                total = int(round(mon * 30.44))
                basis = f'기준정보 {mon:g}개월'
            else:
                total = 0
        rec = {'품목코드': code, '품명': a['name'], '제조일': a['mfg'], '소비기한': exp,
               '총일수': total, '남은일수': left,
               '잔존율': (left / total) if total > 0 else None,
               '분모기준': basis, '출고가능(Box)': a['qty'],
               'Inventory': ','.join(sorted(a['invs']))}
        if exp < base_day:
            expired_rows.append(rec)
        elif rec['잔존율'] is None:
            bad_rows.append(rec)
        else:
            rows.append(rec)

    def _mk(rs):
        d = pd.DataFrame(rs)
        return d.sort_values(['품목코드', '소비기한']).reset_index(drop=True) if not d.empty else d

    return _mk(rows), _mk(expired_rows), _mk(bad_rows), skipped


# ─────────────────────────── 판정 ───────────────────────────
def judge(lots, need):
    if not lots:
        return None
    first = lots[0]
    if first['잔존율'] is not None and first['잔존율'] >= need:
        return {'판정': '현행유지', '권장소비기한': first['소비기한'],
                '권장잔존율': first['잔존율'], '권장가용': first['출고가능(Box)'],
                '사유': '최선(FIFO) 로트가 기준 충족'}
    ok = [l for l in lots if l['잔존율'] is not None and l['잔존율'] >= need]
    if ok:
        t = ok[0]
        fr = first['잔존율']
        why = (f"최선 로트 잔존율 {fr * 100:.1f}% < 기준" if fr is not None
               else '최선 로트 잔존율 산출 불가')
        return {'판정': '변경필요', '권장소비기한': t['소비기한'],
                '권장잔존율': t['잔존율'], '권장가용': t['출고가능(Box)'], '사유': why}
    return {'판정': '출고불가', '권장소비기한': None, '권장잔존율': None,
            '권장가용': 0.0, '사유': '기준 만족 로트 없음'}


@st.cache_data(show_spinner=False)
def build_result(cust_df, stock_df, item_names):
    """거래처 × 품목 전체 조합 판정. (50% 하프도달은 항상 제외)

    거래처를 바꿔가며 주문을 확인할 때 매번 재계산되지 않도록 캐시한다.
    """
    cust = cust_df[cust_df['잔존율'].notna()]
    cust = cust[cust['잔존율'] != HALF]

    lots_by_code = {c: g.sort_values('소비기한').to_dict('records')
                    for c, g in stock_df.groupby('품목코드', sort=False)}
    out = []
    for c in cust.itertuples(index=False):
        need = c.잔존율
        for code, lots in lots_by_code.items():
            j = judge(lots, need)
            if j is None:
                continue
            first = lots[0]
            out.append({
                '채널': c.채널, '거래처코드': c.거래처코드, '거래처': c.거래처,
                '기준잔존율(%)': round(need * 100, 1),
                '품목코드': code, '품명': item_names.get(code) or first['품명'],
                '판정': j['판정'],
                '현재최선_소비기한': first['소비기한'],
                '현재최선_잔존율(%)': (round(first['잔존율'] * 100, 1)
                                 if first.get('잔존율') is not None else None),
                '권장_소비기한': j['권장소비기한'],
                '권장_잔존율(%)': (round(j['권장잔존율'] * 100, 1)
                               if j.get('권장잔존율') is not None else None),
                '권장_출고가능(Box)': j['권장가용'],
                '보유로트수': len(lots), '사유': j['사유'],
            })
    return pd.DataFrame(out)


@st.cache_data(show_spinner=False)
def build_designation(result_df, stock_df):
    """로트(품목 × 소비기한)별 '지정 출고처' 목록 — 변경필요 건만 뒤집는다.

    기본 출고분(IC930 최빠른 유통기한)이 그 거래처 기준에 미달이라
    '더 나중 유통기한 로트로 바꿔 내보내야 하는'(변경필요) 거래처만 모은다.
    → "이 로트(더 나중 유통기한)는 이 거래처들로 내보내야 한다" 를 바로 읽을 수 있다.
    정상(현행유지)·출고불가는 제외한다.
    """
    if result_df.empty:
        return result_df
    df = result_df[result_df['판정'] == '변경필요']

    # 로트 (품목,소비기한) → 제조일
    mfg_map = {}
    for _, s in stock_df.iterrows():
        mfg_map[(s['품목코드'], s['소비기한'])] = s['제조일']

    rows = []
    for (code, exp), g in df.groupby(['품목코드', '권장_소비기한'], sort=False):
        cust_parts, collapsed = set(), []
        for ch, gch in g.groupby('채널', sort=True):
            if ch in COLLAPSE_CHANNELS:            # 대리점(GT)은 하나로 묶기
                collapsed.append(CHANNEL_NAMES.get(ch, ch))
            else:                                   # 나머지는 대표명으로 병합
                cust_parts.update(base_cust_name(n) for n in gch['거래처'])
        cust_parts = sorted(cust_parts)
        rows.append({
            '품목코드': code,
            '품명': g['품명'].iloc[0],
            '제조일자': mfg_map.get((code, exp)),
            '소비기한': exp,
            '★ 지정 출고처 ★': ', '.join(cust_parts + collapsed),
        })
    out = pd.DataFrame(rows)
    return out.sort_values(['품목코드', '소비기한']).reset_index(drop=True) if not out.empty else out


def to_excel(named):
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine='openpyxl') as w:
        for name, df in named:
            df.to_excel(w, sheet_name=name[:31], index=False)
    bio.seek(0)
    return bio


# ─────────────────────────── 주문 확인 (2단계) ───────────────────────────
# 제품별_리스트 열: E=품목코드, F=품목명, H=주문(BOX), Q=재고수량, A=출하Inv
O_CODE, O_NAME, O_BOX, O_QTY, O_INV = 5, 6, 8, 17, 1


@st.cache_data(show_spinner=False)
def load_orders(file_bytes):
    """제품별_리스트(주문) → 품목코드별 주문수량 집계."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    agg = defaultdict(lambda: {'box': 0.0, 'name': '', 'invs': set()})
    for r in range(2, ws.max_row + 1):
        code = code_key(ws.cell(r, O_CODE).value)
        if not code:
            continue
        box = to_float(ws.cell(r, O_BOX).value)
        a = agg[code]
        a['box'] += box
        if not a['name']:
            nm = ws.cell(r, O_NAME).value
            a['name'] = str(nm).strip() if nm else ''
        inv = ws.cell(r, O_INV).value
        if inv:
            a['invs'].add(str(inv).strip())
    wb.close()
    rows = [{'품목코드': c, '주문품명': a['name'], '주문(BOX)': a['box'],
             '출하Inv': ','.join(sorted(a['invs']))} for c, a in agg.items()]
    df = pd.DataFrame(rows)
    return df.sort_values('품목코드').reset_index(drop=True) if not df.empty else df


def check_orders(order_df, stock_df, need, item_names):
    """주문 품목만 그 거래처 기준(need)으로 판정."""
    lots_by = {c: g.sort_values('소비기한').to_dict('records')
               for c, g in stock_df.groupby('품목코드', sort=False)}
    out = []
    for _, o in order_df.iterrows():          # 한글 컬럼명은 itertuples에서 위치명이 되므로 iterrows 사용
        code = o['품목코드']
        ord_box = o['주문(BOX)']
        lots = lots_by.get(code)
        if not lots:
            out.append({'품목코드': code, '품명': item_names.get(code) or o['주문품명'],
                        '주문(BOX)': ord_box, '판정': '재고없음',
                        '현재최선_소비기한': None, '현재최선_잔존율(%)': None,
                        '권장_소비기한': None, '권장_잔존율(%)': None,
                        '권장_출고가능(Box)': 0.0, '충분여부': '-',
                        '사유': '출고가능 재고 없음(기한경과 등 제외됨)'})
            continue
        j = judge(lots, need)
        first = lots[0]
        avail = j['권장가용'] or 0.0
        enough = ('충분' if avail >= ord_box else f'부족 {ord_box - avail:,.0f}') \
            if j['판정'] != '출고불가' else '-'
        out.append({
            '품목코드': code, '품명': item_names.get(code) or o['주문품명'],
            '주문(BOX)': ord_box, '판정': j['판정'],
            '현재최선_소비기한': first['소비기한'],
            '현재최선_잔존율(%)': (round(first['잔존율'] * 100, 1)
                             if first.get('잔존율') is not None else None),
            '권장_소비기한': j['권장소비기한'],
            '권장_잔존율(%)': (round(j['권장잔존율'] * 100, 1)
                           if j.get('권장잔존율') is not None else None),
            '권장_출고가능(Box)': avail, '충분여부': enough, '사유': j['사유'],
        })
    return pd.DataFrame(out)


# ─────────────────────────── UI ───────────────────────────
st.title('📅 거래처별 잔존율')
st.caption(f'거래처별 잔존율 기준으로 “어느 소비기한 재고로 출고해야 하는지” 판정합니다. '
           f'**매일 로케이션별 재고조회만 올리면 됩니다.** '
           f'(기본 출고분 = **{SHIP_INVENTORY}** 인벤토리 재고 기준)')

cust_df, cust_up = load_customers()
item_names, item_months, item_up = load_items()

with st.sidebar:
    st.header('⚙️ 설정')
    base_day = st.date_input('기준일 (잔존율 계산 기준)', value=_today_kst())
    st.caption('잔존율 = (소비기한 − 기준일) ÷ (소비기한 − 제조일)')
    st.caption('※ 50%(하프도달)는 전 거래처 공통이라 판정에서 항상 제외됩니다.')
    st.divider()
    st.markdown('**📚 기준정보 (내장)**')
    st.caption(f"거래처 잔존율: {len(cust_df):,}곳 · 갱신 {cust_up or '-'}")
    st.caption(f"품목 기준정보: {len(item_names):,}품목 · 갱신 {item_up or '-'}")

    with st.expander('🔄 기준정보 갱신 (가끔)'):
        st.caption('Item 기준정보는 월 1회 정도만 갱신하면 됩니다.')
        if st.button('🗂️ 공용 허브에서 갱신', key='janjon_hub', width='stretch'):
            try:
                _hb = hub.item_raw_session()
                if not _hb:
                    st.warning('공용 허브에 Item 기준정보가 없습니다.')
                else:
                    n, up = rebuild_items(_hb, '공용 허브')
                    st.success(f'✅ 허브에서 갱신: {n:,}품목 ({up})')
                    st.cache_data.clear()
            except Exception as e:
                st.error(f'실패: {e}')
        st.caption('품목 기준정보(품명·소비기한월)는 공용 기준정보 관리에서 올린 걸 위 버튼으로 반영합니다.')
        st.divider()
        st.caption('통합 마스터는 고정입니다. 원본이 바뀐 경우에만 올리세요.')
        up_cust = st.file_uploader('통합 마스터.xlsx 업로드', type=['xlsx'], key='up_cust')
        if up_cust is not None:
            try:
                n, up = rebuild_customers(up_cust.getvalue(), up_cust.name)
                st.success(f'✅ 거래처 기준 갱신: {n:,}곳 ({up}) — 새로고침하세요')
            except Exception as e:
                st.error(f'실패: {e}')

if cust_df.empty:
    st.error('거래처 기준정보가 없습니다. 사이드바 → 기준정보 갱신에서 통합 마스터를 올려주세요.')
    st.stop()

# ── 거래처 잔존율 조정 ──
with st.expander('🎛 거래처 잔존율 조정 (직접 수정 · 저장하면 다음부터 계속 적용)'):
    st.caption('잔존율 칸을 고치면 그 거래처 기준이 바뀝니다. 비우면 판정 대상에서 빠집니다. '
               '(0.75 = 75%)')
    edit_src = cust_df[['채널', '거래처코드', '거래처', '잔존율', '기준원문']].copy()
    edited = st.data_editor(
        edit_src, width='stretch', hide_index=True, num_rows='fixed',
        column_config={
            '잔존율': st.column_config.NumberColumn(
                '잔존율(0~1)', min_value=0.0, max_value=1.0, step=0.05, format='%.2f'),
            '기준원문': st.column_config.TextColumn('원문(참고)', disabled=True),
            '채널': st.column_config.TextColumn(disabled=True),
            '거래처코드': st.column_config.TextColumn(disabled=True),
            '거래처': st.column_config.TextColumn(disabled=True),
        },
        key='cust_editor',
    )
    if st.button('💾 잔존율 저장', width='stretch'):
        merged = cust_df.copy()
        merged['잔존율'] = edited['잔존율'].values
        up = save_customers(merged)
        st.success(f'저장 완료 ({up}) — 아래 결과에 즉시 반영됩니다.')
        cust_df = merged

# ── 매일: 재고조회 업로드 ──
st.subheader('① 오늘의 로케이션별 재고조회 업로드')
f_stock = st.file_uploader('로케이션별 재고조회 .xlsx', type=['xlsx'], key='u_stock',
                           label_visibility='collapsed')
if not f_stock:
    st.info('로케이션별 재고조회 파일을 올리면 바로 판정합니다.')
    st.stop()

with st.spinner('재고 읽는 중…'):
    stock_df, expired_df, bad_df, skipped = load_stock(
        f_stock.getvalue(), base_day, item_months)

if stock_df.empty:
    st.error('출고가능 재고가 없습니다. 파일을 확인하세요.')
    st.stop()

tgt = cust_df[(cust_df['잔존율'].notna()) & (cust_df['잔존율'] != HALF)]
m1, m2, m3, m4 = st.columns(4)
m1.metric('거래처(전체)', f'{len(cust_df):,}')
m2.metric('판정 대상 거래처', f'{len(tgt):,}')
m3.metric('재고 품목', f"{stock_df['품목코드'].nunique():,}")
m4.metric('재고 로트', f'{len(stock_df):,}')

_notes = []
if skipped:
    _notes.append(f'제조일·소비기한 없음 {skipped}건')
if not expired_df.empty:
    _notes.append(f'소비기한 경과 {len(expired_df)}로트(제외)')
if not bad_df.empty:
    _notes.append(f'잔존율 산출불가 {len(bad_df)}로트(제외)')
if _notes:
    st.caption('⚠️ 제외 내역 — ' + ' · '.join(_notes) + ' (엑셀에 별도 시트로 포함)')

with st.spinner(f"판정 중… (거래처 {len(tgt):,} × 품목 {stock_df['품목코드'].nunique():,})"):
    res = build_result(cust_df, stock_df, item_names)

if res.empty:
    st.warning('판정 결과가 없습니다. (잔존율 기준이 있는 거래처가 없는지 확인하세요)')
    st.stop()

st.success(f'✅ 판정 완료 — 총 {len(res):,}행')
cnt = res['판정'].value_counts()
k1, k2, k3 = st.columns(3)
k1.metric('현행유지', f"{int(cnt.get('현행유지', 0)):,}")
k2.metric('변경필요', f"{int(cnt.get('변경필요', 0)):,}")
k3.metric('출고불가', f"{int(cnt.get('출고불가', 0)):,}")

st.divider()
st.subheader('② 지정 출고 대상 (기준 미달 → 유통기한 변경)')

_desig_dl = build_designation(res, stock_df)
q = st.text_input('검색 (품목코드 / 품명 / 거래처)', key='q_desig')
dview = _desig_dl
if q:
    s = q.strip().lower()
    dview = dview[dview['품목코드'].str.lower().str.contains(s, na=False)
                  | dview['품명'].str.lower().str.contains(s, na=False)
                  | dview['★ 지정 출고처 ★'].str.lower().str.contains(s, na=False)]
st.caption(f'표시 {len(dview):,}로트 / 전체 {len(_desig_dl):,}로트 · '
           '**정상출고(FIFO로 그냥 나가도 되는 품목)는 제외**하고, 잔존율 기준 미달로 '
           '**이 로트(더 나중 유통기한)로 바꿔 내보내야 하는 거래처**만 표시합니다. '
           '대리점(GT)은 **“대리점”** 하나로 묶고, 나머지는 거래처명 그대로 표기합니다.')
st.dataframe(
    dview.head(2000), width='stretch', hide_index=True,
    column_config={
        '제조일자': st.column_config.DateColumn('제조일자', format='YYYY-MM-DD'),
        '소비기한': st.column_config.DateColumn('소비기한', format='YYYY-MM-DD'),
        '★ 지정 출고처 ★': st.column_config.TextColumn(width='large'),
    })
if len(dview) > 2000:
    st.caption('※ 화면에는 상위 2,000로트만 표시됩니다. 전체는 아래에서 다운로드하세요.')
view = dview

st.divider()
d1, d2 = st.columns(2)
_sheets = [('변경필요', res[res['판정'] == '변경필요']),
           ('출고불가', res[res['판정'] == '출고불가']),
           ('로트별_지정출고처', _desig_dl), ('판정결과(상세)', res),
           ('거래처기준', cust_df), ('재고로트', stock_df)]
if not expired_df.empty:
    _sheets.append(('기한경과(제외)', expired_df))
if not bad_df.empty:
    _sheets.append(('데이터오류(제외)', bad_df))
with d1:
    st.download_button('⬇ 전체 결과 다운로드 (엑셀)', data=to_excel(_sheets),
                       file_name=f'잔존율_출고판정_{base_day:%Y%m%d}.xlsx',
                       mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                       width='stretch')
with d2:
    st.download_button('⬇ 화면 표시분만 다운로드', data=to_excel([('필터결과', view)]),
                       file_name=f'잔존율_출고판정_필터_{base_day:%Y%m%d}.xlsx',
                       mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                       width='stretch')

# ─────────────────────────── ③ 주문 확인 (채널/거래처) ───────────────────────────
st.divider()
st.subheader('③ 주문 확인 (채널·거래처)')
st.caption('기준이 같은 거래처는 **채널로 묶어 한 번에** 확인하고, 채널 안에서 기준이 다르면 '
           '**기준 그룹별로** 나눠서 확인합니다. 주문 파일(제품별_리스트) 하나로 반복 사용할 수 있습니다.')


def _safe(s):
    return re.sub(r'[\\/:*?"<>|]', '_', str(s)).strip()[:40]


mode = st.radio('선택 단위', ['🗂 채널 카테고리', '🏢 개별 거래처'],
                horizontal=True, key='ord_mode')

need_sel = None       # 판정 기준(잔존율)
scope_title = ''      # 화면 표기
scope_file = ''       # 파일명
scope_members = None  # 이 그룹에 해당하는 거래처 목록(DataFrame)

if mode.startswith('🗂'):
    # ── 채널 단위: 기준이 통일되면 채널 하나로, 갈리면 기준 그룹으로 ──
    ch_info = (tgt.groupby('채널')['잔존율']
               .agg(cnt='size', kinds='nunique').reset_index()
               .sort_values('채널'))
    ch_labels, ch_map = [], {}
    for r in ch_info.itertuples(index=False):
        if r.kinds == 1:
            lab = f'{ch_disp(r.채널)} · {r.cnt}곳 · 기준 통일'
        else:
            lab = f'{ch_disp(r.채널)} · {r.cnt}곳 · 기준 {r.kinds}종 ⚠ (그룹 선택)'
        ch_labels.append(lab)
        ch_map[lab] = r.채널
    pick_ch_lab = st.selectbox('채널 선택', ch_labels,
                               index=0 if ch_labels else None, key='ord_ch')
    if pick_ch_lab:
        ch = ch_map[pick_ch_lab]
        sub = tgt[tgt['채널'] == ch].sort_values('거래처')
        rates = sorted(sub['잔존율'].unique())
        if len(rates) == 1:
            need_sel = rates[0]
            scope_members = sub
            scope_title = f"채널 **{ch_disp(ch)}** 전체 · **{len(sub)}곳** · 기준 잔존율 **{need_sel*100:.0f}%** (통일)"
            scope_file = f'{ch}_{need_sel*100:.0f}p'
        else:
            st.warning(f'⚠️ 채널 **{ch_disp(ch)}** 는 기준이 {len(rates)}종입니다 '
                       f"({', '.join(f'{x*100:.0f}%' for x in rates)}). 기준 그룹을 골라 개별로 확인하세요.")
            grp_labels, grp_map = [], {}
            for rt in rates:
                mem = sub[sub['잔존율'] == rt]
                gl = f'{rt*100:.0f}% · {len(mem)}곳'
                grp_labels.append(gl)
                grp_map[gl] = (rt, mem)
            pick_grp = st.selectbox('기준 그룹 선택', grp_labels, key='ord_grp')
            need_sel, scope_members = grp_map[pick_grp]
            scope_title = (f"채널 **{ch_disp(ch)}** · 기준 **{need_sel*100:.0f}%** 그룹 · "
                           f"**{len(scope_members)}곳**")
            scope_file = f'{ch}_{need_sel*100:.0f}p'
else:
    # ── 개별 거래처 ──
    tgt_sorted = tgt.sort_values(['채널', '거래처'])
    labels = [f"[{r.채널}] {r.거래처} — {r.잔존율 * 100:.0f}%"
              for r in tgt_sorted.itertuples(index=False)]
    lab2row = {lab: r for lab, r in zip(labels, tgt_sorted.itertuples(index=False))}
    pick = st.selectbox('거래처 선택', labels, index=0 if labels else None, key='ord_cust')
    if pick:
        sel = lab2row[pick]
        need_sel = sel.잔존율
        scope_title = f"**{sel.거래처}** · 채널 {ch_disp(sel.채널)} · 기준 잔존율 **{need_sel*100:.0f}%**"
        scope_file = _safe(sel.거래처)
        scope_members = tgt[(tgt['채널'] == sel.채널) & (tgt['거래처'] == sel.거래처)]

f_ord = st.file_uploader('주문 파일 (제품별_리스트 .xlsx)', type=['xlsx'], key='u_ord')

if need_sel is None or not f_ord:
    st.info('채널(또는 거래처)을 고르고 주문 파일을 올리면 주문 품목만 판정합니다.')
    st.stop()

order_df = load_orders(f_ord.getvalue())
if order_df.empty:
    st.error('주문 데이터를 읽지 못했습니다. (E=품목코드, H=주문(BOX) 형식인지 확인)')
    st.stop()

ores = check_orders(order_df, stock_df, need_sel, item_names)

st.markdown('▶ ' + scope_title)
if scope_members is not None and len(scope_members) > 1:
    with st.expander(f'이 기준으로 함께 확인되는 거래처 {len(scope_members)}곳 보기'):
        st.dataframe(scope_members[['채널', '거래처코드', '거래처', '잔존율']],
                     width='stretch', hide_index=True)
oc = ores['판정'].value_counts()
o1, o2, o3, o4 = st.columns(4)
o1.metric('주문 품목', f'{len(ores):,}')
o2.metric('현행유지', f"{int(oc.get('현행유지', 0)):,}")
o3.metric('변경필요', f"{int(oc.get('변경필요', 0)):,}")
o4.metric('출고불가+재고없음',
          f"{int(oc.get('출고불가', 0)) + int(oc.get('재고없음', 0)):,}")

need_act = ores[ores['판정'].isin(['변경필요', '출고불가', '재고없음'])]
if need_act.empty:
    st.success('🎉 모든 주문 품목이 현행 그대로 출고 가능합니다.')
else:
    st.warning(f'⚠️ 조치 필요 {len(need_act):,}건 — 변경필요 '
               f"{int(oc.get('변경필요', 0))} · 출고불가 {int(oc.get('출고불가', 0))} · "
               f"재고없음 {int(oc.get('재고없음', 0))}")

only_act = st.checkbox('조치 필요한 것만 보기', value=True, key='ord_only')
st.dataframe(need_act if only_act else ores, width='stretch', hide_index=True)

st.download_button(
    '⬇ 이 주문 확인 결과 (엑셀)',
    data=to_excel([('주문확인', ores), ('조치필요', need_act),
                   ('대상거래처', scope_members if scope_members is not None else cust_df.head(0))]),
    file_name=f"주문확인_{scope_file}_{base_day:%Y%m%d}.xlsx",
    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    width='stretch')
