"""
입고 파레트 구분기 (Streamlit 웹앱)

규칙:
1. 한 파레트 내 Plt_1차 합 ≤ 1 (절대 기준)
2. 5박스 이내 품목: 한 파레트에 최대 10품목
3. 6~24박스 품목: 한 파레트에 최대 3품목
4. 25박스 이상 품목: 단독 파레트 (Plt_1차 ≥ 1이면 분할)
"""

import io
import json
import shutil
from copy import copy as _copy
from datetime import datetime, date, time as dtime
from pathlib import Path

import openpyxl
import pandas as pd
import streamlit as st
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


APP_DIR = Path(__file__).parent
TEMPLATE_DIR = APP_DIR / 'templates'
DATA_DIR = APP_DIR / 'data'
DATA_DIR.mkdir(exist_ok=True)
TEMPLATE_DIR.mkdir(exist_ok=True)
FCJ_TEMPLATE = TEMPLATE_DIR / 'FCJ.xlsx'
MASTER_FILE = DATA_DIR / 'master_barcodes.json'


# ---------- 핵심 로직 ----------
DEFAULTS = {
    'small_max': 5,       # 이 박스 수 이하 → 소형 (규칙 2)
    'medium_max': 24,     # small_max 초과 ~ 이 값 이하 → 중형 (규칙 3)
    'small_cap': 10,      # 소형 파레트당 최대 품목 수
    'medium_cap': 3,      # 중형 파레트당 최대 품목 수
    'plt_sum_max': 1.0,   # Plt_1차 합 상한 (절대 기준)
}


def categorize(items, cfg):
    rule4 = [i for i in items if i['box'] > cfg['medium_max']]
    rule3 = [i for i in items if cfg['small_max'] < i['box'] <= cfg['medium_max']]
    rule2 = [i for i in items if i['box'] <= cfg['small_max']]
    return rule4, rule3, rule2


def pack_first_fit(items_list, max_count, max_sum):
    """박스 수 오름차순 First-Fit 빈패킹.
    작은 박스끼리 먼저 한 파레트에 채워, 박스 수 큰 품목이 마지막 빈(단독에 가까움)으로 남도록.
    """
    sorted_items = sorted(items_list, key=lambda x: (x['box'], -x['plt1']))
    bins = []
    for it in sorted_items:
        placed = False
        for b in bins:
            if len(b['items']) < max_count and b['sum'] + it['plt1'] <= max_sum + 1e-9:
                b['items'].append(it)
                b['sum'] += it['plt1']
                placed = True
                break
        if not placed:
            bins.append({'sum': it['plt1'], 'items': [it]})
    return bins


def build_pallets(items, cfg):
    rule4, rule3, rule2 = categorize(items, cfg)
    plt_max = cfg['plt_sum_max']
    pallets = []
    plt_no = 0

    # 규칙 4: 단독 파레트 (Plt_1차 ≥ plt_max이면 분할). 박스 수 많은 순으로 처리.
    for it in sorted(rule4, key=lambda x: (-x['box'], -x['plt1'])):
        plt = it['plt1']
        bpp = it['plt_conv']
        if plt >= plt_max - 1e-9 and bpp:
            full_int = int(plt / plt_max + 1e-9)
            remainder_plt = plt - full_int * plt_max
            remaining_box = it['box']
            box_per_pallet = bpp * plt_max  # 한 파레트에 들어가는 박스 수
            for _ in range(full_int):
                plt_no += 1
                pallets.append({
                    'plt_no': plt_no,
                    'rule': '4',
                    'items': [{**it,
                               'box': int(box_per_pallet),
                               'qty': int(box_per_pallet) * (it['ipsu'] or 0),
                               'plt1': plt_max}],
                })
                remaining_box -= int(box_per_pallet)
            if remainder_plt > 1e-9 and remaining_box > 0:
                plt_no += 1
                pallets.append({
                    'plt_no': plt_no,
                    'rule': '4',
                    'items': [{**it, 'box': remaining_box,
                               'qty': remaining_box * (it['ipsu'] or 0),
                               'plt1': remaining_box / bpp}],
                })
        else:
            plt_no += 1
            pallets.append({'plt_no': plt_no, 'rule': '4', 'items': [it]})

    # 규칙 3: 중형, max_count=medium_cap
    for b in pack_first_fit(rule3, cfg['medium_cap'], plt_max):
        plt_no += 1
        pallets.append({'plt_no': plt_no, 'rule': '3', 'items': b['items']})

    # 규칙 2: 소형, max_count=small_cap
    for b in pack_first_fit(rule2, cfg['small_cap'], plt_max):
        plt_no += 1
        pallets.append({'plt_no': plt_no, 'rule': '2', 'items': b['items']})

    return pallets


# ---------- 입력 파싱 ----------
def parse_workbook(uploaded_file):
    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    return wb


def find_header_row(ws, max_search=10):
    for r in range(1, min(ws.max_row + 1, max_search + 1)):
        for c in range(1, min(ws.max_column + 1, 5)):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip().lower() in ('item code', 'itemcode'):
                return r
    return None


def extract_title(ws):
    """1행 타이틀에서 날짜 추출 (예: '2026년 4월 28일 입고예정정보')"""
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v is None:
            continue
        if isinstance(v, datetime):
            return v.date(), str(v.date())
        if isinstance(v, date):
            return v, str(v)
        if isinstance(v, str) and '입고' in v:
            return None, v
    return None, ''


def extract_items(ws, header_row):
    items = []
    for r in range(header_row + 1, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, 12)]
        if row[0] is None:
            continue
        try:
            items.append({
                'item_code': row[0],
                'name': row[1],
                'ipsu': row[2],
                'box': int(row[3]) if row[3] is not None else 0,
                'qty': row[4],
                'expiry': row[5],
                'plt_conv': row[6],
                'plt1': float(row[7]) if row[7] is not None else 0.0,
            })
        except (TypeError, ValueError):
            continue
    return items


# ---------- 출력 엑셀 작성 ----------
HEADERS = ['Item code', 'Item', '입수', '박스', '낱개', '소비기한',
           'plt환산', 'Plt_1차', 'PLT 번호', '비 고', '순번']


def build_excel(pallets, title_date, totals):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '입고정보'

    thin = Side(border_style='thin', color='888888')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(bold=True)
    header_fill = PatternFill('solid', fgColor='D9E1F2')

    # 1행: 타이틀
    ws.cell(1, 1, title_date if title_date else '')
    ws.cell(1, 2, ' 입고예정정보')
    ws.cell(1, 4, totals['box'])
    ws.cell(1, 5, totals['qty'])
    ws.cell(1, 8, totals['plt'])
    for c in range(1, 12):
        ws.cell(1, c).font = Font(bold=True, size=12)

    # 2행: 헤더
    for i, h in enumerate(HEADERS, 1):
        cell = ws.cell(2, i, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # 3행~: 데이터
    r = 3
    for p in pallets:
        for it in p['items']:
            ws.cell(r, 1, it['item_code'])
            ws.cell(r, 2, it['name'])
            ws.cell(r, 3, it['ipsu'])
            ws.cell(r, 4, it['box'])
            ws.cell(r, 5, it['qty'])
            ws.cell(r, 6, it.get('expiry'))
            ws.cell(r, 7, it['plt_conv'])
            ws.cell(r, 8, it['plt1'])
            ws.cell(r, 9, p['plt_no'])
            for c in range(1, 12):
                ws.cell(r, c).border = border
            r += 1

    # 컬럼 너비
    widths = [12, 42, 6, 7, 8, 11, 9, 11, 9, 14, 7]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------- 바코드 마스터 관리 ----------
def load_master():
    if MASTER_FILE.exists():
        try:
            return json.loads(MASTER_FILE.read_text(encoding='utf-8'))
        except Exception:
            pass
    return {'mapping': {}, 'updated_at': None, 'count': 0, 'source_name': None}


def save_master(mapping, source_name):
    data = {
        'mapping': mapping,
        'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'count': len(mapping),
        'source_name': source_name,
    }
    MASTER_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding='utf-8')
    return data


def parse_master_excel(uploaded_file):
    """업로드한 엑셀에서 품목코드↔바코드 매핑 추출.
    헤더 행을 자동 감지하여 코드/바코드 컬럼을 찾는다.
    """
    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    ws = wb.active

    code_keywords = {'item code', 'itemcode', 'item_code', '품목코드', '제품코드', 'sku', 'code', '코드'}
    bc_keywords = {'barcode', '바코드', 'gtin', 'ean', 'jan'}

    header_row = None
    code_col = None
    bc_col = None

    for r in range(1, min(ws.max_row + 1, 10)):
        cands_code = None
        cands_bc = None
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            s = str(v).strip().lower()
            if cands_code is None and any(k in s for k in code_keywords):
                cands_code = c
            if cands_bc is None and any(k in s for k in bc_keywords):
                cands_bc = c
        if cands_code is not None and cands_bc is not None:
            header_row = r
            code_col = cands_code
            bc_col = cands_bc
            break

    if header_row is None:
        raise ValueError("'품목코드'와 '바코드' 컬럼을 자동으로 찾지 못했습니다. 헤더에 두 컬럼이 모두 있는지 확인해주세요.")

    mapping = {}
    for r in range(header_row + 1, ws.max_row + 1):
        code = ws.cell(r, code_col).value
        bc = ws.cell(r, bc_col).value
        if code is None or bc is None:
            continue
        # 키 정규화: 정수면 정수문자열로, 아니면 그대로
        try:
            code_key = str(int(code))
        except (TypeError, ValueError):
            code_key = str(code).strip()
        bc_str = str(bc).strip()
        if not bc_str:
            continue
        # 엑셀의 바코드 앞 작은따옴표 제거
        if bc_str.startswith("'"):
            bc_str = bc_str[1:]
        mapping[code_key] = bc_str

    return mapping, header_row, code_col, bc_col


def _master_key(item_code):
    if item_code is None or item_code == '':
        return None
    try:
        return str(int(item_code))
    except (TypeError, ValueError):
        return str(item_code).strip()


def lookup_barcode(item_code, master):
    if not item_code:
        return ''
    key = _master_key(item_code)
    return master.get('mapping', {}).get(key, '') if key else ''


# ---------- 유통기한 마스터 관리 ----------
def normalize_expiry(v):
    """유통기한을 YYYYMMDD 정수로 정규화. 실패 시 None."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return int(v.strftime('%Y%m%d'))
    if isinstance(v, date):
        return int(v.strftime('%Y%m%d'))
    if isinstance(v, (int, float)):
        n = int(v)
        if 19000101 <= n <= 21001231:
            return n
        return None
    if isinstance(v, str):
        s = v.strip().replace('-', '').replace('/', '').replace('.', '').replace(' ', '')
        try:
            n = int(s)
            if 19000101 <= n <= 21001231:
                return n
        except ValueError:
            pass
    return None


EMPTY_EXPIRY_MASTER = {
    'mapping': {},        # {code: max_expiry}  — 인벤토리 무관 최댓값(호환용)
    'by_inv': {},         # {code: {inventory: max_expiry}} — 인벤토리별
    'inventories': [],    # 등장한 모든 인벤토리 (정렬됨)
    'duplicates': {},     # {code: [(inv, expiry), ...]} 늦은 순
    'updated_at': None,
    'count': 0,
    'source_name': None,
}

NO_INVENTORY = '(미지정)'


def parse_expiry_master(uploaded_file):
    """업로드 엑셀에서 품목코드↔유통기한 매핑 추출.
    Inventory 컬럼이 있으면 (코드, 인벤토리)별로 매핑.
    같은 (코드, 인벤토리)에 여러 행이면 가장 늦은 유통기한 사용.
    """
    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    ws = wb.active

    code_keywords = {'item code', 'itemcode', 'item_code', '품목코드', '제품코드', 'sku', 'code', '코드'}
    exp_keywords = {'유통기한', '소비기한', '유효기한', 'expiry', 'expiration', 'best by', 'use by', 'exp date'}
    inv_keywords = {'inventory', '인벤토리'}

    header_row = None
    code_col = None
    exp_col = None
    inv_col = None  # 옵셔널

    for r in range(1, min(ws.max_row + 1, 10)):
        cands_code = None
        cands_exp = None
        cands_inv = None
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            s = str(v).strip().lower()
            if cands_code is None and any(k in s for k in code_keywords):
                cands_code = c
            if cands_exp is None and any(k in s for k in exp_keywords):
                cands_exp = c
            if cands_inv is None and any(k in s for k in inv_keywords):
                cands_inv = c
        if cands_code is not None and cands_exp is not None:
            header_row = r
            code_col = cands_code
            exp_col = cands_exp
            inv_col = cands_inv  # None일 수 있음
            break

    if header_row is None:
        raise ValueError("'품목코드'와 '유통기한' 컬럼을 자동으로 찾지 못했습니다. 헤더에 두 컬럼이 모두 있는지 확인해주세요.")

    by_inv = {}        # {code: {inv: max_expiry}}
    flat_max = {}      # {code: max_expiry}
    inv_set = set()
    skipped = 0

    for r in range(header_row + 1, ws.max_row + 1):
        code = ws.cell(r, code_col).value
        exp = ws.cell(r, exp_col).value
        if code is None or exp is None:
            continue
        code_key = _master_key(code)
        exp_int = normalize_expiry(exp)
        if exp_int is None or code_key is None:
            skipped += 1
            continue

        if inv_col is not None:
            inv_v = ws.cell(r, inv_col).value
            inv_key = str(inv_v).strip() if inv_v not in (None, '') else NO_INVENTORY
        else:
            inv_key = NO_INVENTORY
        inv_set.add(inv_key)

        cd = by_inv.setdefault(code_key, {})
        if inv_key not in cd or exp_int > cd[inv_key]:
            cd[inv_key] = exp_int

        if code_key not in flat_max or exp_int > flat_max[code_key]:
            flat_max[code_key] = exp_int

    # 한 코드에 인벤토리가 여러 개면 duplicates에 정리 (늦은 순)
    duplicates = {}
    for code, inv_dict in by_inv.items():
        if len(inv_dict) > 1:
            duplicates[code] = sorted(inv_dict.items(), key=lambda x: -x[1])

    inventories = sorted(inv_set)
    return flat_max, by_inv, inventories, duplicates, header_row, code_col, exp_col, inv_col, skipped


def lookup_expiry_master(item_code, expiry_master, inventory=None):
    """마스터에서 유통기한 룩업.
    inventory가 지정되면 해당 인벤토리 값을, 아니면 모든 인벤토리 중 가장 늦은 값.
    """
    key = _master_key(item_code)
    if key is None:
        return None
    if inventory is not None and inventory != '':
        return expiry_master.get('by_inv', {}).get(key, {}).get(inventory)
    return expiry_master.get('mapping', {}).get(key)


DEFAULT_INVENTORY = 'IC930'
ALLOWED_INVENTORIES = ['IC930', 'IC920', 'IC100']  # 선택 가능 풀 + 자동 선택 우선순위


def default_inventory_for(item_code, expiry_master, priority=ALLOWED_INVENTORIES):
    """기본 인벤토리 자동 선택. ALLOWED_INVENTORIES 순서대로 보유 여부 확인,
    첫 매치 반환. 셋 다 보유 안 하면 None.
    """
    key = _master_key(item_code)
    if key is None:
        return None
    inv_dict = expiry_master.get('by_inv', {}).get(key, {})
    if not inv_dict:
        return None
    for inv in priority:
        if inv in inv_dict:
            return inv
    return None


def selectable_inventories_for(item_code, expiry_master, allowed=ALLOWED_INVENTORIES):
    """ALLOWED_INVENTORIES 중에서 해당 품목이 실제로 보유한 인벤토리 리스트."""
    key = _master_key(item_code)
    if key is None:
        return []
    inv_dict = expiry_master.get('by_inv', {}).get(key, {})
    return [inv for inv in allowed if inv in inv_dict]


def held_inventories_for(item_code, expiry_master):
    """해당 품목이 보유 중인 인벤토리 목록 (늦은 유통기한 순)."""
    key = _master_key(item_code)
    if key is None:
        return []
    inv_dict = expiry_master.get('by_inv', {}).get(key, {})
    return [inv for inv, _ in sorted(inv_dict.items(), key=lambda x: -x[1])]


# ---------- FCJ 양식 출력 (파레트별 시트) ----------
SHIPPER_NAME = '샘표식품'
DEFAULT_TIME = dtime(9, 0)


DATA_ROW_HEIGHT_MULTIPLIER = 2.5  # 양식 14행 높이의 N배로 데이터 행 높이를 키움


def _capture_row_style(ws, row, ncols=12, height_multiplier=DATA_ROW_HEIGHT_MULTIPLIER):
    """원본 양식의 데이터 첫 행 스타일을 캡처. 행 높이는 multiplier만큼 곱함."""
    template = []
    for c in range(1, ncols + 1):
        src = ws.cell(row, c)
        template.append({
            'font': _copy(src.font),
            'fill': _copy(src.fill),
            'border': _copy(src.border),
            'alignment': _copy(src.alignment),
            'number_format': src.number_format,
        })
    base_h = ws.row_dimensions[row].height if row in ws.row_dimensions else None
    height = base_h * height_multiplier if base_h else None
    return template, height


def _apply_row_style(ws, row, template, height, ncols=12):
    for c in range(1, ncols + 1):
        cell = ws.cell(row, c)
        t = template[c - 1]
        cell.font = _copy(t['font'])
        cell.fill = _copy(t['fill'])
        cell.border = _copy(t['border'])
        cell.alignment = _copy(t['alignment'])
        cell.number_format = t['number_format']
    if height is not None:
        ws.row_dimensions[row].height = height


_BLANK_BORDER = Border()


def _copy_row_block(src_ws, src_start, src_end, dst_ws, dst_start, ncols=12):
    """src_ws[src_start..src_end]의 값/스타일/행 높이를 dst_ws[dst_start..]에 복사."""
    for i in range(src_end - src_start + 1):
        src_r = src_start + i
        dst_r = dst_start + i
        for c in range(1, ncols + 1):
            sc = src_ws.cell(src_r, c)
            dc = dst_ws.cell(dst_r, c)
            dc.value = sc.value
            if sc.has_style:
                dc.font = _copy(sc.font)
                dc.fill = _copy(sc.fill)
                dc.border = _copy(sc.border)
                dc.alignment = _copy(sc.alignment)
                dc.number_format = sc.number_format
        if src_r in src_ws.row_dimensions:
            dst_ws.row_dimensions[dst_r].height = src_ws.row_dimensions[src_r].height


def _copy_merges_in_range(src_ws, dst_ws, src_start, src_end, row_offset):
    """src_ws의 src_start~src_end 행에 속한 병합 영역을 row_offset만큼 옮겨 dst_ws에 적용."""
    for mr in list(src_ws.merged_cells.ranges):
        if src_start <= mr.min_row and mr.max_row <= src_end:
            dst_ws.merge_cells(
                start_row=mr.min_row + row_offset, start_column=mr.min_col,
                end_row=mr.max_row + row_offset, end_column=mr.max_col,
            )


def _copy_column_widths(src_ws, dst_ws):
    for col_letter, dim in src_ws.column_dimensions.items():
        if dim.width:
            dst_ws.column_dimensions[col_letter].width = dim.width


def _setup_print_layout(ws, last_data_row, ncols=12, fit_to_height=1):
    """A4 가로, 가로 1페이지에 맞춤. fit_to_height=1이면 세로도 1페이지에 압축,
    0이면 자동(수동 페이지 나누기/row_breaks 우선)."""
    ws.page_setup.paperSize = ws.PAPERSIZE_A4    # A4
    ws.page_setup.orientation = 'landscape'       # 가로
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = fit_to_height
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.page_margins.header = 0.2
    ws.page_margins.footer = 0.2

    ws.print_options.horizontalCentered = True

    last_col_letter = openpyxl.utils.get_column_letter(ncols)
    ws.print_area = f'A1:{last_col_letter}{last_data_row}'


def _clear_data_area(ws, start_row=14, end_row=None, ncols=12):
    """데이터 영역의 값과 테두리를 깨끗이 비움 + 행 높이 초기화."""
    if end_row is None:
        end_row = max(ws.max_row, 30)
    for r in range(start_row, end_row + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(r, c)
            cell.value = None
            cell.border = _BLANK_BORDER
        if r in ws.row_dimensions:
            ws.row_dimensions[r].height = None


def fill_fcj_sheet(ws, plt_items, header_info, master, expiry_map,
                   plt_no, total_plts, style_template, template_height):
    # 상단 정보 영역
    ws.cell(6, 3).value = SHIPPER_NAME              # 화주명 (고정)
    ws.cell(6, 10).value = header_info['date']      # 입고예정일자
    ws.cell(7, 10).value = DEFAULT_TIME             # 입고예정시간 (고정 09:00)
    ws.cell(8, 3).value = header_info['po_no']      # PO NO
    ws.cell(8, 10).value = header_info['vehicle']   # 차량번호

    # 비고 자리(우측 상단 비어있는 K8:L8 병합)에 파레트 표기
    ws.cell(8, 11).value = f'PLT {plt_no} / {total_plts}'

    # 1) 데이터 영역(14행~) 깨끗이 비움 (값 + 테두리 + 행 높이)
    _clear_data_area(ws, start_row=14)

    # 2) 품목 수만큼 양식 행 스타일 복제 + 값 채우기
    for i, it in enumerate(plt_items, 1):
        r = 13 + i
        _apply_row_style(ws, r, style_template, template_height)
        ws.cell(r, 1, i)
        ws.cell(r, 2, lookup_barcode(it['item_code'], master) or '')
        ws.cell(r, 3, it['item_code'])
        ws.cell(r, 4, it['name'])
        ws.cell(r, 5, it['qty'])     # 예정수량 = 총 낱개
        ws.cell(r, 6, it['ipsu'])    # 입수
        ws.cell(r, 7, it['box'])     # 박스
        ws.cell(r, 8, 0)             # 낱개(잔량)
        # 유통기한: expiry_map(코드→값) 우선, 없으면 입력의 소비기한
        key = _master_key(it['item_code'])
        expiry = (expiry_map or {}).get(key) if key else None
        if expiry is None:
            expiry = normalize_expiry(it.get('expiry'))
        ws.cell(r, 9, expiry)
        ws.cell(r, 10, 'N')
        ws.cell(r, 11, '상온')

    # 3) 인쇄 페이지 설정 (A4 가로, 1페이지에 맞춤)
    last_row = 13 + len(plt_items)
    _setup_print_layout(ws, last_row)


def _fill_pallet_block(ws, plt_items, header_info, master, expiry_map,
                       plt_no, total_plts, block_start, style_template, template_height):
    """ws의 block_start행부터 한 PLT 블록(상단 1~13행 양식 + 데이터)을 채움.
    상단 1~13행 양식은 이미 복사돼 있다고 가정. 헤더 정보·데이터·스타일만 채움.
    Returns: 데이터 마지막 행 번호.
    """
    # 상단 정보 (1행 기준 6,3 / 6,10 / 7,10 / 8,3 / 8,10 / 8,11)
    ws.cell(block_start + 5, 3).value = SHIPPER_NAME
    ws.cell(block_start + 5, 10).value = header_info['date']
    ws.cell(block_start + 6, 10).value = DEFAULT_TIME
    ws.cell(block_start + 7, 3).value = header_info['po_no']
    ws.cell(block_start + 7, 10).value = header_info['vehicle']
    ws.cell(block_start + 7, 11).value = f'PLT {plt_no} / {total_plts}'

    data_start = block_start + 13
    for i, it in enumerate(plt_items, 1):
        r = data_start + i - 1
        _apply_row_style(ws, r, style_template, template_height)
        ws.cell(r, 1, i)
        ws.cell(r, 2, lookup_barcode(it['item_code'], master) or '')
        ws.cell(r, 3, it['item_code'])
        ws.cell(r, 4, it['name'])
        ws.cell(r, 5, it['qty'])
        ws.cell(r, 6, it['ipsu'])
        ws.cell(r, 7, it['box'])
        ws.cell(r, 8, 0)
        key = _master_key(it['item_code'])
        expiry = (expiry_map or {}).get(key) if key else None
        if expiry is None:
            expiry = normalize_expiry(it.get('expiry'))
        ws.cell(r, 9, expiry)
        ws.cell(r, 10, 'N')
        ws.cell(r, 11, '상온')

    return data_start + len(plt_items) - 1


def build_all_in_one_sheet(wb, pallets, header_info, master, expiry_map,
                           template_ws, style_template, template_height):
    """모든 PLT 양식+데이터를 한 시트에 페이지 단위로 통합. '전체' 시트로 만들고 맨 앞에 배치."""
    from openpyxl.worksheet.pagebreak import Break

    all_ws = wb.create_sheet('전체', 0)
    _copy_column_widths(template_ws, all_ws)

    total = len(pallets)
    current_row = 1
    last_data_row = 1

    for idx, p in enumerate(pallets):
        # 양식 1~13행(헤더 + 표 헤더) 복사
        _copy_row_block(template_ws, 1, 13, all_ws, current_row)
        _copy_merges_in_range(template_ws, all_ws, 1, 13, current_row - 1)

        last_data_row = _fill_pallet_block(
            all_ws, p['items'], header_info, master, expiry_map,
            p['plt_no'], total, current_row, style_template, template_height,
        )

        # 마지막 PLT가 아니면 페이지 나누기 추가
        if idx < total - 1:
            all_ws.row_breaks.append(Break(id=last_data_row))

        current_row = last_data_row + 1

    # '전체' 시트는 가로만 1페이지에 맞추고 세로는 페이지 나누기(row_breaks)에 맡김
    _setup_print_layout(all_ws, last_data_row, fit_to_height=0)
    return all_ws


def build_fcj_workbook(pallets, header_info, master, expiry_map=None,
                      template_path=FCJ_TEMPLATE):
    if not Path(template_path).exists():
        raise FileNotFoundError(f"FCJ 양식 파일이 없습니다: {template_path}")
    wb = openpyxl.load_workbook(template_path)
    base = wb.worksheets[0]

    # 양식의 14행 스타일 캡처
    style_template, template_height = _capture_row_style(base, 14)

    total = len(pallets)

    # 깨끗한 양식(1~13행 헤더 영역)을 별도로 보존: 첫 시트가 변경되기 전에 두 번째 워크북으로 다시 로드
    template_wb = openpyxl.load_workbook(template_path)
    template_ws = template_wb.worksheets[0]

    # 첫 시트는 PLT_01에 사용
    base.title = f'PLT_{pallets[0]["plt_no"]:02d}'
    fill_fcj_sheet(base, pallets[0]['items'], header_info, master, expiry_map,
                   pallets[0]['plt_no'], total, style_template, template_height)

    # 나머지는 시트 복제
    for p in pallets[1:]:
        new_ws = wb.copy_worksheet(base)
        new_ws.title = f'PLT_{p["plt_no"]:02d}'
        fill_fcj_sheet(new_ws, p['items'], header_info, master, expiry_map,
                       p['plt_no'], total, style_template, template_height)

    # 모든 PLT를 한 시트에 통합 ('전체' 시트, 맨 앞)
    build_all_in_one_sheet(wb, pallets, header_info, master, expiry_map,
                           template_ws, style_template, template_height)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------- 결과 ↔ DataFrame 변환 ----------
EDIT_COLUMNS = ['PLT 번호', 'Item code', 'Item', '입수', '박스', '낱개',
                'plt환산', 'Plt_1차', '소비기한']


def pallets_to_df(pallets):
    rows = []
    for p in pallets:
        for it in p['items']:
            rows.append({
                'PLT 번호': p['plt_no'],
                'Item code': it['item_code'],
                'Item': it['name'],
                '입수': it['ipsu'],
                '박스': it['box'],
                '낱개': it['qty'],
                'plt환산': it['plt_conv'],
                'Plt_1차': round(float(it['plt1']), 6),
                '소비기한': it.get('expiry'),
            })
    return pd.DataFrame(rows, columns=EDIT_COLUMNS)


def recompute_row(row):
    """박스가 변하면 낱개·Plt_1차를 자동 재계산."""
    try:
        box = int(row['박스']) if row['박스'] is not None else 0
    except (TypeError, ValueError):
        box = 0
    try:
        ipsu = int(row['입수']) if row['입수'] is not None else 0
    except (TypeError, ValueError):
        ipsu = 0
    try:
        bpp = float(row['plt환산']) if row['plt환산'] else 0.0
    except (TypeError, ValueError):
        bpp = 0.0
    row['낱개'] = box * ipsu
    row['Plt_1차'] = round(box / bpp, 6) if bpp else 0.0
    return row


def df_to_pallets(df, cfg):
    """편집된 DataFrame을 pallets 구조로 복원. 규칙은 박스 수로 자동 판단."""
    df_clean = df.dropna(subset=['Item code', 'PLT 번호']).copy()
    pallets = []
    for plt_no, group in df_clean.groupby('PLT 번호', sort=True):
        items = []
        for _, r in group.iterrows():
            items.append({
                'item_code': r['Item code'],
                'name': r['Item'],
                'ipsu': r['입수'],
                'box': int(r['박스']) if pd.notna(r['박스']) else 0,
                'qty': int(r['낱개']) if pd.notna(r['낱개']) else 0,
                'expiry': r['소비기한'] if pd.notna(r['소비기한']) else None,
                'plt_conv': r['plt환산'],
                'plt1': float(r['Plt_1차']) if pd.notna(r['Plt_1차']) else 0.0,
            })
        max_box = max((it['box'] for it in items), default=0)
        if max_box > cfg['medium_max']:
            rule = '4'
        elif max_box > cfg['small_max']:
            rule = '3'
        else:
            rule = '2'
        pallets.append({'plt_no': int(plt_no), 'rule': rule, 'items': items})
    return pallets


def validate_pallets(pallets, cfg):
    issues = []
    for p in pallets:
        plt_sum = sum(it['plt1'] for it in p['items'])
        cnt = len(p['items'])
        if plt_sum > cfg['plt_sum_max'] + 1e-6:
            issues.append(f"PLT {p['plt_no']}: Plt 합 {plt_sum:.3f} > 상한 {cfg['plt_sum_max']:.2f}")
        max_box = max((it['box'] for it in p['items']), default=0)
        if max_box > cfg['medium_max']:
            if cnt > 1:
                issues.append(f"PLT {p['plt_no']}: 대형 단독 파레트인데 {cnt}품목")
        elif max_box > cfg['small_max']:
            if cnt > cfg['medium_cap']:
                issues.append(f"PLT {p['plt_no']}: 중형 {cnt}품목 > 한도 {cfg['medium_cap']}")
        else:
            if cnt > cfg['small_cap']:
                issues.append(f"PLT {p['plt_no']}: 소형 {cnt}품목 > 한도 {cfg['small_cap']}")
    return issues


# ---------- Streamlit UI ----------
try:  # 단독 실행 시에만 (통합 Home.py에서 실행되면 무시)
    st.set_page_config(page_title='BNF 파레트 구분기', page_icon='📦', layout='wide')
except Exception:
    pass

st.title('📦 BNF 파레트 구분기')
st.caption('입고예정 Excel을 업로드하면 파레트 단위로 자동 구분합니다.')

with st.sidebar:
    st.header('🎛 기준값 설정')
    st.caption('변경하면 즉시 재계산됩니다.')

    if st.button('🔄 기본값으로 되돌리기', use_container_width=True):
        for k, v in DEFAULTS.items():
            st.session_state[k] = v

    small_max = st.number_input(
        '소형 박스 기준 (이 값 이하)', min_value=1, max_value=100,
        value=st.session_state.get('small_max', DEFAULTS['small_max']),
        step=1, key='small_max',
        help='이 박스 수 이하인 품목은 소형으로 분류됩니다.',
    )
    medium_max = st.number_input(
        '중형 박스 기준 (이 값 이하)', min_value=int(small_max) + 1, max_value=1000,
        value=max(st.session_state.get('medium_max', DEFAULTS['medium_max']), int(small_max) + 1),
        step=1, key='medium_max',
        help='소형 기준 초과 ~ 이 값 이하는 중형. 이 값을 초과하면 단독 파레트(대형).',
    )
    small_cap = st.number_input(
        '소형 파레트당 최대 품목 수', min_value=1, max_value=50,
        value=st.session_state.get('small_cap', DEFAULTS['small_cap']),
        step=1, key='small_cap',
    )
    medium_cap = st.number_input(
        '중형 파레트당 최대 품목 수', min_value=1, max_value=20,
        value=st.session_state.get('medium_cap', DEFAULTS['medium_cap']),
        step=1, key='medium_cap',
    )
    plt_sum_max = st.number_input(
        'Plt_1차 합 상한 (절대 기준)', min_value=0.1, max_value=2.0,
        value=float(st.session_state.get('plt_sum_max', DEFAULTS['plt_sum_max'])),
        step=0.05, format='%.2f', key='plt_sum_max',
    )

    cfg = {
        'small_max': int(small_max),
        'medium_max': int(medium_max),
        'small_cap': int(small_cap),
        'medium_cap': int(medium_cap),
        'plt_sum_max': float(plt_sum_max),
    }

    st.divider()
    st.markdown(f"""
**현재 적용 중인 규칙**
| 박스 | 정책 |
|---|---|
| > {cfg['medium_max']}박스 | 단독 파레트 (Plt합 ≥ {cfg['plt_sum_max']:.2f}이면 분할) |
| {cfg['small_max']+1} ~ {cfg['medium_max']}박스 | 최대 {cfg['medium_cap']}품목/파레트 |
| ≤ {cfg['small_max']}박스 | 최대 {cfg['small_cap']}품목/파레트 |

**절대 기준**: 한 파레트 내 Plt_1차 합 ≤ {cfg['plt_sum_max']:.2f}
""")

    st.divider()
    st.header('📤 출력 형식')
    output_mode = st.radio(
        '출력 형식 선택',
        ['기본 (입고정보 시트)', 'FCJ 양식 (파레트별 시트)'],
        index=0,
        label_visibility='collapsed',
    )

    fcj_header = None
    if output_mode.startswith('FCJ'):
        st.markdown('**입고 정보 입력**')
        st.caption('화주명·시간은 고정값(샘표식품 / 09:00)으로 채워집니다.')
        fcj_date = st.date_input('입고예정일자', value=date.today())
        fcj_po = st.text_input('PO NO', value='')
        fcj_vehicle = st.text_input('차량번호', value='')
        fcj_header = {
            'date': fcj_date,
            'po_no': fcj_po,
            'vehicle': fcj_vehicle,
        }

        st.divider()
        st.markdown('**📚 바코드 마스터**')
        master = load_master()
        if master.get('count'):
            st.success(f"등록됨: {master['count']:,}개 품목 ({master.get('source_name','-')})")
            st.caption(f"마지막 갱신: {master.get('updated_at','-')}")
        else:
            st.warning('등록된 마스터가 없습니다. 바코드는 빈칸으로 출력됩니다.')

        new_master = st.file_uploader(
            '마스터 엑셀 업로드(품목코드 + 바코드 컬럼)', type=['xlsx', 'xls'],
            key='master_uploader',
            help='헤더에 "품목코드"와 "바코드"(또는 영문) 컬럼이 있으면 자동 인식됩니다. 업로드 후 자동 저장.',
        )
        if new_master is not None:
            try:
                mapping, hr, cc, bc = parse_master_excel(new_master)
                if not mapping:
                    st.error('매핑 데이터가 비어있습니다. 파일 내용을 확인해주세요.')
                else:
                    saved = save_master(mapping, new_master.name)
                    st.success(
                        f'✅ 마스터 갱신 완료: {saved["count"]:,}개 품목 '
                        f'(헤더 {hr}행, 코드 {cc}열, 바코드 {bc}열)'
                    )
                    st.caption('다음 출력부터 자동 적용됩니다.')
            except Exception as e:
                st.error(f'마스터 파싱 실패: {e}')

        st.divider()
        st.markdown('**📅 유통기한 마스터** (매번 새로 업로드)')
        expiry_master = st.session_state.get('expiry_master', EMPTY_EXPIRY_MASTER)
        if expiry_master.get('count'):
            st.success(
                f"이번 세션 등록: {expiry_master['count']:,}개 품목 "
                f"({expiry_master.get('source_name','-')})"
            )
            st.caption(f"업로드 시각: {expiry_master.get('updated_at','-')}")
            dup_saved = expiry_master.get('duplicates', {})
            if dup_saved:
                st.caption(f"⚠️ 유통기한 2개 이상: {len(dup_saved)}개 품목 (가장 나중 것 사용)")
        else:
            st.info('💡 유통기한 마스터는 자동 저장되지 않으니 매번 새로 업로드해주세요.')

        new_expiry = st.file_uploader(
            '마스터 엑셀 업로드(품목코드 + 유통기한 컬럼)', type=['xlsx', 'xls'],
            key='expiry_uploader',
            help='헤더에 "품목코드"와 "유통기한"(또는 소비기한) 컬럼이 있으면 자동 인식됩니다. '
                 '같은 품목코드에 여러 유통기한이 있으면 가장 늦은 것을 사용합니다. '
                 '페이지 새로고침/앱 재시작 시 사라지므로 매번 업로드해주세요.',
        )
        if new_expiry is not None:
            try:
                flat_max, by_inv, inventories, dups, hr, cc, ec, inv_col, skipped = \
                    parse_expiry_master(new_expiry)
                if not flat_max:
                    st.error('매핑 데이터가 비어있습니다. 파일 내용을 확인해주세요.')
                else:
                    saved = {
                        'mapping': flat_max,
                        'by_inv': by_inv,
                        'inventories': inventories,
                        'duplicates': dups,
                        'count': len(flat_max),
                        'source_name': new_expiry.name,
                        'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    }
                    st.session_state.expiry_master = saved
                    expiry_master = saved
                    # 마스터가 갱신되면 매칭 결과도 재계산되도록 시그너처 무효화
                    st.session_state.pop('expiry_match_sig', None)
                    inv_msg = f', 인벤토리 {inv_col}열' if inv_col else ' (인벤토리 컬럼 없음)'
                    msg = (f'✅ 유통기한 마스터 등록 완료: {saved["count"]:,}개 품목 '
                           f'(헤더 {hr}행, 코드 {cc}열, 유통기한 {ec}열{inv_msg})')
                    if skipped:
                        msg += f' | 날짜 형식 인식 실패 {skipped}건 스킵'
                    st.success(msg)
                    if inventories and inventories != [NO_INVENTORY]:
                        st.caption(f'📦 보유 인벤토리: {len(inventories)}종 — '
                                   + ', '.join(inventories[:8])
                                   + (' …' if len(inventories) > 8 else ''))
                    if dups:
                        st.warning(f'⚠️ 인벤토리가 2개 이상인 품목: {len(dups)}개 (기본값 IC930, 사용자 변경 가능)')
                        with st.expander(f'중복 품목코드 보기 ({len(dups)}개)'):
                            dup_rows = []
                            for code, inv_exp_pairs in sorted(dups.items()):
                                dup_rows.append({
                                    '품목코드': code,
                                    '인벤토리 수': len(inv_exp_pairs),
                                    '인벤토리별 유통기한': ', '.join(f'{i}={e}' for i, e in inv_exp_pairs),
                                })
                            st.dataframe(pd.DataFrame(dup_rows),
                                         hide_index=True, use_container_width=True)
            except Exception as e:
                st.error(f'유통기한 마스터 파싱 실패: {e}')
    else:
        master = load_master()
        expiry_master = st.session_state.get('expiry_master', EMPTY_EXPIRY_MASTER)

    st.divider()
    st.caption('비고·순번 컬럼은 기본 출력에서 비워둡니다. 필요 시 결과 파일에서 직접 입력하세요.')

uploaded = st.file_uploader('입고예정 Excel 파일 (.xlsx)', type=['xlsx'])

if not uploaded:
    st.info('파일을 업로드하면 자동으로 분석됩니다.')
    st.stop()

# 입력 파일 파싱은 한 번만 (캐시) — 매 rerun마다 다시 파싱하면 느림
file_sig = (uploaded.name, uploaded.size)
if st.session_state.get('input_file_sig') != file_sig:
    try:
        st.session_state.input_file_sig = file_sig
        wb_in = parse_workbook(uploaded)
        st.session_state.cached_sheet_names = wb_in.sheetnames
        # 첫 시트 기준으로 미리 추출
        first_ws = wb_in.worksheets[0]
        first_hr = find_header_row(first_ws)
        if first_hr is None:
            st.session_state.cached_per_sheet = {}
        else:
            t_date, t_str = extract_title(first_ws)
            its = extract_items(first_ws, first_hr)
            st.session_state.cached_per_sheet = {
                first_ws.title: {
                    'header_row': first_hr,
                    'title_date': t_date,
                    'title_str': t_str,
                    'items': its,
                }
            }
        st.session_state.cached_wb = wb_in  # 다른 시트 선택용
    except Exception as e:
        st.error(f'파일을 열 수 없습니다: {e}')
        st.stop()

wb = st.session_state.cached_wb
sheet_names = st.session_state.cached_sheet_names

col_a, col_b = st.columns([1, 3])
with col_a:
    sheet = st.selectbox('시트 선택', sheet_names, index=0)

# 선택된 시트가 캐시에 없으면 파싱해서 저장
per_sheet = st.session_state.cached_per_sheet
if sheet not in per_sheet:
    ws = wb[sheet]
    header_row = find_header_row(ws)
    if header_row is None:
        st.error("헤더 행에서 'Item code'를 찾지 못했습니다. 시트를 확인해주세요.")
        st.stop()
    t_date, t_str = extract_title(ws)
    its = extract_items(ws, header_row)
    per_sheet[sheet] = {
        'header_row': header_row,
        'title_date': t_date,
        'title_str': t_str,
        'items': its,
    }

cur = per_sheet[sheet]
header_row = cur['header_row']
title_date = cur['title_date']
title_str = cur['title_str']
items = cur['items']

if not items:
    st.error('데이터 행이 없습니다.')
    st.stop()

# 입력 요약
total_box = sum(i['box'] for i in items)
total_qty = sum(i['qty'] or 0 for i in items)
total_plt = sum(i['plt1'] for i in items)

c1, c2, c3, c4 = st.columns(4)
c1.metric('품목 수', f'{len(items)}')
c2.metric('총 박스', f'{total_box:,}')
c3.metric('총 낱개', f'{total_qty:,}')
c4.metric('Plt환산 합', f'{total_plt:.2f}')

with st.expander('📥 입력 데이터 미리보기'):
    df_in = pd.DataFrame(items)
    st.dataframe(df_in, use_container_width=True, hide_index=True)

st.divider()

# 자동 파레트 구분 (캐시)
import hashlib
src_signature = hashlib.md5(
    (uploaded.name + str(uploaded.size) + sheet + json.dumps(cfg, sort_keys=True)).encode()
).hexdigest()

if (st.session_state.get('auto_pallets_sig') != src_signature
        or st.session_state.get('cached_auto_pallets') is None):
    st.session_state.auto_pallets_sig = src_signature
    st.session_state.cached_auto_pallets = build_pallets(items, cfg)
auto_pallets = st.session_state.cached_auto_pallets

if (st.session_state.get('source_sig') != src_signature
        or st.session_state.get('edited_df') is None):
    st.session_state.source_sig = src_signature
    st.session_state.edited_df = pallets_to_df(auto_pallets)

st.subheader('✏️ 결과 수정 및 검토')
st.caption(
    'PLT 번호를 같게 하면 같은 파레트, 다르게 하면 다른 파레트가 됩니다. '
    '박스 수를 바꾸면 낱개·Plt_1차가 자동으로 재계산됩니다. '
    '행 추가/삭제도 가능합니다.'
)

reset_col, info_col = st.columns([1, 3])
with reset_col:
    if st.button('🔄 자동 재구분으로 되돌리기', use_container_width=True):
        st.session_state.edited_df = pallets_to_df(auto_pallets)
        st.rerun()
with info_col:
    st.caption(f'자동 구분: 총 {len(auto_pallets)}개 파레트로 분할됨')

# 데이터 편집기
edited_df = st.data_editor(
    st.session_state.edited_df,
    key='pallet_editor',
    use_container_width=True,
    hide_index=True,
    num_rows='dynamic',
    disabled=['Item code', 'Item', '입수', 'plt환산'],
    column_config={
        'PLT 번호': st.column_config.NumberColumn(
            'PLT 번호', min_value=1, step=1, format='%d',
            help='같은 번호끼리 묶이면 같은 파레트입니다.',
        ),
        'Item code': st.column_config.NumberColumn('Item code', format='%d'),
        '박스': st.column_config.NumberColumn('박스', min_value=0, step=1, format='%d'),
        '낱개': st.column_config.NumberColumn('낱개', min_value=0, step=1, format='%d'),
        'Plt_1차': st.column_config.NumberColumn('Plt_1차', format='%.4f'),
        'plt환산': st.column_config.NumberColumn('plt환산', format='%g', disabled=True),
        '소비기한': st.column_config.NumberColumn('소비기한', format='%d'),
    },
)

# 박스 변경 감지 → 낱개/Plt_1차 자동 재계산
prev = st.session_state.edited_df
if not edited_df.equals(prev):
    # 박스가 바뀐 행은 낱개/Plt_1차 재계산
    recomputed = edited_df.copy()
    for idx in recomputed.index:
        if idx not in prev.index:
            # 새로 추가된 행도 재계산
            recomputed.loc[idx] = recompute_row(recomputed.loc[idx].to_dict())
            continue
        try:
            old_box = prev.loc[idx, '박스']
        except KeyError:
            continue
        new_box = recomputed.loc[idx, '박스']
        if pd.notna(new_box) and old_box != new_box:
            recomputed.loc[idx] = recompute_row(recomputed.loc[idx].to_dict())
    st.session_state.edited_df = recomputed
    if not recomputed.equals(edited_df):
        st.rerun()
    edited_df = recomputed

# 편집 결과로 pallets 재구성
edited_pallets = df_to_pallets(edited_df, cfg)
rule_count = {'4': 0, '3': 0, '2': 0}
for p in edited_pallets:
    rule_count[p['rule']] += 1

st.divider()
st.subheader(f'📊 현재: 총 {len(edited_pallets)}개 파레트')

c1, c2, c3 = st.columns(3)
c1.metric(f'규칙 4 (>{cfg["medium_max"]}박스, 단독)', f'{rule_count["4"]}개')
c2.metric(f'규칙 3 ({cfg["small_max"]+1}~{cfg["medium_max"]}박스, {cfg["medium_cap"]}품목)',
          f'{rule_count["3"]}개')
c3.metric(f'규칙 2 (≤{cfg["small_max"]}박스, {cfg["small_cap"]}품목)',
          f'{rule_count["2"]}개')

# 파레트별 요약
summary_rows = []
for p in edited_pallets:
    plt_sum = sum(it['plt1'] for it in p['items'])
    box_sum = sum(it['box'] for it in p['items'])
    summary_rows.append({
        'PLT 번호': p['plt_no'],
        '규칙': p['rule'],
        '품목 수': len(p['items']),
        '총 박스': box_sum,
        'Plt합': round(plt_sum, 4),
        '품목 목록': ' / '.join(it['name'] for it in p['items']),
    })
st.dataframe(pd.DataFrame(summary_rows), use_container_width=True, hide_index=True)

# 검증 메시지
issues = validate_pallets(edited_pallets, cfg)
if issues:
    st.warning('⚠️ 규칙 위반 항목:\n' + '\n'.join(f'- {x}' for x in issues))
else:
    st.success('✅ 모든 파레트가 규칙을 만족합니다.')

# 변경된 pallets를 다운로드 빌드용으로 사용
pallets = edited_pallets

# 📅 유통기한 매칭 단계 (FCJ 모드 한정)
expiry_map_for_output = None
if output_mode.startswith('FCJ'):
    st.divider()
    st.subheader('📅 유통기한 매칭')

    # 마스터 정보
    if expiry_master.get('count'):
        st.caption(f"마스터: {expiry_master['count']:,}개 품목 등록됨 "
                   f"(갱신 {expiry_master.get('updated_at','-')})")
    else:
        st.info('사이드바의 **📅 유통기한 마스터**에서 엑셀을 업로드하면 자동 매칭됩니다. '
                '마스터가 없어도 입력 엑셀의 소비기한이나 직접 입력으로 채울 수 있습니다.')

    # 이번 입고 품목(중복 코드 1줄로 통합)
    unique_items = {}
    for p in pallets:
        for it in p['items']:
            key = _master_key(it['item_code'])
            if key and key not in unique_items:
                unique_items[key] = it

    # 선택 가능한 인벤토리는 IC100/IC920/IC930 셋으로 고정
    inv_options = list(ALLOWED_INVENTORIES)

    def _decide_applied(input_v, master_v):
        """적용값과 출처를 결정."""
        if input_v is not None and master_v is not None:
            return max(input_v, master_v), '둘 다 → 늦은 값'
        if input_v is not None:
            return input_v, '입력값'
        if master_v is not None:
            return master_v, '마스터'
        return None, '미매핑'

    # 자동 매칭 결과 빌드
    def build_auto_expiry_df():
        rows = []
        for key, it in unique_items.items():
            input_v = normalize_expiry(it.get('expiry'))
            default_inv = default_inventory_for(it['item_code'], expiry_master) or ''
            # 인벤토리 컬럼이 없는 마스터(예: 출고진행현황)는 default_inv=''가 되므로,
            # 이 경우 전체 최댓값(flat-max)으로 조회한다. (인벤토리 있으면 해당 값 사용)
            master_v = lookup_expiry_master(it['item_code'], expiry_master, default_inv)
            applied, source = _decide_applied(input_v, master_v)
            selectable = selectable_inventories_for(it['item_code'], expiry_master)
            if selectable:
                note = f"선택 가능: {', '.join(selectable)}"
            else:
                held_all = held_inventories_for(it['item_code'], expiry_master)
                if held_all and held_all != [NO_INVENTORY]:
                    note = f"⚠ IC100/920/930에 재고 없음 (보유: {', '.join(held_all[:3])})"
                elif held_all == [NO_INVENTORY]:
                    note = '인벤토리 정보 없음'
                else:
                    note = '마스터 없음'
            rows.append({
                '품목코드': it['item_code'],
                '품목명': it['name'],
                '인벤토리': default_inv,
                '입력값': input_v,
                '마스터값': master_v,
                '적용 유통기한': applied,
                '출처': source,
                '비고': note,
            })
        return pd.DataFrame(rows)

    # session_state로 매칭 결과 관리 (입력/cfg/마스터 변경 시 재계산)
    match_sig = hashlib.md5(
        (src_signature + str(expiry_master.get('updated_at', ''))
         + str(expiry_master.get('count', 0))).encode()
    ).hexdigest()
    if (st.session_state.get('expiry_match_sig') != match_sig
            or st.session_state.get('expiry_match_df') is None):
        st.session_state.expiry_match_sig = match_sig
        st.session_state.expiry_match_df = build_auto_expiry_df()

    reset_col2, info_col2 = st.columns([1, 3])
    with reset_col2:
        if st.button('🔄 자동 매칭으로 되돌리기', use_container_width=True,
                     key='expiry_reset_btn'):
            st.session_state.expiry_match_df = build_auto_expiry_df()
            st.rerun()
    with info_col2:
        st.caption('적용 유통기한 셀을 직접 수정하면 출처가 "수기"로 변경됩니다. '
                   'YYYYMMDD 형식 8자리 정수를 입력하세요. (예: 20271231)')

    inv_select_options = [''] + inv_options if inv_options else ['']

    # data_editor 호출 *전*에 위젯의 누적 편집(edited_rows)을 적용해서
    # 마스터값/적용/출처를 미리 갱신. 이렇게 하면 별도 st.rerun() 호출 없이
    # 한 번의 rerun에서 모든 갱신이 화면에 반영됨 → 매칭 무한 반복 방지 + 빠름.
    editor_state = st.session_state.get('expiry_editor', {})
    edited_rows = editor_state.get('edited_rows', {}) if isinstance(editor_state, dict) else {}

    if edited_rows:
        df_pre = st.session_state.expiry_match_df.copy()
        for idx_key, changes in edited_rows.items():
            try:
                idx = int(idx_key)
            except (TypeError, ValueError):
                idx = idx_key
            if idx not in df_pre.index:
                continue
            if '인벤토리' in changes:
                new_inv = changes['인벤토리']
                code = df_pre.loc[idx, '품목코드']
                input_v = df_pre.loc[idx, '입력값']
                input_v = int(input_v) if pd.notna(input_v) else None
                # 인벤토리를 비우면(''): 전체 최댓값(flat-max)으로 조회
                master_v = lookup_expiry_master(code, expiry_master, new_inv)
                applied, source = _decide_applied(input_v, master_v)
                df_pre.loc[idx, '인벤토리'] = new_inv if new_inv else ''
                df_pre.loc[idx, '마스터값'] = master_v
                df_pre.loc[idx, '적용 유통기한'] = applied
                df_pre.loc[idx, '출처'] = source
            if '적용 유통기한' in changes:
                new_app = changes['적용 유통기한']
                df_pre.loc[idx, '적용 유통기한'] = new_app
                df_pre.loc[idx, '출처'] = '수기' if pd.notna(new_app) else df_pre.loc[idx, '출처']
        st.session_state.expiry_match_df = df_pre

    edited_expiry = st.data_editor(
        st.session_state.expiry_match_df,
        key='expiry_editor',
        use_container_width=True,
        hide_index=True,
        num_rows='fixed',
        disabled=['품목코드', '품목명', '입력값', '마스터값', '출처', '비고'],
        column_config={
            '품목코드': st.column_config.NumberColumn('품목코드', format='%d'),
            '인벤토리': st.column_config.SelectboxColumn(
                '인벤토리', options=inv_select_options,
                help='행별 인벤토리 선택. 변경 시 마스터값/적용 유통기한이 자동 재계산됩니다.',
            ),
            '입력값': st.column_config.NumberColumn('입력값', format='%d'),
            '마스터값': st.column_config.NumberColumn('마스터값', format='%d'),
            '적용 유통기한': st.column_config.NumberColumn(
                '적용 유통기한', format='%d', help='YYYYMMDD 형식 (예: 20271231)'),
        },
    )

    # 코드 → 적용 유통기한 dict 생성
    expiry_map_for_output = {}
    for _, row in edited_expiry.iterrows():
        code = row['품목코드']
        applied = row['적용 유통기한']
        if pd.notna(applied):
            key = _master_key(code)
            if key:
                expiry_map_for_output[key] = int(applied)

    # 알림: 미매핑 / 마스터 중복
    miss_cnt = edited_expiry['적용 유통기한'].isna().sum()
    dup_in_input = edited_expiry[edited_expiry['비고'] != '']
    c_e1, c_e2, c_e3 = st.columns(3)
    c_e1.metric('총 품목', f'{len(edited_expiry)}개')
    c_e2.metric('적용됨', f'{len(edited_expiry) - miss_cnt}개')
    c_e3.metric('미매핑', f'{miss_cnt}개')

    if miss_cnt:
        miss_names = edited_expiry[edited_expiry['적용 유통기한'].isna()]['품목명'].tolist()
        st.warning(f"⚠️ 미매핑 {miss_cnt}개 — 직접 입력하거나 마스터를 추가해주세요: "
                   + ', '.join(n[:14] for n in miss_names[:10]))
    if len(dup_in_input):
        with st.expander(f'ℹ️ 마스터에 유통기한 2개 이상이었던 품목 ({len(dup_in_input)}개) — 가장 나중 값 사용'):
            st.dataframe(dup_in_input[['품목코드', '품목명', '적용 유통기한', '비고']],
                         hide_index=True, use_container_width=True)

# 다운로드
today_str = datetime.now().strftime('%y%m%d')

def _pallets_signature(pallets):
    """파레트 데이터 변경 감지를 위한 가벼운 시그너처."""
    return tuple(
        (p['plt_no'], tuple((it['item_code'], it['box'], it.get('qty'), it.get('plt1'))
                            for it in p['items']))
        for p in pallets
    )


if output_mode.startswith('FCJ'):
    if not FCJ_TEMPLATE.exists():
        st.error(f'FCJ 양식 파일이 없습니다: {FCJ_TEMPLATE}')
    else:
        try:
            # 다운로드 빌드 캐시: pallets/header/master/expiry_map 시그너처가 같으면 재사용
            fcj_sig = hashlib.md5(
                (str(_pallets_signature(pallets))
                 + json.dumps({k: str(v) for k, v in (fcj_header or {}).items()}, sort_keys=True)
                 + str(master.get('updated_at', '')) + str(master.get('count', 0))
                 + json.dumps(expiry_map_for_output or {}, sort_keys=True)
                ).encode()
            ).hexdigest()
            if (st.session_state.get('fcj_sig') != fcj_sig
                    or st.session_state.get('fcj_bytes') is None):
                st.session_state.fcj_sig = fcj_sig
                st.session_state.fcj_bytes = build_fcj_workbook(
                    pallets, fcj_header, master, expiry_map=expiry_map_for_output)
            xlsx_bytes = st.session_state.fcj_bytes

            date_tag = fcj_header['date'].strftime('%y%m%d') if fcj_header.get('date') else today_str
            default_name = f'FCJ_입고예정_{date_tag}_파레트{len(pallets)}개.xlsx'

            # 마스터에 없는 품목 수 알려주기
            if master.get('count'):
                missing = [it for p in pallets for it in p['items']
                           if not lookup_barcode(it['item_code'], master)]
                if missing:
                    st.warning(f"바코드 미매핑 품목 {len(missing)}개: " +
                               ', '.join(sorted({it['name'][:14] for it in missing}))[:200])

            st.download_button(
                '⬇️ FCJ 양식 Excel 다운로드 (파레트별 시트)',
                data=xlsx_bytes,
                file_name=default_name,
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                type='primary',
                use_container_width=True,
            )
        except Exception as e:
            st.error(f'FCJ 파일 생성 실패: {e}')
else:
    totals = {'box': total_box, 'qty': total_qty, 'plt': total_plt}
    basic_sig = hashlib.md5(
        (str(_pallets_signature(pallets)) + str(title_date) + json.dumps(totals)).encode()
    ).hexdigest()
    if (st.session_state.get('basic_sig') != basic_sig
            or st.session_state.get('basic_bytes') is None):
        st.session_state.basic_sig = basic_sig
        st.session_state.basic_bytes = build_excel(pallets, title_date, totals)
    xlsx_bytes = st.session_state.basic_bytes
    default_name = f'입고정보_파레트구분_{today_str}.xlsx'

    st.download_button(
        '⬇️ 결과 Excel 다운로드',
        data=xlsx_bytes,
        file_name=default_name,
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        type='primary',
        use_container_width=True,
    )
