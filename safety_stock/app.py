# -*- coding: utf-8 -*-
"""안전재고 계산기 (Streamlit 페이지)

매월 하는 일: **그 달 CJ출고실적 1개만 업로드** → 히스토리에 누적 → 최근 12개월로 안전재고 재계산.
  · 비네이버(네이버·토스 제외)만 자동 추출, 행사no로 딜 식별(σ에서 제외)
  · 롤링 12개월이라 계절 변화(여름 급증 등) 자동 반영
  · 결과: 품목별 안전재고·Min·Max·파레트 → 엑셀 다운로드
"""
import io
import re
import sys
from datetime import datetime, timezone, timedelta
from pathlib import Path

import openpyxl
import pandas as pd
import streamlit as st

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from safety_stock import core, store  # noqa: E402
from master_hub import store as hub  # noqa: E402


def _apply_hub_master(master: dict, history: dict) -> dict:
    """공용 기준정보 허브의 입수·하대(배면×배단)·품명을 safety_stock 마스터에 덮어씌움.

    - 대상: 기존 마스터 코드 ∪ 히스토리 코드 (숫자 코드만).
    - 하대/입수: 비어있으면 채우고, 값이 다르면 허브값으로 교정(허브=ERP 원본, 정답).
    - 허브에 없는 코드(예: 8000xxx 특수코드)는 그대로 보존.
    반환: 갱신 통계.
    """
    hip, hha, hnm = hub.ipsu_map(), hub.hadae_map(), hub.name_map()
    if not hha and not hip:
        return {}
    codes = set(master) | {str(c) for c in history}
    st_ = {"하대교정": 0, "하대신규": 0, "입수교정": 0, "입수신규": 0, "신규품목": 0}
    for k in codes:
        if not str(k).isdigit():
            continue
        ci = int(k)
        h_plt, h_ip, h_nm = hha.get(ci), hip.get(ci), hnm.get(ci)
        if h_plt is None and h_ip is None:
            continue
        m = master.get(k)
        if m is None:
            master[k] = {"ip": int(h_ip) if h_ip else None,
                         "plt": int(h_plt) if h_plt else None,
                         "cat": "", "nm": h_nm or ""}
            st_["신규품목"] += 1
            continue
        if h_plt is not None:
            cur = m.get("plt")
            if cur in (None, 0, ""):
                m["plt"] = int(h_plt); st_["하대신규"] += 1
            elif int(cur) != int(h_plt):
                m["plt"] = int(h_plt); st_["하대교정"] += 1
        if h_ip is not None:
            cur = m.get("ip")
            if cur in (None, 0, ""):
                m["ip"] = int(h_ip); st_["입수신규"] += 1
            elif int(cur) != int(h_ip):
                m["ip"] = int(h_ip); st_["입수교정"] += 1
        if not m.get("nm") and h_nm:
            m["nm"] = h_nm
    return st_


def _stock_codes(file) -> dict:
    """BNF 상품별재고현황(.xls/.xlsx)에서 {코드: 현재고} 최소 추출(신규품목 후보 탐색용).
    transfer.parse_stock 와 동일한 헤더·코드 규칙(출고가능량 우선)."""
    df = pd.read_excel(io.BytesIO(file.getvalue()), header=None, dtype=object)
    hrow = 0
    for i in range(min(6, len(df))):
        vals = [str(x).strip() for x in df.iloc[i].tolist()]
        if any(v in ("상품코드", "품목코드", "CJ코드", "제품코드") for v in vals):
            hrow = i
            break
    hdr = [str(x).strip() for x in df.iloc[hrow].tolist()]

    def find(names, d=None):
        for j, h in enumerate(hdr):
            if h in names:
                return j
        return d
    ci = find(("상품코드", "품목코드", "CJ코드", "제품코드", "코드"), 0)
    cur_i = find(("출고가능량", "출고가능", "실가용재고", "가용재고"))
    if cur_i is None:
        cur_i = find(("재고수량", "현재고", "현재고(EA)", "재고EA"), 1)
    out = {}
    for r in range(hrow + 1, len(df)):
        row = df.iloc[r].tolist()
        c = row[ci] if ci < len(row) else None
        if c is None:
            continue
        c = str(int(c)) if isinstance(c, float) else str(c).strip()
        if not (c.isdigit() or c.startswith("P")):
            continue
        cur = None
        if cur_i is not None and cur_i < len(row):
            try:
                cur = float(row[cur_i])
            except (TypeError, ValueError):
                cur = None
        out[c] = cur
    return out


def _name_toks(s):
    """이름을 의미토큰으로(2자↑, 숫자·용량 제외). 브랜드·제품군 매칭용."""
    return [t for t in re.split(r"[\s/()·,\-_]+", str(s or ""))
            if len(t) >= 2 and not any(ch.isdigit() for ch in t)]


def _name_sim(a, b):
    """이름 유사도 점수(공유 토큰 수 + 첫 토큰(브랜드) 가중)."""
    ta = _name_toks(a)
    sb = set(_name_toks(b))
    if not ta or not sb:
        return 0
    score = sum(1 for t in ta if t in sb)
    if ta[0] in sb:
        score += 2
    return score


KST = timezone(timedelta(hours=9))

st.title("📐 BNF(비네이버) 안전재고 산출")
st.caption("매월 CJ출고실적 1개만 올리면 12개월 누적으로 안전재고를 재계산합니다. "
           "비네이버(네이버·토스 제외)·딜 제외·박스/파레트 기준.")

# ---------- 히스토리 로드 ----------
if "ss_history" not in st.session_state:
    h, src = store.load_history()
    st.session_state.ss_history = h
    st.session_state.ss_src = src
history = st.session_state.ss_history
master = store.load_master()

# 공용 기준정보 허브의 입수·하대 자동 적용 (기본 ON) — 한 번 올리면 이 페이지도 최신 반영
use_hub = st.checkbox("🗂️ 공용 기준정보 허브의 입수·하대 자동 적용 (권장)", value=True,
                      help="공용 기준정보 관리에서 올린 ERP Item 마스터의 입수·하대(배면×배단)를 "
                           "이 계산에 반영합니다. 끄면 내장 마스터값을 사용합니다.")
if use_hub:
    _hstat = _apply_hub_master(master, history)
    if _hstat and sum(_hstat.values()):
        st.caption("🗂️ 허브 반영 — " + " · ".join(f"{k} {v}건" for k, v in _hstat.items() if v))

# 품목코드 승계(입수변경 등으로 코드 변경) — 구코드 이력·마스터를 신코드로 이전
remap = store.load_code_remap()
remap_log = core.apply_code_remap(history, master, remap)
new_items = store.load_new_items()   # 신규품목 초기기준(유사품 기반)

meta = store.history_meta()
c1, c2, c3 = st.columns(3)
dmax = max((d for it in history.values() for d in it.keys()), default="—")
c1.metric("누적 품목수", f"{len(history):,}")
c2.metric("최신 데이터", dmax)
c3.metric("저장소", "Supabase" if store.use_supabase() else "로컬/세션")

# ---------- 설정 ----------
_saved_cfg = store.load_calc_settings()   # 영구 저장값(있으면 기본값으로)
def _cfg(k):
    return _saved_cfg.get(k, core.DEFAULT_SETTINGS[k])
with st.expander("⚙️ 설정 (리드타임·발주주기·서비스레벨)", expanded=False):
    cc = st.columns(4)
    lead = cc[0].number_input("리드타임(일)", 1, 30, int(_cfg("lead_time")))
    cycle = cc[1].number_input("발주주기(일)", 1, 30, int(_cfg("cycle")))
    _zopts = ["99% (2.33)", "98% (2.05)", "95% (1.65)", "99.5% (2.58)"]
    _zmap = {"99% (2.33)": 2.33, "98% (2.05)": 2.05, "95% (1.65)": 1.65, "99.5% (2.58)": 2.58}
    _zidx = next((i for i, o in enumerate(_zopts) if _zmap[o] == _cfg("z")), 0)
    zsel = cc[2].selectbox("서비스레벨", _zopts, index=_zidx)
    batch = cc[3].number_input("발주배치(일)", 1, 60, int(_cfg("batch")))
    z = _zmap[zsel]
    st.caption(f"노출기간 = 리드타임+발주주기 = **{lead+cycle}일**")
    if st.button("💾 이 설정을 기본값으로 저장 (다음에도 유지)", key="save_calc_cfg"):
        ok = store.save_calc_settings({"lead_time": int(lead), "cycle": int(cycle),
                                       "z": float(z), "batch": int(batch)})
        st.success("저장됨 — 다음에도 이 값으로 시작합니다." if ok
                   else "로컬(Supabase 미설정)에서는 이번 세션만 적용됩니다.")
settings = {"lead_time": lead, "cycle": cycle, "z": z, "batch": batch}

# ---------- 품목코드 승계 (입수변경 등으로 코드가 바뀐 경우) ----------
with st.expander(f"🔀 품목코드 승계 (입수변경) — 적용 {len(remap)}건", expanded=False):
    st.caption("입수량이 바뀌며 품목코드가 변경된 경우, **구코드의 판매이력(낱개 기준)**을 "
               "신코드로 승계해 안전재고를 이어서 산출합니다. 낱개 상품이 동일할 때만 사용하세요.")
    if remap_log:
        st.dataframe(pd.DataFrame(remap_log), width="stretch", hide_index=True)
    else:
        st.info("등록된 승계 규칙이 없습니다.")
    with st.form("remap_add", clear_on_submit=True):
        rc = st.columns([1, 1, 1, 2])
        r_old = rc[0].text_input("구코드")
        r_new = rc[1].text_input("신코드")
        r_ip = rc[2].number_input("신입수", 1, 9999, 24)
        r_plt = rc[3].number_input("신하대박스수(선택,0=구코드유지)", 0, 9999, 0)
        if st.form_submit_button("➕ 승계규칙 추가", type="secondary"):
            if r_old.strip() and r_new.strip():
                ok = store.add_code_remap(r_old.strip(), r_new.strip(), r_ip,
                                          r_plt or None)
                st.success("추가됨" + (" (Supabase)" if ok else " (로컬은 시드파일로만 반영)")
                           + " — 재계산하면 반영됩니다.")
                st.rerun()
            else:
                st.error("구코드·신코드를 입력하세요.")

# ---------- 신규품목 초기기준 (유사품 기반, 이력 없는 신제품) ----------
with st.expander(f"🆕 신규품목 초기기준 (유사품 기반) — 등록 {len(new_items)}건", expanded=False):
    st.caption("판매이력이 없는 신제품은 **비슷한 기존 품목**의 수요패턴(일평균·변동)을 "
               "계수배해 초기 Min/Max를 잡습니다. 판매일수가 임계(기본 30일)를 넘으면 "
               "**자기 통계로 자동 전환**(자립)됩니다. 신규품목은 '재고이동 계획'에서 재고는 있는데 "
               "기준이 없으면 ⚠️미등록 경고로 알려줍니다.")
    if new_items:
        prev, _pinfo = core.compute(dict(history), dict(master), settings, None, new_items)
        seeded = {r["품목코드"]: r for r in prev
                  if str(r.get("비고", "")).startswith("신규(유사")
                  or "신규(자립)" in str(r.get("비고", ""))}
        st.dataframe(pd.DataFrame([{
            "신규코드": n["code"], "품목명": n.get("nm", ""),
            "유사품목": n["analog"], "계수": n.get("factor", 1.0),
            "입수": n.get("ip"),
            "현재상태": ("자립(자기통계)" if "신규(자립)" in str(seeded.get(n["code"], {}).get("비고", ""))
                     else ("시드적용" if n["code"] in seeded else "유사품 이력없음")),
            "Min": seeded.get(n["code"], {}).get("Min(발주점,EA)"),
            "Max": seeded.get(n["code"], {}).get("Max(목표재고,EA)"),
        } for n in new_items]), width="stretch", hide_index=True)
    else:
        st.info("등록된 신규품목이 없습니다.")

    # 재고파일로 신규품목 자동 추출 — 재고 있는데 이력 없는 코드 → 유사품만 지정
    st.markdown("**📄 재고파일로 신규품목 자동 추출** (재고엔 있고 이력 없는 것)")
    up_ni = st.file_uploader("BNF 상품별재고현황(.xls/.xlsx) 업로드 → 신규품목 후보 자동 추출",
                             type=["xls", "xlsx"], key="ni_stock")
    if up_ni is not None:
        try:
            scodes = _stock_codes(up_ni)
            reg = {str(n["code"]) for n in new_items}
            cand = [c for c, cur in scodes.items()
                    if c not in history and c not in reg and (cur or 0) > 0]
            if not cand:
                st.success("재고파일에 미등록 신규품목이 없습니다 (모두 기준 보유).")
            else:
                # 후보 품명·입수·하대를 허브에서 보강(품명미상 방지 — 신규코드는 master에 아직 없음)
                try:
                    _hip, _hha, _hnm = hub.ipsu_map(), hub.hadae_map(), hub.name_map()
                except Exception:
                    _hip = _hha = _hnm = {}
                for c in cand:
                    if not c.isdigit():
                        continue
                    ci = int(c)
                    mm = dict(master.get(c) or {})
                    if not mm.get("nm") and _hnm.get(ci):
                        mm["nm"] = _hnm[ci]
                    if not mm.get("ip") and _hip.get(ci):
                        mm["ip"] = int(_hip[ci])
                    if not mm.get("plt") and _hha.get(ci):
                        mm["plt"] = int(_hha[ci])
                    if mm:
                        master[c] = mm
                hist_names = {h: master.get(h, {}).get("nm", "") for h in history}
                st.caption(f"{len(cand)}개 후보 — **유사품과 계수만** 정하면 됩니다. "
                           "입수·품명은 허브값 자동. 유사품은 **이름 비슷한 순**으로 정렬됩니다(입력해 검색도 가능).")
                with st.form("ni_bulk", clear_on_submit=True):
                    picks = {}
                    for c in cand:
                        m = master.get(c, {})
                        cnm = m.get("nm") or ""
                        # 유사품 옵션을 후보 이름과 유사한 순으로 정렬
                        ranked = sorted(history, key=lambda h: (-_name_sim(cnm, hist_names[h]), h))
                        an_opts = ["(선택 안함)"] + [f"{h} {hist_names[h]}".strip() for h in ranked]
                        cc = st.columns([1, 2, 2, 1])
                        cc[0].markdown(f"`{c}`")
                        cc[1].markdown(f"{cnm or '(품명 미상)'} · 입수 {m.get('ip') or '?'}")
                        an = cc[2].selectbox("유사품", an_opts, key=f"ni_an_{c}",
                                             label_visibility="collapsed")
                        fac = cc[3].number_input("계수", 0.1, 5.0, 1.0, 0.1,
                                                 key=f"ni_fac_{c}", label_visibility="collapsed")
                        picks[c] = (an, fac)
                    if st.form_submit_button("➕ 선택한 후보 등록", type="primary"):
                        n = 0
                        for c, (an, fac) in picks.items():
                            if an and an != "(선택 안함)":
                                acode = an.split()[0]
                                m = master.get(c, {})
                                store.add_new_item(c, acode, m.get("ip"), m.get("plt"),
                                                   m.get("nm", ""), fac)
                                n += 1
                        if n:
                            st.success(f"{n}건 등록 — 아래 2️⃣ 재계산 후 '적용'하면 이동계획에 반영됩니다.")
                            st.rerun()
                        else:
                            st.warning("유사품을 하나도 선택하지 않았습니다.")
        except Exception as e:
            st.error(f"후보 추출 실패: {e}")

    st.markdown("**✏️ 직접 등록**")
    with st.form("newitem_add", clear_on_submit=True):
        nc = st.columns([1, 1.4, 1, 1, 1, 0.8])
        n_code = nc[0].text_input("신규코드")
        n_nm = nc[1].text_input("품목명")
        n_analog = nc[2].text_input("유사품목코드")
        n_ip = nc[3].number_input("입수", 1, 9999, 24)
        n_plt = nc[4].number_input("하대박스수(선택)", 0, 9999, 0)
        n_factor = nc[5].number_input("계수", 0.1, 5.0, 1.0, 0.1)
        if st.form_submit_button("➕ 신규품목 등록", type="secondary"):
            if n_code.strip() and n_analog.strip():
                if n_analog.strip() not in history:
                    st.warning(f"유사품목 {n_analog}의 판매이력이 없습니다. 이력 있는 코드를 지정하세요.")
                else:
                    ok = store.add_new_item(n_code.strip(), n_analog.strip(), n_ip,
                                            n_plt or None, n_nm.strip(), n_factor)
                    st.success("등록됨" + (" (Supabase)" if ok else " (로컬은 시드파일로만)")
                               + " — 재계산하면 반영됩니다.")
                    st.rerun()
            else:
                st.error("신규코드·유사품목코드를 입력하세요.")

st.divider()

# ---------- 월별 CJ출고실적 업로드 → 누적 ----------
st.subheader("1️⃣ 이번 달 실적 반영")
up = st.file_uploader("CJ출고실적 xlsx 업로드 (그 달 1개, 'raw' 시트 포함)", type=["xlsx"], key="cj")


def _find_cols(ws):
    """헤더행에서 필요한 열 인덱스 자동 탐색."""
    hdr = [ws.cell(1, j).value for j in range(1, ws.max_column + 1)]
    want = {"ship_date": ["출고일자"], "code": ["품목코드"], "qty": ["출고수량"],
            "channel": ["판매채널"], "event_no": ["행사no", "행사NO", "행사번호"]}
    col = {}
    for key, names in want.items():
        for i, h in enumerate(hdr):
            if h is not None and str(h).strip() in names:
                col[key] = i
                break
    return col


if up is not None:
    if st.button("➕ 이 실적을 히스토리에 반영", type="primary"):
        try:
            wb = openpyxl.load_workbook(io.BytesIO(up.getvalue()), data_only=True, read_only=True)
            ws = wb["raw"] if "raw" in wb.sheetnames else wb[wb.sheetnames[0]]
            col = _find_cols(ws)
            need = {"ship_date", "code", "qty", "channel", "event_no"}
            if not need.issubset(col):
                st.error(f"필요한 열을 못 찾았습니다: 누락 {need - set(col)}. (출고일자·품목코드·출고수량·판매채널·행사no)")
            else:
                rows = ws.iter_rows(min_row=2, values_only=True)
                chunk = core.extract_cj_month(rows, col)
                wb.close()
                before = dmax
                core.merge_history(history, chunk)
                core.trim_history(history, months=14)
                st.session_state.ss_history = history
                saved = store.save_history(history)
                new_dmax = max((d for it in history.values() for d in it.keys()), default="—")
                nqty = sum(v[0] for days in chunk.values() for v in days.values())
                st.success(f"반영 완료: {len(chunk)}품목, 비네이버 {nqty:,.0f}개. "
                           f"최신 {before} → {new_dmax}. "
                           + ("Supabase 저장됨" if saved else "세션에만 반영(로컬)"))
        except Exception as e:
            st.error(f"처리 실패: {e}")

# ---------- 재계산 ----------
st.subheader("2️⃣ 안전재고 재계산")
if st.button("🔄 최근 12개월로 재계산", type="primary"):
    rows, info = core.compute(history, master, settings, None, new_items)
    st.session_state["ss_result"] = {"rows": rows, "info": info, "settings": settings}

res = st.session_state.get("ss_result")
if res:
    rows, info = res["rows"], res["info"]
    if not rows:
        st.warning("히스토리가 비어있습니다. 실적을 먼저 반영하세요.")
    else:
        by = {r["품목코드"]: r for r in rows}
        df = pd.DataFrame(rows).drop(columns=["months"], errors="ignore")  # dict컬럼 표시 제외
        tot = df["안전재고_EA"].sum()
        i1, i2, i3, i4 = st.columns(4)
        i1.metric("대상 품목", f"{info['품목수']:,}")
        i2.metric("안전재고 합계(연 기준)", f"{tot:,.0f} EA")
        i3.metric("계산 기간", f"{info['시작']} ~ {info['종료']}")
        i4.metric("노출기간", f"{info['노출기간']}일")
        if info.get("신규시드") or info.get("신규자립") or info.get("신규_유사품없음"):
            msg = (f"신규품목: 유사품 시드 {info.get('신규시드',0)}건 · "
                   f"자립(자기통계) {info.get('신규자립',0)}건")
            if info.get("신규_유사품없음"):
                msg += f" · ⚠️ 유사품 이력없음 {len(info['신규_유사품없음'])}건({', '.join(info['신규_유사품없음'][:5])})"
            st.info(msg)
        if info.get("신규최소SS"):
            st.info(f"🆕 신규(이력짧음)&활발 품목 {info['신규최소SS']}개 → 이력 없는 달 안전재고 "
                    f"최소 **{info.get('최소SS박스',0)}박스** 적용(비고 '신규최소SS'). "
                    "그 달 실적이 올라오면 자동으로 실제값 계산, 이력 6개월 넘으면 바닥값 해제.")
        if info.get("연중최소SS"):
            st.info(f"🔁 연중판매형(판매 달력월 8개월↑)&활발 품목 {info['연중최소SS']}개 → "
                    f"이력 없는 달만 안전재고 최소 **{info.get('연중최소SS박스',0)}박스** 적용(비고 '연중최소SS'). "
                    "이력 있는 달은 실제 계산 그대로. 판매월 적은 계절상품은 0 유지(유통기한 보호).")
        st.dataframe(df, width="stretch", height=460)

        # 계절(월별) 프로파일 뷰어
        with st.expander("📅 계절(월별) 프로파일 보기 — 성수기↑ 비수기↓ 반영 확인"):
            sel = st.selectbox("품목 선택", [r["품목코드"] for r in rows],
                               format_func=lambda c: f"{c}  {by[c]['품목명'][:26]}")
            mp = by[sel].get("months") or {}
            mdf = pd.DataFrame([{"월": f"{m}월", "일평균": mp.get(m, {}).get("rate"),
                                 "안전재고": mp.get(m, {}).get("ss"),
                                 "Min": mp.get(m, {}).get("mn"),
                                 "Max": mp.get(m, {}).get("mx")} for m in range(1, 13)])
            st.dataframe(mdf, width="stretch", hide_index=True)
            st.caption(f"연 고정: SS {by[sel]['안전재고_EA']} · Max {by[sel]['Max(목표재고,EA)']} "
                       "(참고). 이동계획은 계획일자의 **그 달** 값을 사용합니다.")

        # 엑셀 다운로드
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as xw:
            df.to_excel(xw, index=False, sheet_name="안전재고")
        dl, sv = st.columns([1, 1])
        dl.download_button("📥 안전재고 산출표 다운로드 (xlsx)", buf.getvalue(),
                           file_name=f"안전재고_산출_{datetime.now(KST):%y%m%d}.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           width="stretch")
        if sv.button("✅ 이 기준을 재고이동계획에 적용", type="secondary", width="stretch"):
            ok = store.save_baseline(res["rows"], res["settings"])
            st.success("재고이동계획 기준으로 저장됨(계절 월별 포함)" + (" (Supabase)" if ok else " (로컬 시드)"))
        st.caption("안전재고·Min·Max는 **계절(월별)** 로 저장됩니다. 이동계획은 계획일자 달의 값을 사용. "
                   "딜(행사)은 제외된 평상시 기준이며, 딜 물량은 이동계획의 이벤트/행사일정으로 별도 반영합니다.")
