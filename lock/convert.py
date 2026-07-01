"""
Lock 재고 엑셀 변환기
원본 (조회1_xxx.xlsx) -> 편집본 (락.xlsx)

변환 규칙:
1. 'sheet' 시트:
   - 두 번째 '출고진행' 컬럼(J열, 비어있음) 삭제
   - 필터: Location에 '-' 없음 + Lock 사유명 != '불가_실사차이'
   - 정렬: Location -> Item ID
2. 'Sheet1' 시트 신규 (피벗):
   - A3 헤더: Location | Item ID | Item | 유통기한 | 합계 : 재고수량2
   - 그룹: (Location, Item ID, 유통기한), 값: 재고수량(박스 수) 합계
   - Location은 그룹 첫 행에만 표시
   - 마지막에 '총합계' 행

사용법:
    python convert.py <입력파일.xlsx> [출력파일.xlsx]
    출력 파일을 생략하면 입력파일과 같은 폴더에 '입력파일_변환.xlsx' 로 저장
"""
import sys
from collections import defaultdict
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment

EXCLUDE_REASON = "불가_실사차이"


def keep_row(row):
    """필터: Location에 '-' 없음 + Lock 사유명 != 불가_실사차이"""
    location = row[5]  # F열
    reason = row[17]   # R열 (Lock 사유명)
    if location is None:
        return False
    if "-" in str(location):
        return False
    if reason == EXCLUDE_REASON:
        return False
    return True


def to_int(v):
    """1010357.0 -> 1010357. None은 그대로."""
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return int(v)
    return v


def parse_box_count(v):
    """'1 / 0 (12EA)' -> 1. 숫자면 그대로 정수화."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).strip()
    # 첫 토큰을 숫자로 (공백/슬래시 앞)
    head = s.split("/")[0].strip()
    try:
        return int(float(head))
    except ValueError:
        return v  # 못 파싱하면 원문 유지


def transform(src_path: Path, dst_path: Path):
    src_wb = load_workbook(src_path, data_only=True)
    if "sheet" not in src_wb.sheetnames:
        raise ValueError(f"입력 파일에 'sheet' 시트가 없습니다. 시트 목록: {src_wb.sheetnames}")
    src_ws = src_wb["sheet"]

    rows = list(src_ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("입력 파일이 비어 있습니다.")
    headers = list(rows[0])
    data = rows[1:]

    # 1) 첫 번째 '출고진행' 컬럼 (인덱스 8, 0 값으로 채워진 컬럼) 삭제
    DROP_IDX = 8
    new_headers = headers[:DROP_IDX] + headers[DROP_IDX + 1:]

    # 2) 필터 + 숫자 정수 변환
    filtered = []
    for r in data:
        if not keep_row(r):
            continue
        new_row = list(r[:DROP_IDX]) + list(r[DROP_IDX + 1:])
        new_row[2] = to_int(new_row[2])  # Item ID
        new_row[6] = to_int(new_row[6])  # 재고수량 (EA)
        new_row[7] = parse_box_count(new_row[7])  # 재고수량 (박스수)
        filtered.append(new_row)

    # 3) Location -> Item ID 정렬 (안정 정렬)
    filtered.sort(key=lambda r: (str(r[5]), r[2] if r[2] is not None else 0))

    # 출력 워크북 만들기 (Sheet1을 먼저, sheet를 나중에)
    out_wb = Workbook()
    # 기본 시트를 Sheet1로
    pivot_ws = out_wb.active
    pivot_ws.title = "Sheet1"

    # 4) 피벗 만들기 — 편집본 기준 컬럼 인덱스 (DROP 후)
    # 0:Lock 1:Inventory 2:Item ID 3:Item 4:단위 5:Location
    # 6:재고수량(EA) 7:재고수량(BOX수) 8:출고진행 9:출고가능 10:출고가능
    # 11:입고일 12:재고이동 13:제조일자 14:유통기한 ...
    pivot = defaultdict(int)
    item_name = {}  # (loc, item_id) -> item name
    for r in filtered:
        loc = r[5]
        item_id = r[2]
        item = r[3]
        exp = r[14]  # 유통기한
        qty = r[7] or 0  # 박스 수량
        key = (loc, item_id, exp)
        pivot[key] += qty
        item_name[(loc, item_id)] = item

    # 정렬: Location, Item ID, 유통기한
    pivot_keys = sorted(
        pivot.keys(), key=lambda k: (str(k[0]), k[1] if k[1] is not None else 0, str(k[2]))
    )

    # Sheet1: 헤더는 R3, 데이터는 R4부터
    header_font = Font(bold=True)
    pivot_ws["A3"] = "Location"
    pivot_ws["B3"] = "Item ID"
    pivot_ws["C3"] = "Item"
    pivot_ws["D3"] = "유통기한"
    pivot_ws["E3"] = "합계 : 재고수량2"
    for col in "ABCDE":
        pivot_ws[f"{col}3"].font = header_font

    last_loc = None
    row_idx = 4
    for k in pivot_keys:
        loc, item_id, exp = k
        item = item_name.get((loc, item_id))
        # Location은 그룹 첫 행에만 표시
        loc_cell = loc if loc != last_loc else None
        last_loc = loc
        pivot_ws.cell(row=row_idx, column=1, value=loc_cell)
        pivot_ws.cell(row=row_idx, column=2, value=item_id)
        pivot_ws.cell(row=row_idx, column=3, value=item)
        pivot_ws.cell(row=row_idx, column=4, value=exp)
        pivot_ws.cell(row=row_idx, column=5, value=pivot[k])
        row_idx += 1

    # 총합계 행
    pivot_ws.cell(row=row_idx, column=1, value="총합계").font = header_font
    pivot_ws.cell(row=row_idx, column=5, value=sum(pivot.values())).font = header_font

    # 컬럼 폭
    pivot_ws.column_dimensions["A"].width = 12
    pivot_ws.column_dimensions["B"].width = 12
    pivot_ws.column_dimensions["C"].width = 50
    pivot_ws.column_dimensions["D"].width = 12
    pivot_ws.column_dimensions["E"].width = 18

    # 5) 'sheet' 시트 작성
    data_ws = out_wb.create_sheet("sheet")
    data_ws.append(new_headers)
    for r in filtered:
        data_ws.append(r)
    for cell in data_ws[1]:
        cell.font = header_font

    out_wb.save(dst_path)
    print(f"[OK] 변환 완료: {dst_path}")
    print(f"  - 입력 행: {len(data)}개")
    print(f"  - 필터 후 행: {len(filtered)}개")
    print(f"  - 피벗 그룹: {len(pivot_keys)}개, 총합계: {sum(pivot.values())}")


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    src = Path(sys.argv[1])
    if not src.exists():
        print(f"[ERR] 파일이 없습니다: {src}")
        sys.exit(1)
    if len(sys.argv) >= 3:
        dst = Path(sys.argv[2])
    else:
        dst = src.with_name(src.stem + "_변환.xlsx")
    transform(src, dst)


if __name__ == "__main__":
    main()
