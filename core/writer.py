from datetime import datetime
import os
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
BOLD = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center")

# 거래처명 파싱: "...이름...(수량)" 에서 마지막 괄호 숫자 = 수량 (이름 안의 괄호 보존)
_CUST_RE = re.compile(r"^(.*)\((\d+)\)\s*$")
# 농협 장성/군위/경남 = 리스트 맨 아래
_BOTTOM_NH = ("장성", "군위", "경남")

LONG_COLS = ["Item ID", "Item", "유통기한", "매칭수량", "거래처명"]


def _is_bottom(cust: str) -> bool:
    c = str(cust)
    return ("농협" in c or "[NH]" in c) and any(b in c for b in _BOTTOM_NH)


def to_long(df: pd.DataFrame) -> pd.DataFrame:
    """wide 매칭결과 → (품목 × 거래처) 1행씩 long.
    거래처명 열("이름(수량), 이름(수량)…")을 파싱해 펼친다.
    정렬: 농협 장성/군위/경남 맨 아래, 그 외 거래처명, 같은 거래처 안에서 Item ID/유통기한.
    """
    rows = []
    for _, r in df.iterrows():
        raw = r.get("거래처명")
        if raw is None or (isinstance(raw, float) and pd.isna(raw)) or str(raw).strip() == "":
            continue
        for part in str(raw).split(", "):
            part = part.strip()
            if not part:
                continue
            m = _CUST_RE.match(part)
            name = m.group(1).strip() if m else part
            qty = int(m.group(2)) if m else None
            rows.append({
                "Item ID": r.get("Item ID"),
                "Item": r.get("Item"),
                "유통기한": r.get("유통기한"),
                "매칭수량": qty,
                "거래처명": name,
            })
    out = pd.DataFrame(rows, columns=LONG_COLS)
    if out.empty:
        return out
    out["_b"] = out["거래처명"].map(lambda c: 1 if _is_bottom(c) else 0)
    out["_id"] = pd.to_numeric(out["Item ID"], errors="coerce")
    out = out.sort_values(["_b", "거래처명", "_id", "유통기한"], kind="stable")
    return out.drop(columns=["_b", "_id"]).reset_index(drop=True)


def write_summary(df: pd.DataFrame, out_dir: str,
                  prefix: str = "지정출고_매칭결과") -> str:
    """거래처별 long 양식으로 저장 (Item ID·Item·유통기한·매칭수량·거래처명)."""
    os.makedirs(out_dir, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    path = os.path.join(out_dir, f"{prefix}_{ts}.xlsx")

    long_df = to_long(df)

    wb = Workbook()
    ws = wb.active
    ws.title = "정리"

    # 헤더: 3행 (1·2행 비움)
    for col_idx, name in enumerate(LONG_COLS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=name)
        cell.font = BOLD
        cell.fill = HEADER_FILL
        cell.alignment = CENTER

    for r, row in enumerate(long_df.itertuples(index=False), start=4):
        for c, value in enumerate(row, start=1):
            if value is pd.NA or (isinstance(value, float) and pd.isna(value)):
                value = None
            elif LONG_COLS[c - 1] in ("Item ID", "매칭수량") and value is not None:
                try:
                    value = int(value)
                except (ValueError, TypeError):
                    pass
            ws.cell(row=r, column=c, value=value)

    widths = {"Item ID": 11, "Item": 38, "유통기한": 12, "매칭수량": 10, "거래처명": 34}
    for i, name in enumerate(LONG_COLS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(name, 12)

    ws.freeze_panes = "A4"
    wb.save(path)
    return path
