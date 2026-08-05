"""
유통기한 재고 출고 모니터
- 로컬 실행: watchlist.json 사용
- 클라우드 배포: Supabase DB 사용 (st.secrets에 SUPABASE_URL, SUPABASE_KEY 설정)
"""

import streamlit as st
import pandas as pd
import io
import json
import re
import base64
from datetime import date, timedelta
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


_XLSX_COLS = ["창고", "품목코드", "품목명", "등록 유통기한", "리뉴얼구분", "창고입고", "리뉴얼(락)",
              "합친수량(Box)", "일평균출고", "가용일수", "예상소진일",
              "출고진행 유통기한", "최신 출고 유통기한", "상태", "비고", "담당자", "확인여부"]
_XLSX_FILL = {"red": "FFCCCC", "orange": "FFE0B2", "gray": "D9D9D9", "": "FFFFFF"}


def build_result_xlsx(df) -> bytes:
    """결과 DataFrame → 색상 입힌 엑셀 bytes (창고별로 정렬)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "유통기한 모니터"
    ws.append(_XLSX_COLS)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="305496")
        c.alignment = Alignment(horizontal="center", vertical="center")
    if df is not None and not df.empty:
        for _, r in df.sort_values(["창고", "상태"]).iterrows():
            ws.append([r.get(c, "") for c in _XLSX_COLS])
            fill = PatternFill("solid", fgColor=_XLSX_FILL.get(r.get("_color", ""), "FFFFFF"))
            for cell in ws[ws.max_row]:
                cell.fill = fill
    widths = [8, 11, 30, 13, 10, 18, 24, 13, 11, 10, 12, 20, 16, 14, 40, 12, 8]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ─── Supabase 감지 ─────────────────────────────────────────────────────────
# 우선순위: JAEGO_SUPABASE_* (재고모니터 전용/기존 프로젝트) > SUPABASE_* (통합 공용)
def _sb_url():
    try:
        return st.secrets.get("JAEGO_SUPABASE_URL") or st.secrets.get("SUPABASE_URL")
    except Exception:
        return None


def _sb_key():
    try:
        return st.secrets.get("JAEGO_SUPABASE_KEY") or st.secrets.get("SUPABASE_KEY")
    except Exception:
        return None


def _supabase_configured() -> bool:
    return bool(_sb_url()) and bool(_sb_key())

USE_SUPABASE = _supabase_configured()

if USE_SUPABASE:
    from supabase import create_client

    @st.cache_resource
    def _sb():
        return create_client(_sb_url(), _sb_key())

# ─── 컬럼 인덱스 ──────────────────────────────────────────────────────────
COL_품목코드   = 4
COL_품목명     = 5
COL_유통기한   = 9
COL_현재고Box = 11
COL_출고예정   = 14
COL_출고가능Box = 18
COL_LOCK      = 23

APP_DIR        = Path(__file__).parent
WATCHLIST_FILE = APP_DIR / "watchlist.json"


# ─── 워치리스트 CRUD ────────────────────────────────────────────────────────

WATCHLIST_TABLE_MISSING = False  # watchlist 테이블 미생성 감지 플래그


def load_watchlist() -> list[dict]:
    global WATCHLIST_TABLE_MISSING
    if USE_SUPABASE:
        try:
            resp = _sb().table("watchlist").select("*").order("id").execute()
            return resp.data or []
        except Exception:
            WATCHLIST_TABLE_MISSING = True   # 테이블 없음 → 빈 목록(안내는 UI에서)
            return []
    # 로컬 폴백
    if WATCHLIST_FILE.exists():
        with open(WATCHLIST_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return []


def add_item(code: str, expiry: str) -> bool:
    """추가. 중복이면 False 반환."""
    wl = load_watchlist()
    if any(str(r["code"]) == code and str(r["expiry"]) == expiry for r in wl):
        return False
    if USE_SUPABASE:
        try:
            _sb().table("watchlist").insert({"code": code, "expiry": expiry}).execute()
        except Exception:
            return False
    else:
        wl.append({"code": code, "expiry": expiry})
        _save_local(wl)
    return True


def delete_item(row_id) -> None:
    if USE_SUPABASE:
        try:
            _sb().table("watchlist").delete().eq("id", row_id).execute()
        except Exception:
            pass
    else:
        wl = load_watchlist()
        wl = [r for r in wl if r.get("id") != row_id]
        _save_local(wl)


def _save_local(wl: list) -> None:
    # 로컬 전용 - id가 없으면 부여
    for i, r in enumerate(wl):
        if "id" not in r:
            r["id"] = i + 1
    with open(WATCHLIST_FILE, "w", encoding="utf-8") as f:
        json.dump(wl, f, ensure_ascii=False, indent=2)


# ─── 담당자 매핑 (품목코드 → 담당자) ─────────────────────────────────────────

def _fetch_담당자_supabase() -> dict:
    """통합 Supabase app_settings['inventory_담당자']에서 로드 (재고분석기와 공유)."""
    try:
        url = st.secrets.get("SUPABASE_URL")
        key = st.secrets.get("SUPABASE_KEY")
        if not (url and key):
            return {}
        from supabase import create_client
        cli = create_client(url, key)
        r = cli.table("app_settings").select("value").eq("key", "inventory_담당자").execute()
        raw = (r.data[0].get("value") or {}) if r.data else {}
        return {str(k).strip(): str(v).strip() for k, v in raw.items()}
    except Exception:
        return {}


def _parse_담당자_xlsx(file) -> dict:
    """업로드 담당자 엑셀 → {품목코드: 담당자} (1열 코드, 2열 담당자)."""
    try:
        d = pd.read_excel(file, dtype=str, header=0)
        if d.shape[1] < 2:
            return {}
        cc, nc = d.columns[0], d.columns[1]
        m = {}
        for _, r in d.iterrows():
            c = str(r[cc]).strip()
            n = str(r[nc]).strip()
            if c and c.lower() != "nan" and n and n.lower() != "nan":
                m[c] = n
        return m
    except Exception:
        return {}


# ─── 확인 체크 (영속) — 기존 app_settings(jsonb)에 저장 → 별도 테이블/SQL 불필요 ──

CHECKS_FILE = APP_DIR / "checks.json"
CHECKS_KEY  = "jaego_checks"


def _settings_client():
    """app_settings 테이블이 있는 통합 Supabase(SUPABASE_*) 클라이언트."""
    try:
        url = st.secrets.get("SUPABASE_URL")
        key = st.secrets.get("SUPABASE_KEY")
        if not (url and key):
            return None
        from supabase import create_client
        return create_client(url, key)
    except Exception:
        return None


def load_checks() -> dict:
    """{key: True} — 확인 완료된 항목."""
    cli = _settings_client()
    if cli is not None:
        try:
            r = cli.table("app_settings").select("value").eq("key", CHECKS_KEY).execute()
            v = (r.data[0].get("value") or {}) if r.data else {}
            return {k: True for k, val in v.items() if val}
        except Exception:
            pass
    if CHECKS_FILE.exists():
        try:
            with open(CHECKS_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


def set_check(key: str, checked: bool) -> None:
    cli = _settings_client()
    if cli is not None:
        try:
            r = cli.table("app_settings").select("value").eq("key", CHECKS_KEY).execute()
            cur = (r.data[0].get("value") or {}) if r.data else {}
            if checked:
                cur[key] = True
            else:
                cur.pop(key, None)
            cli.table("app_settings").upsert(
                {"key": CHECKS_KEY, "value": cur}, on_conflict="key").execute()
            return
        except Exception:
            pass
    # 로컬 폴백
    data = load_checks()
    if checked:
        data[key] = True
    else:
        data.pop(key, None)
    with open(CHECKS_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


# ─── 리뉴얼 구분(전량/부분) — watchlist 스키마 변경 없이 app_settings에 저장 ──

SCOPES_FILE = APP_DIR / "scopes.json"
SCOPES_KEY  = "jaego_scopes"


def _scope_key(code: str, expiry: str) -> str:
    return f"{str(code).strip()}|{str(expiry).strip()}"


def load_scopes() -> dict:
    """{code|expiry: '전량'|'부분'} — 등록 시 지정한 리뉴얼 구분."""
    cli = _settings_client()
    if cli is not None:
        try:
            r = cli.table("app_settings").select("value").eq("key", SCOPES_KEY).execute()
            return (r.data[0].get("value") or {}) if r.data else {}
        except Exception:
            pass
    if SCOPES_FILE.exists():
        try:
            with open(SCOPES_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


def set_scope(key: str, scope: str) -> None:
    cli = _settings_client()
    if cli is not None:
        try:
            r = cli.table("app_settings").select("value").eq("key", SCOPES_KEY).execute()
            cur = (r.data[0].get("value") or {}) if r.data else {}
            if scope:
                cur[key] = scope
            else:
                cur.pop(key, None)
            cli.table("app_settings").upsert(
                {"key": SCOPES_KEY, "value": cur}, on_conflict="key").execute()
            return
        except Exception:
            pass
    data = load_scopes()
    if scope:
        data[key] = scope
    else:
        data.pop(key, None)
    with open(SCOPES_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


# ─── ABC 프로그램 출고 데이터(일평균출고) 연계 — 같은 Supabase(app_settings) ──

ABC_INDEX_KEY = "abc_monthly_index"


def _abc_month_key(ym: str) -> str:
    return f"abc_monthly_b64_{ym.replace('-', '_')}"


def _norm_code(v) -> str:
    return re.sub(r"\.0$", "", str(v).strip())


@st.cache_data(show_spinner=False)
def load_abc_daily_avg(n_months: int = 3):
    """ABC 프로그램의 최근 n개월 '일평균출고'를 품번별 평균으로 반환.
    반환: ({품번: 일평균출고(float)}, 사용월 리스트)  ·  없으면 ({}, [])."""
    cli = _settings_client()
    if cli is None:
        return {}, []
    try:
        r = cli.table("app_settings").select("value").eq("key", ABC_INDEX_KEY).execute()
        idx = (r.data[0].get("value") or []) if r.data else []
    except Exception:
        return {}, []
    months = sorted(set(idx))[-n_months:]           # 최근 n개월
    frames = []
    for ym in months:
        try:
            rr = cli.table("app_settings").select("value").eq(
                "key", _abc_month_key(ym)).execute()
            v = rr.data[0]["value"] if rr.data else None
            if not v:
                continue
            d = pd.read_excel(io.BytesIO(base64.b64decode(v)), sheet_name=0)
            d.columns = [str(c).strip() for c in d.columns]
            if "품번" not in d.columns or "일평균출고" not in d.columns:
                continue
            dd = d[["품번", "일평균출고"]].copy()
            dd["품번"] = dd["품번"].map(_norm_code)
            dd["일평균출고"] = pd.to_numeric(dd["일평균출고"], errors="coerce").fillna(0)
            frames.append(dd)
        except Exception:
            continue
    if not frames:
        return {}, months
    allm = pd.concat(frames, ignore_index=True)
    avg = allm.groupby("품번")["일평균출고"].mean()   # 등장한 월들의 평균
    return {k: float(v) for k, v in avg.items()}, months


# ─── 수량 파싱 ─────────────────────────────────────────────────────────────

def has_qty(val) -> bool:
    if pd.isna(val) or str(val).strip() in ("", "nan", "None", "NaN"):
        return False
    tokens = re.findall(r"[\d,]+", str(val))
    return any(float(t.replace(",", "")) > 0 for t in tokens)


def to_exp_str(raw) -> str | None:
    if pd.isna(raw) or str(raw).strip() in ("", "nan"):
        return None
    try:
        return str(int(float(str(raw).strip())))
    except Exception:
        return str(raw).strip()


# ─── 분석 로직 ─────────────────────────────────────────────────────────────

def _to_num(v) -> float:
    if pd.isna(v):
        return 0.0
    s = str(v).replace(",", "").strip()
    if s == "" or s.lower() in ("nan", "none"):
        return 0.0
    try:
        return float(s)
    except Exception:
        return 0.0


def analyze_item(df: pd.DataFrame, code: str, target_exp: str, scope: str = "미지정") -> dict:
    """등록 (품목,유통기한) 분석 결과 dict 반환.
    keys: 품목명, status, color, note, 출고중, 입고, 락, 락누락
      · 입고   : 등록 유통기한 재고가 이 창고에 입고됐는지
      · 락     : 리뉴얼 구분(전량/부분)에 비춰 본 락 상태
      · 락누락 : 전량으로 등록됐는데 락 안 걸린 재고가 있으면 True
    """
    mask    = df.iloc[:, COL_품목코드].astype(str).str.strip() == code.strip()
    item_df = df[mask]

    _EMPTY_QTY = {"선입가용": 0.0, "동일미락": 0.0, "합친수량": 0.0}
    if item_df.empty:
        return {"품목명": "-", "status": "미입고", "color": "gray",
                "note": "품목코드가 이 창고 파일에 없음", "출고중": [],
                "입고": "❌ 미입고(품목없음)", "락": "-", "락누락": False, **_EMPTY_QTY}

    품목명 = str(item_df.iloc[0, COL_품목명])

    exp_info: dict[str, dict] = {}
    for _, row in item_df.iterrows():
        exp = to_exp_str(row.iloc[COL_유통기한])
        if exp is None:
            continue
        e = exp_info.setdefault(exp, {"has_lock_free": False, "in_출고": False,
                                      "qty": 0.0, "avail": 0.0})
        if not has_qty(row.iloc[COL_LOCK]):
            e["has_lock_free"] = True
        if has_qty(row.iloc[COL_출고예정]):
            e["in_출고"] = True
        e["qty"]   += _to_num(row.iloc[COL_현재고Box])
        e["avail"] += _to_num(row.iloc[COL_출고가능Box])   # 출고가능(Box) = 락·이동 제외 가용

    all_exps    = sorted(exp_info.keys())
    출고중_목록 = [e for e in all_exps if exp_info[e]["in_출고"]]

    if target_exp not in all_exps:
        return {"품목명": 품목명, "status": "미입고", "color": "gray",
                "note": f"등록 유통기한({target_exp}) 재고가 이 창고에 없음", "출고중": 출고중_목록,
                "입고": "❌ 미입고(유통기한 없음)", "락": "-", "락누락": False, **_EMPTY_QTY}

    t = exp_info[target_exp]
    입고 = "✅ 입고" if t["qty"] > 0 else "✅ 입고(현재고 0)"
    has_free = t["has_lock_free"]           # 락 안 걸린(출고가능) 재고 존재 여부
    락누락 = False
    if scope == "전량":
        if has_free:
            락, 락누락 = "🔴 락 누락 의심 (전량인데 미락 재고 있음)", True
        else:
            락 = "🔒 전량 락 (정상)"
    elif scope == "부분":
        락 = "🔓 부분 (락 해제분 정상)" if has_free else "🔒 전량 락"
    else:  # 미지정 — 구분 등록 안 됨
        락 = "🔒 전량 락" if not has_free else "🔓 락 일부해제"

    idx         = all_exps.index(target_exp)
    before_exps = all_exps[:idx]
    after_exps  = all_exps[idx + 1:]
    직전        = before_exps[-1] if before_exps else None
    이후_출고   = [e for e in after_exps if exp_info[e]["in_출고"]]

    # 지금 나갈 수 있는 재고: 빠른 유통기한 가용 + 동일 유통기한 락 안 걸린 가용
    선입가용 = sum(exp_info[e]["avail"] for e in before_exps)
    동일미락 = t["avail"]
    합친수량 = 선입가용 + 동일미락

    if 이후_출고:
        status, color, note = "위험", "red", f"이후 유통기한 출고진행 중: {', '.join(이후_출고)}"
    elif t["in_출고"]:
        status, color, note = "주의(락 점검)", "orange", f"동일 유통기한({target_exp}) 출고진행 중"
    elif not t["has_lock_free"] and 직전 and exp_info[직전]["in_출고"]:
        status, color, note = ("주의(락 점검)", "orange",
                               f"직전 유통기한({직전}) 출고진행 중  /  락 없는 동일 유통기한 재고 없음")
    else:
        status, color = "정상", ""
        note = "락 없는 동일 유통기한 재고 있음" if t["has_lock_free"] else "출고진행 없음"

    return {"품목명": 품목명, "status": status, "color": color, "note": note,
            "출고중": 출고중_목록, "입고": 입고, "락": 락, "락누락": 락누락,
            "선입가용": 선입가용, "동일미락": 동일미락, "합친수량": 합친수량}


# ─── UI ───────────────────────────────────────────────────────────────────

try:  # 단독 실행 시에만 (통합 Home.py에서 실행되면 무시)
    st.set_page_config(page_title="유통기한 재고 모니터", layout="wide", page_icon="📦")
except Exception:
    pass
st.title("📦 리뉴얼 재고 출고 시점 점검")
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent))
from page_help import show_help  # noqa: E402
show_help({
    "목적": "신규 파우치 워치리스트를 관리하고, 재고 파일을 올려서 유통기한 임박 여부를 점검.",
    "필요한 파일": "재고조회 xlsx (Location별). 워치리스트는 Supabase(공유) 또는 로컬 json.",
    "사용 순서": "1. 워치리스트 확인/편집 (신규 파우치 품번·품명·기준일)\n"
                 "2. 재고 파일 업로드\n"
                 "3. 각 품번별 잔여 유통기한·소진 예상 표시\n"
                 "4. 필요 시 결과 다운로드",
    "참고": "클라우드 배포에선 Supabase 공유 리스트를 사용.",
})

# watchlist 테이블 미생성 감지 → 1회 안내 (테이블 만들면 사라짐)
if USE_SUPABASE:
    load_watchlist()  # 플래그 세팅용
    if WATCHLIST_TABLE_MISSING:
        st.warning(
            "⚠ `watchlist` 테이블이 없습니다. 방법 2가지:\n\n"
            "**A) 기존 재고모니터 데이터 그대로 쓰기** — 앱 Secrets에 기존 재고모니터 "
            "Supabase 값을 `JAEGO_SUPABASE_URL` / `JAEGO_SUPABASE_KEY` 로 추가 (이관 불필요)\n\n"
            "**B) 이 통합 Supabase에 새로 만들기** — SQL Editor에서 1회 실행:\n"
            "```sql\ncreate table watchlist (\n  id serial primary key,\n"
            "  code varchar(50) not null,\n  expiry varchar(8) not null,\n"
            "  unique(code, expiry)\n);\n```")

# ── 품목 등록 (숨김 토글) ──────────────────────────────────────────────────
col_cb, col_desc = st.columns([2, 5])
with col_cb:
    show_register = st.checkbox("⚙️ 품목 등록/삭제 열기", value=False)
with col_desc:
    st.markdown("**신규 파우치 품목코드 유통기한 등록**")

if show_register:
    with st.container(border=True):
        st.subheader("🔖 품목 등록 / 삭제")
        left, right = st.columns([1, 1])

        # 등록
        with left:
            with st.form("add_form", clear_on_submit=True):
                inp_code   = st.text_input("품목코드", placeholder="예: 2061438")
                inp_expiry = st.text_input("유통기한 (YYYYMMDD)", placeholder="예: 20270827")
                inp_scope  = st.radio(
                    "리뉴얼 구분", ["전량", "부분"], horizontal=True,
                    help="전량 = 이 유통기한 재고 전부를 락 걸어야 함(락 누락 감지). "
                         "부분 = 일부만 신규 파우치라 락 해제분이 있어도 정상")
                submitted  = st.form_submit_button("➕ 등록", width='stretch')

            if submitted:
                if inp_code and inp_expiry:
                    ok = add_item(inp_code.strip(), inp_expiry.strip())
                    set_scope(_scope_key(inp_code, inp_expiry), inp_scope)  # 구분은 항상 최신값 저장
                    if ok:
                        st.success(f"등록 완료: **{inp_code.strip()}** / `{inp_expiry.strip()}` · {inp_scope}")
                        st.rerun()
                    else:
                        st.warning(f"이미 등록된 항목입니다 (리뉴얼 구분은 **{inp_scope}** 로 갱신됨)")
                        st.rerun()
                else:
                    st.error("품목코드와 유통기한을 모두 입력해주세요")

        # 삭제
        with right:
            wl = load_watchlist()
            _scopes = load_scopes()
            if not wl:
                st.info("등록된 품목이 없습니다")
            else:
                st.markdown("**등록 목록** (구분 미지정은 재등록해 지정하세요 · 🗑 삭제)")
                for item in wl:
                    sc = _scopes.get(_scope_key(item["code"], item["expiry"]), "미지정")
                    _badge = {"전량": "🔒 전량", "부분": "🔓 부분"}.get(sc, "⚪ 미지정")
                    c1, c2 = st.columns([5, 1])
                    c1.markdown(f"**{item['code']}** &nbsp; `{item['expiry']}` &nbsp; {_badge}")
                    if c2.button("🗑", key=f"del_{item['id']}"):
                        delete_item(item["id"])
                        set_scope(_scope_key(item["code"], item["expiry"]), "")  # 구분도 삭제
                        st.rerun()

st.divider()

# ── 담당자 데이터 (분석결과에 품목 담당자 표시) ──────────────────────────────
with st.expander("👤 담당자 데이터 (분석결과에 품목 담당자 표시)"):
    st.caption("품목코드 → 담당자 매핑. 재고분석기에 등록된 담당자(Supabase)를 자동으로 불러오며, "
               "파일을 올리면 그 파일이 우선합니다. (1열: 품목코드, 2열: 담당자)")
    up_dam = st.file_uploader("담당자 로우데이터 xlsx", type=["xlsx"], key="jaego_dam")
    if up_dam is not None:
        st.session_state["_dam_map"] = _parse_담당자_xlsx(up_dam)
    dam_map = st.session_state.get("_dam_map") or _fetch_담당자_supabase()
    if dam_map:
        _src = "업로드 파일" if st.session_state.get("_dam_map") else "Supabase(재고분석기 공유)"
        st.success(f"담당자 {len(dam_map)}명 로드됨 · 출처: {_src}")
    else:
        st.info("담당자 데이터 없음 — '담당자' 열은 비어서 표시됩니다.")

# ── 파일 업로드 및 분석 ────────────────────────────────────────────────────
st.header("📂 파일 업로드")
uploaded = st.file_uploader(
    "로케이션별 재고조회 Excel 파일 (.xlsx)",
    type=["xlsx", "xls"],
)

if uploaded:
    wl = load_watchlist()
    if not wl:
        st.warning("⚠️ 등록된 품목이 없습니다. 위 '품목 등록/삭제 열기'를 체크하여 품목을 등록해주세요.")
    else:
        df = pd.read_excel(uploaded, dtype=str, header=0)
        # ── 창고 선택: 전체창고 파일도 창고별로 분리 (섞이지 않음) ──
        _whs_all = sorted({str(v).strip() for v in df.iloc[:, 0].dropna()
                           if str(v).strip()})
        # 기본 체크: IC930·IC920·IC906 (파일에 있는 것만)
        _default = [w for w in ("IC930", "IC920", "IC906") if w in _whs_all] or _whs_all
        if len(_whs_all) > 1:
            sel_whs = st.multiselect(
                "🏬 창고 선택 (창고별로 분리 표시 — 보고 싶은 창고만 체크, 섞이지 않음)",
                _whs_all, default=_default)
        else:
            sel_whs = _whs_all
        if not sel_whs:
            sel_whs = _whs_all
        scope_map = load_scopes()
        abc_avg, abc_months = load_abc_daily_avg(3)
        _today = date.today()
        with st.spinner("분석 중..."):
            rows = []
            for wh in sel_whs:
                df_wh = df[df.iloc[:, 0].astype(str).str.strip() == wh]
                for item in wl:
                    sc = scope_map.get(_scope_key(item["code"], item["expiry"]), "미지정")
                    a = analyze_item(df_wh, str(item["code"]), str(item["expiry"]), sc)
                    # 미입고(품목/유통기한 없음)는 IC930에서만 표시 — 다른 창고는 없어도 무시
                    if a["color"] == "gray" and wh != "IC930":
                        continue
                    출고중 = a["출고중"]
                    # 지금 나갈 수 있는 재고(합친수량) ÷ 3개월 일평균출고 → 가용일수·예상소진일
                    합 = a["합친수량"]
                    daily = abc_avg.get(_norm_code(item["code"]))
                    if 합 > 0 and daily and daily > 0:
                        days = 합 / daily
                        가용일수_s = f"{days:.1f}일"
                        예상소진일 = (_today + timedelta(days=round(days))).strftime("%Y-%m-%d")
                    else:
                        가용일수_s = "-"
                        예상소진일 = "-"
                    rows.append(
                        {
                            "창고":            wh,
                            "품목코드":         str(item["code"]),
                            "품목명":           a["품목명"],
                            "등록 유통기한":     str(item["expiry"]),
                            "리뉴얼구분":       sc,
                            "창고입고":         a["입고"],
                            "리뉴얼(락)":       a["락"],
                            "합친수량(Box)":    round(합),
                            "일평균출고":       (round(daily, 1) if daily else "-"),
                            "가용일수":         가용일수_s,
                            "예상소진일":       예상소진일,
                            "출고진행 유통기한":  ", ".join(출고중) if 출고중 else "-",
                            "최신 출고 유통기한": max(출고중) if 출고중 else "-",
                            "상태":            a["status"],
                            "비고":            a["note"],
                            "담당자":          dam_map.get(str(item["code"]).strip(), ""),
                            "_color":          a["color"],
                            "_lockmiss":       a["락누락"],
                        }
                    )

        result_df = pd.DataFrame(rows)
        if result_df.empty:
            st.info("선택한 창고에서 등록 품목을 찾지 못했습니다. 창고 체크를 확인하세요.")
            st.stop()
        BG = {"red": "#FFCCCC", "orange": "#FFE0B2", "gray": "#E0E0E0", "": "#FFFFFF"}

        # 전체 요약 (미입고는 IC930에서만 집계되므로 분모는 실제 표시 건수)
        _n_reg = len(result_df)
        _n_in  = int((result_df["창고입고"].astype(str).str.startswith("✅")).sum())
        c1, c2, c3, c4, c5 = st.columns(5)
        c1.metric("📦 창고입고", f"{_n_in}/{_n_reg}",
                  help="등록(품목×창고) 중 해당 유통기한 재고가 입고된 건")
        c2.metric("🔴 락누락", int(result_df["_lockmiss"].sum()),
                  help="전량 등록인데 락 안 걸린 재고가 있음 — 락 확인 필요")
        c3.metric("🔴 위험", int((result_df["_color"] == "red").sum()),
                  help="이후 유통기한 출고진행 중")
        c4.metric("🟠 주의", int((result_df["_color"] == "orange").sum()),
                  help="동일/직전 유통기한 출고진행 중")
        c5.metric("❌ 미입고", int((result_df["_color"] == "gray").sum()),
                  help="품목/등록 유통기한 재고가 이 창고에 없음")

        if abc_months:
            st.caption(f"📈 **가용일수 = 합친수량(선입가용+동일미락) ÷ 일평균출고** · "
                       f"일평균출고 기준: ABC 최근 {len(abc_months)}개월 {', '.join(abc_months)} 평균 "
                       f"· 예상소진일 = 오늘({_today:%Y-%m-%d}) + 가용일수")
        else:
            st.caption("📈 가용일수·예상소진일: ABC 월별 출고 데이터가 없어 계산 불가(–). "
                       "ABC분석 프로그램에 월별 데이터를 등록하면 자동 반영됩니다.")

        # ── 확인 체크 로드 (+ 엑셀용 확인여부 열) ──
        checks = load_checks()

        def _mk_key(r):
            return "|".join([str(r["창고"]), str(r["품목코드"]), str(r["등록 유통기한"]),
                             str(r["상태"]), str(r["출고진행 유통기한"])])
        result_df["확인여부"] = ["✔" if checks.get(_mk_key(r), False) else ""
                              for _, r in result_df.iterrows()]

        # 비정상만 보기 / 확인완료 숨기기 / 엑셀 다운로드
        _f1, _f2, _f3 = st.columns([2, 2, 2])
        with _f1:
            abn_only = st.toggle("⚠ 비정상(위험·주의·미입고·락누락)만 보기", value=False)
        with _f2:
            hide_checked = st.toggle("✔ 확인완료 숨기기", value=False,
                                     help="이미 확인(체크)한 항목을 목록에서 숨겨 반복 작업 방지")
        view_df = (result_df[result_df["_color"].isin(["red", "orange", "gray"])
                             | result_df["_lockmiss"]]
                   if abn_only else result_df)
        with _f3:
            st.download_button(
                "📥 결과 엑셀 다운로드",
                build_result_xlsx(view_df),
                f"유통기한모니터_{date.today():%Y%m%d}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width='stretch')

        if view_df.empty:
            st.info("표시할 결과가 없습니다 (비정상 없음).")
        st.caption("✔ '확인' 열을 체크하면 저장되어, 다음에 다시 분석해도 유지됩니다 (동일 작업 반복 방지).")
        st.subheader("📊 분석 결과 (창고별)")

        _SEV = {"red": 0, "orange": 1, "": 2, "gray": 3}
        _EMO = {"red": "🔴", "orange": "🟠", "": "✅", "gray": "❌"}
        for wh in sel_whs:
            wh_df = view_df[view_df["창고"] == wh].copy()
            if wh_df.empty:
                continue
            wh_df = wh_df.sort_values("_color", key=lambda s: s.map(_SEV),
                                      kind="stable").reset_index(drop=True)
            # 행별 key + 확인 seed (+ 확인완료 숨기기 필터)
            keys, seed, keep = [], [], []
            for _, r in wh_df.iterrows():
                k = _mk_key(r)
                chk = checks.get(k, False)
                if hide_checked and chk:
                    keep.append(False)
                    continue
                keep.append(True)
                keys.append(k)
                seed.append(chk)
            wh_df = wh_df[keep].reset_index(drop=True)
            if wh_df.empty:
                continue
            _r = int((wh_df["_color"] == "red").sum())
            _o = int((wh_df["_color"] == "orange").sum())
            _k = int((wh_df["_color"] == "").sum())
            _m = int((wh_df["_color"] == "gray").sum())
            _lm = int(wh_df["_lockmiss"].sum())
            st.markdown(f"#### 🏬 {wh}  —  🔴 위험 {_r} · 🟠 주의 {_o} · ✅ 정상 {_k} · "
                        f"❌ 미입고 {_m} · 🔴 락누락 {_lm}")
            disp = wh_df.drop(columns=["_color", "_lockmiss", "창고", "확인여부"]).copy()
            disp["상태"] = [f'{_EMO.get(c, "")} {s}'
                          for c, s in zip(wh_df["_color"], wh_df["상태"])]
            disp.insert(0, "확인", seed)
            _bg = wh_df["_color"].tolist()          # 행별 색 (disp 행 순서와 일치)
            _lmm = wh_df["_lockmiss"].tolist()

            def _row_style(row, bg=_bg, lm=_lmm):
                # 락누락 행은 색 없음이어도 옅은 분홍으로 강조
                c = BG.get(bg[row.name], "#FFFFFF")
                if lm[row.name] and c == "#FFFFFF":
                    c = "#FFD6D6"
                return [f"background-color: {c}"] * len(row)

            # Styler로 행 전체 배경색 유지 + 체크박스 편집 가능
            styled = disp.style.apply(_row_style, axis=1).map(
                lambda _: "font-weight: bold", subset=["등록 유통기한"])
            edited = st.data_editor(
                styled, key=f"ed_{wh}", hide_index=True, width='stretch',
                height=min(600, 60 + len(disp) * 38),
                column_config={"확인": st.column_config.CheckboxColumn(
                    "확인", help="확인 완료 시 체크 — 저장되어 다음 분석에도 유지", default=False)},
                disabled=[c for c in disp.columns if c != "확인"])
            # 변경분만 저장 (data_editor가 편집 시 자동 rerun하므로 강제 rerun 불필요 —
            # 강제 rerun은 저장 실패(테이블 없음) 시 무한 루프를 유발하므로 제거)
            for i, k in enumerate(keys):
                newv = bool(edited.iloc[i]["확인"])
                if newv != seed[i]:
                    set_check(k, newv)

        st.markdown("""
---
**범례**

| 색상 | 의미 |
|------|------|
| 🔴 빨간 (위험) — 역순출고 우려 | 등록 유통기한보다 **이후** 유통기한이 출고 진행 중 |
| 🟠 주황 (주의) — 신규파우치 출고 임박 (락해제 점검 필요) | **동일** 유통기한 출고 진행 중  ·  또는  ·  락 없는 동일 유통기한 재고 없고 **직전** 유통기한 출고 진행 중 |
| ⬜ 색 없음 (정상) | 락 없는 동일 유통기한 재고 있음 |
| ⬛ 회색 (미입고) | 등록 품목/유통기한 재고가 이 창고에 **없음** |
| 🟥 분홍 (락누락) | **전량** 등록인데 락 안 걸린 재고가 있음 — 락 확인 필요 |

**추가 열**
- **리뉴얼구분** : 등록 시 지정한 전량/부분 (미지정이면 재등록해 지정)
- **창고입고** : 등록 (품목·유통기한) 재고가 이 창고에 입고됐는지 (✅ 입고 / ❌ 미입고)
- **리뉴얼(락)** : 리뉴얼구분에 비춘 락 상태
  - **전량** → 🔒 전량 락(정상) / 🔴 **락 누락 의심**(전량인데 미락 재고 있음)
  - **부분** → 🔓 부분(락 해제분 정상)
- **합친수량(Box)** : 지금 나갈 수 있는 재고 = 선입가용(등록보다 빠른 유통기한 출고가능) + 동일미락(동일 유통기한 락 안 걸린 출고가능)
- **일평균출고** : ABC분석 프로그램의 최근 3개월 일평균출고 평균 (품번별)
- **가용일수 / 예상소진일** : 합친수량 ÷ 일평균출고 = 며칠치 남았는지 · 오늘 + 가용일수 = 대략 소진 시점
        """)

        st.markdown("""
<div style="
    background-color: #FFF3CD;
    border: 3px solid #FF0000;
    border-radius: 8px;
    padding: 18px 24px;
    margin-top: 16px;
    text-align: center;
">
    <span style="font-size: 22px; font-weight: 900; color: #CC0000;">
        ⚠️ 락 해제 시 메일 발송 필수
    </span><br>
    <span style="font-size: 17px; font-weight: 700; color: #333;">
        수신 : 영업팀 전부 &nbsp;|&nbsp; 참조 : SCM
    </span>
</div>
""", unsafe_allow_html=True)
