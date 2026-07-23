# -*- coding: utf-8 -*-
"""ABC 재배치 분석 — Streamlit 페이지.

로컬 ABC 재배치 분석기(tkinter)의 웹 이식본.
데이터는 Supabase(`app_settings` 테이블 + base64) 영구 저장.
"""
from __future__ import annotations

import base64
import io
import re
import sys
import traceback
from datetime import datetime
from pathlib import Path

import pandas as pd
import streamlit as st

HERE = Path(__file__).parent
sys.path.insert(0, str(HERE))
import core  # noqa: E402

TMP = HERE / "_tmp"
TMP.mkdir(exist_ok=True)
MIME_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

MASTER_META_KEY = "abc_master_meta"
MASTER_B64_KEY = "abc_master_b64"
INDEX_KEY = "abc_monthly_index"


def _month_key(ym: str) -> str:
    return f"abc_monthly_b64_{ym.replace('-', '_')}"


# ---------- Supabase ----------
try:
    _SB_OK = bool(st.secrets.get("SUPABASE_URL")) and bool(st.secrets.get("SUPABASE_KEY"))
except Exception:
    _SB_OK = False


@st.cache_resource(show_spinner=False)
def _sb():
    from supabase import create_client
    return create_client(st.secrets["SUPABASE_URL"], st.secrets["SUPABASE_KEY"])


def _fetch(key):
    if not _SB_OK:
        return None
    try:
        r = _sb().table("app_settings").select("value").eq("key", key).execute()
        return r.data[0]["value"] if r.data else None
    except Exception:
        return None


def _upsert(key, value):
    if not _SB_OK:
        return False
    try:
        _sb().table("app_settings").upsert(
            {"key": key, "value": value}, on_conflict="key").execute()
        return True
    except Exception as e:
        st.error(f"Supabase 저장 실패: {e}")
        return False


def _delete(key):
    if not _SB_OK:
        return False
    try:
        _sb().table("app_settings").delete().eq("key", key).execute()
        return True
    except Exception:
        return False


def _fetch_master_meta():
    return _fetch(MASTER_META_KEY)


@st.cache_data(show_spinner=False)
def _fetch_master_bytes(version):
    v = _fetch(MASTER_B64_KEY)
    return base64.b64decode(v) if v else None


def _store_master(data: bytes, count: int):
    ok = _upsert(MASTER_B64_KEY, base64.b64encode(data).decode())
    if ok:
        _upsert(MASTER_META_KEY, {
            "updated": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "n": int(count),
        })
    _fetch_master_bytes.clear()
    return ok


def _fetch_index() -> list:
    idx = _fetch(INDEX_KEY) or []
    return sorted(set(idx))


def _store_index(index_list):
    return _upsert(INDEX_KEY, sorted(set(index_list)))


@st.cache_data(show_spinner=False)
def _fetch_month_bytes(ym: str, version):
    v = _fetch(_month_key(ym))
    return base64.b64decode(v) if v else None


def _store_month(ym: str, data: bytes):
    ok = _upsert(_month_key(ym), base64.b64encode(data).decode())
    if ok:
        idx = _fetch_index()
        if ym not in idx:
            idx.append(ym)
            _store_index(idx)
    _fetch_month_bytes.clear()
    return ok


def _delete_month(ym: str):
    _delete(_month_key(ym))
    idx = [x for x in _fetch_index() if x != ym]
    _store_index(idx)
    _fetch_month_bytes.clear()


# ---------- 헬퍼 ----------
def _guess_ym(filename: str) -> str:
    m = re.search(r"(20\d{2})[-_]?(\d{2})", filename)
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    return ""


def _load_workbook_count(data: bytes) -> int:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        n = sum(1 for _ in wb.active.iter_rows(min_row=2, values_only=True)) or 0
        wb.close()
        return n
    except Exception:
        return 0


MASTER_COLS = ["품번", "품명", "하대(박스/팔레트)", "출시월", "현재로케"]


def _normalize_month_file(data: bytes) -> bytes:
    """원본 xlsx를 3컬럼(품번·품명·일평균출고)으로 정규화.
    - 헤더에 '품번'/'품명'/'일평균출고' 있으면 그 컬럼 사용
    - 없으면 C/D/M열(0-based: 2/3/12) 폴백
    - 이미 3컬럼 편집본이면 통과.
    """
    df = pd.read_excel(io.BytesIO(data), sheet_name=0)
    df.columns = [str(c).strip() for c in df.columns]

    def pick(name, fallback_idx):
        for c in df.columns:
            if c == name:
                return c
        if fallback_idx < len(df.columns):
            return df.columns[fallback_idx]
        return None

    c_code = pick("품번", 2)
    c_name = pick("품명", 3)
    c_avg = pick("일평균출고", 12)
    if not (c_code and c_name and c_avg):
        # 원본 그대로 반환 (core가 알아서 처리)
        return data

    out = df[[c_code, c_name, c_avg]].copy()
    out.columns = ["품번", "품명", "일평균출고"]
    out = out[out["품번"].notna()]
    out["품번"] = _norm_code(out["품번"])
    out["일평균출고"] = pd.to_numeric(out["일평균출고"], errors="coerce").fillna(0)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        out.to_excel(w, sheet_name="sheet", index=False)
    return buf.getvalue()


def _norm_code(s):
    return (
        pd.Series(s).astype(str).str.strip()
        .str.replace(r"\.0$", "", regex=True)
    )


def _auto_add_new_items(month_bytes: bytes, ym: str) -> int:
    """월 파일에서 신규 품번을 감지해 마스터에 자동 추가.
    출시월 = 등록되는 월의 YYMM. 반환값: 추가된 신규 품목 수."""
    m = _fetch_master_meta()
    if not m:
        return 0
    try:
        df_monthly = pd.read_excel(io.BytesIO(month_bytes), sheet_name=0)
    except Exception:
        return 0
    df_monthly.columns = [str(c).strip() for c in df_monthly.columns]
    if "품번" not in df_monthly.columns:
        return 0

    master_bytes = _fetch_master_bytes(m.get("updated", ""))
    if not master_bytes:
        return 0
    df_master = pd.read_excel(io.BytesIO(master_bytes))
    df_master.columns = [str(c).strip() for c in df_master.columns]
    for col in MASTER_COLS:
        if col not in df_master.columns:
            df_master[col] = ""
    df_master = df_master[MASTER_COLS]
    df_master["품번"] = _norm_code(df_master["품번"])

    codes = _norm_code(df_monthly["품번"])
    names = (
        df_monthly["품명"].astype(str).str.strip()
        if "품명" in df_monthly.columns else pd.Series([""] * len(codes))
    )
    df_pairs = pd.DataFrame({"품번": codes, "품명": names})
    df_pairs = df_pairs[
        (df_pairs["품번"].str.len() > 0) & (df_pairs["품번"] != "nan")
    ].drop_duplicates(subset=["품번"], keep="first")

    existing = set(df_master["품번"])
    new_pairs = df_pairs[~df_pairs["품번"].isin(existing)].copy()
    if len(new_pairs) == 0:
        return 0

    m = re.match(r"^(\d{4})-(\d{2})$", ym)
    yymm = m.group(1)[2:] + m.group(2) if m else ym
    new_pairs["하대(박스/팔레트)"] = ""
    new_pairs["출시월"] = yymm
    new_pairs["현재로케"] = ""
    new_pairs = new_pairs[MASTER_COLS]

    df_combined = pd.concat([df_master, new_pairs], ignore_index=True)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        df_combined.to_excel(w, sheet_name="location_master", index=False)
    _store_master(buf.getvalue(), len(df_combined))
    return len(new_pairs)


def _apply_현재로케_update(update_bytes: bytes) -> dict:
    """업로드된 (품번, 현재로케) 2컬럼 xlsx를 마스터에 반영.
    반환: {"applied": N, "unmatched": [품번...]}"""
    m = _fetch_master_meta()
    if not m:
        return {"error": "마스터가 등록되지 않았습니다. 먼저 마스터를 업로드하세요."}

    try:
        df_new = pd.read_excel(io.BytesIO(update_bytes), sheet_name=0)
    except Exception as e:
        return {"error": f"파일 읽기 실패: {e}"}
    df_new.columns = [str(c).strip() for c in df_new.columns]
    need = ["품번", "현재로케"]
    miss = [c for c in need if c not in df_new.columns]
    if miss:
        return {"error": f"필수 컬럼 누락: {miss} (양식: 품번, 현재로케)"}

    master_bytes = _fetch_master_bytes(m.get("updated", ""))
    df_master = pd.read_excel(io.BytesIO(master_bytes))
    df_master.columns = [str(c).strip() for c in df_master.columns]
    if "품번" not in df_master.columns or "현재로케" not in df_master.columns:
        return {"error": "마스터에 '품번' 또는 '현재로케' 컬럼이 없습니다."}

    df_new = df_new[need].dropna(subset=["품번"]).copy()
    df_new["품번"] = _norm_code(df_new["품번"])
    df_new["현재로케"] = df_new["현재로케"].astype(str).str.strip()
    df_new = df_new[df_new["현재로케"] != ""]
    df_master["품번"] = _norm_code(df_master["품번"])

    update_map = dict(zip(df_new["품번"], df_new["현재로케"]))
    master_codes = set(df_master["품번"])
    matched = set(update_map) & master_codes
    unmatched = sorted(set(update_map) - master_codes)

    if not matched:
        return {"applied": 0, "unmatched": unmatched,
                "error": "마스터와 일치하는 품번이 하나도 없습니다."}

    mask = df_master["품번"].isin(update_map)
    df_master.loc[mask, "현재로케"] = df_master.loc[mask, "품번"].map(update_map)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        df_master.to_excel(w, sheet_name="location_master", index=False)
    _store_master(buf.getvalue(), len(df_master))
    return {"applied": len(matched), "unmatched": unmatched}


def _make_현재로케_template() -> bytes:
    """품번-현재로케 2컬럼 빈 양식 xlsx 바이트 생성."""
    df = pd.DataFrame(columns=["품번", "현재로케"])
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        df.to_excel(w, sheet_name="현재로케_업데이트", index=False)
    return buf.getvalue()


# ---------- UI ----------
st.title("📊 ABC 재배치 분석")

if not _SB_OK:
    st.error("Supabase 자격증명이 설정되지 않았습니다. `st.secrets`에 SUPABASE_URL/KEY를 등록하세요.")
    st.stop()


# ────────── ① 마스터 관리 ──────────
st.header("① 로케 마스터")
meta = _fetch_master_meta()
c1, c2 = st.columns([3, 2])
with c1:
    if meta:
        st.success(f"등록됨 · 마지막 업데이트 **{meta.get('updated','?')}** · {meta.get('n',0):,}품목")
    else:
        st.warning("아직 등록되지 않았습니다.")
with c2:
    if meta:
        b = _fetch_master_bytes(meta.get("updated", ""))
        if b:
            st.download_button("📥 현재 마스터 다운로드", b, file_name="location_master.xlsx",
                               mime=MIME_XLSX, width='stretch')

up_master = st.file_uploader("교체할 마스터 xlsx 업로드", type=["xlsx"], key="upl_master")
if up_master:
    data = up_master.getvalue()
    n = _load_workbook_count(data)
    if st.button(f"✓ 마스터 등록 ({n:,}행)", type="primary", width='stretch'):
        if _store_master(data, n):
            st.success("등록 완료")
            st.rerun()

# 현재로케만 일괄 업데이트 (재배치 이동 후 반영용)
with st.expander("🔄 현재로케만 일괄 업데이트 (품번·현재로케 2컬럼)"):
    st.caption("실제 창고 이동 후 (품번, 새 현재로케) 목록을 업로드하면 마스터의 해당 품번의 현재로케만 갱신됩니다.")
    st.download_button("📥 빈 양식 다운로드", data=_make_현재로케_template(),
                       file_name="현재로케_업데이트_양식.xlsx", mime=MIME_XLSX,
                       width='stretch', key="dl_loc_tmpl")
    up_loc = st.file_uploader("품번-현재로케 xlsx 업로드", type=["xlsx"], key="upl_loc_update")
    if up_loc and st.button("✓ 현재로케 갱신 적용", type="primary", width='stretch', key="btn_loc_apply"):
        res = _apply_현재로케_update(up_loc.getvalue())
        if res.get("error"):
            st.error(res["error"])
        else:
            st.success(f"적용 {res['applied']:,}개, 미매칭 {len(res['unmatched']):,}개")
            if res["unmatched"]:
                st.caption("미매칭 품번 (앞 20개): " + ", ".join(res["unmatched"][:20]))
            st.rerun()

st.divider()

# ────────── ② 월별 데이터 등록 ──────────
st.header("② 월별 출고 데이터")

# ① 기준월 먼저 선택
now = datetime.now()
default_ym_pick = f"{now.year}-{now.month:02d}"
ym = st.text_input("① 기준월 (YYYY-MM)", value=default_ym_pick,
                   max_chars=7, key="ym_input")
valid = bool(re.match(r"^20\d{2}-\d{2}$", ym or ""))
if not valid:
    st.warning("YYYY-MM 형식으로 입력하세요 (예: 2026-05)")

# ② 파일 업로드
up_month = st.file_uploader(f"② 월별 xlsx 업로드 (기준월 {ym})",
                            type=["xlsx"], key="upl_month",
                            disabled=not valid)
if up_month and valid:
    guess = _guess_ym(up_month.name)
    if guess and guess != ym:
        st.caption(f"⚠ 파일명 추측 `{guess}` ≠ 선택 기준월 `{ym}` — 다시 확인하세요.")
    if st.button(f"✓ {ym} 로 등록/교체", type="primary", width='stretch'):
        try:
            data = _normalize_month_file(up_month.getvalue())
        except Exception as e:
            st.error(f"파일 정규화 실패: {e}")
            st.stop()
        if _store_month(ym, data):
            st.success(f"{ym} 저장 완료 (품번·품명·일평균출고 3컬럼으로 정규화)")
            n_new = _auto_add_new_items(data, ym)
            if n_new > 0:
                st.info(f"신규 품목 {n_new:,}건 자동 추가 (출시월 {ym.replace('-','')[-4:]} 설정)")
            st.rerun()

# 등록된 월 리스트
index = _fetch_index()
if index:
    st.caption(f"등록된 월: **{len(index)}개** ({index[0]} ~ {index[-1]})")
    with st.expander("등록된 월 목록 (삭제 가능)"):
        del_target = st.selectbox("삭제할 월 선택", ["(선택 안함)"] + index[::-1], key="del_sel")
        if del_target != "(선택 안함)":
            if st.button(f"🗑 {del_target} 삭제", key="del_btn"):
                _delete_month(del_target)
                st.success(f"{del_target} 삭제됨")
                st.rerun()
        # 목록 표
        st.dataframe(pd.DataFrame({"기준월": index}), width='stretch',
                     height=min(300, 80 + len(index) * 32), hide_index=True)
else:
    st.info("아직 등록된 월이 없습니다.")

st.divider()

# ────────── ③ 분석 실행 ──────────
st.header("③ ABC 재분류 + 히트맵 생성")

if not (meta and index):
    st.warning("마스터와 최소 1개월 이상 데이터가 필요합니다.")
    st.stop()

ref_ym = st.selectbox("기준월 선택", index[::-1], index=0, key="ref_sel")

if st.button("▶ 분석 실행", type="primary", width='stretch'):
    log_lines = []
    def log(m):
        log_lines.append(m)

    try:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        sess = TMP / f"session_{ts}"
        (sess / "monthly").mkdir(parents=True)
        (sess / "output").mkdir()

        with st.spinner("Supabase에서 데이터 다운로드..."):
            # 마스터
            master_path = sess / "location_master.xlsx"
            master_path.write_bytes(_fetch_master_bytes(meta.get("updated", "")))
            # 월별 (기준월까지만 사용해도 core가 알아서 처리)
            for ym in index:
                b = _fetch_month_bytes(ym, meta.get("updated", ""))
                if b:
                    (sess / "monthly" / f"{ym}.xlsx").write_bytes(b)

        with st.spinner(f"분석 실행 중 (기준월 {ref_ym})..."):
            result = core.run_analysis(
                str(sess / "monthly"),
                str(master_path),
                str(sess / "output"),
                log=log,
            )

        st.success("분석 완료")
        with st.expander("실행 로그", expanded=False):
            st.code("\n".join(log_lines))

        # 결과 파일
        out_files = sorted((sess / "output").glob("*.xlsx"))
        if not out_files:
            st.error("결과 파일이 생성되지 않았습니다.")
        else:
            for f in out_files:
                st.download_button(f"📥 {f.name} 다운로드",
                                   data=f.read_bytes(), file_name=f.name,
                                   mime=MIME_XLSX, width='stretch',
                                   key=f"dl_{f.name}")
            # 요약 시트 미리보기
            try:
                xls = pd.ExcelFile(out_files[0])
                if "요약" in xls.sheet_names:
                    st.subheader("요약")
                    st.dataframe(xls.parse("요약"), width='stretch', hide_index=True)
                if "ABC_리스트" in xls.sheet_names:
                    st.subheader("ABC_리스트 (상위 30)")
                    st.dataframe(xls.parse("ABC_리스트").head(30),
                                 width='stretch', hide_index=True)
            except Exception:
                pass
    except Exception as e:
        st.error(f"오류: {e}")
        st.code(traceback.format_exc())
        with st.expander("실행 로그"):
            st.code("\n".join(log_lines))
