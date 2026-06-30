# -*- coding: utf-8 -*-
"""쿠팡 지정출고 자동매칭 - 매칭 로직 (청크 기반 FEFO 자동배분)

  analyze(inv_path, out_path)         -> 분석 결과(dict)
  export(analysis, sel1, sel2, path)  -> 사용자 선택 반영하여 xlsx 저장
  build(inv_path, out_path)           -> analyze + export(자동선택)  [CLI 호환]

입력:
  1) 로케이션별 재고조회_*.xlsx   -> OV5 락재고 + 현재고 출고가능
  2) 출고진행현황_전체탭_*.xlsx    -> 쿠팡 센터별 주문

부가 입력(자동 탐색, 통합 양식(쿠팡) 파일):
  - 기준정보 시트 -> 배면(하대) 매핑
  - 재고정보 시트 -> OV5 팔레트ID 매핑
"""
import glob
import json
import os
import re
import sys
import warnings
from collections import defaultdict
from datetime import datetime

# ERP export 파일이 기본 스타일 메타데이터를 빼고 저장해서 나오는 경고 무시
warnings.filterwarnings(
    "ignore", message="Workbook contains no default style.*",
    module="openpyxl")

import openpyxl  # noqa: E402
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DOWNLOADS = os.path.join(os.path.expanduser("~"), "Downloads")


def _resolve_output_dir():
    """결과 저장 폴더. 기존 경로가 있으면 그대로, 없으면 ~/Desktop/지정출고."""
    primary = r"C:\Users\sempio\Desktop\새 폴더\지정출고"
    if os.path.isdir(primary):
        return primary
    fallback = os.path.join(os.path.expanduser("~"), "Desktop", "지정출고")
    try:
        os.makedirs(fallback, exist_ok=True)
    except Exception:
        pass
    return fallback


OUTPUT_DIR = _resolve_output_dir()

TARGET_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                           "대상품목.json")
# 마스터(하대) 캐시 — Item_*.xlsx 업로드 시 여기에 영구 저장
MASTER_ITEM_CACHE = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                 "master_item.xlsx")
DEFAULT_TARGETS = [
    {"code": 1015267, "name": "토장450g_8입"},
    {"code": 1015310, "name": "토장찌개양념 910g_온라인전용"},
    {"code": 1015455, "name": "조선고추장 175g_증정용"},
    {"code": 1016160, "name": "질러 부드러운육포 20g"},
    {"code": 1019141, "name": "순작 우리아이 첫 보리차 192g(R_25)"},
    {"code": 2031156, "name": "샘표 우리엄마 돼지고기장조림 110g(R_16)_16개입"},
    {"code": 2031185, "name": "샘표 김치찌개전용꽁치 400g(대만산)_24개입"},
    {"code": 2031186, "name": "샘표 꽁치원터치 400g(대만산)_24개입"},
    {"code": 2054403, "name": "연두링 멸치디포리 140g"},
    {"code": 2091348, "name": "미소된장국 128g(8gx16)"},
    {"code": 2091349, "name": "미소된장국 8gX5"},
]

# 재고조회 컬럼 인덱스
INV_FIXLOC, INV_LOC = 2, 3
INV_CODE, INV_NAME, INV_UOM, INV_IPSU = 4, 5, 6, 7
INV_EXP = 9
INV_AVAIL_BOX = 18
INV_LOCK_BOX = 24

# 출고진행현황 컬럼 인덱스
OUT_CODE, OUT_NAME = 1, 2
OUT_BOX = 4
OUT_SHIPNO = 7
OUT_CUST = 11
OUT_DELIV = 19

# 엑셀 스타일
HDR_FILL = PatternFill("solid", fgColor="2F5496")
HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
SEL_FILL = PatternFill("solid", fgColor="FFF2CC")
WARN_FILL = PatternFill("solid", fgColor="FCE4D6")
OK_FILL = PatternFill("solid", fgColor="E2EFDA")
GRAY_FILL = PatternFill("solid", fgColor="F2F2F2")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

NO_CUST = "(잔여/미배정)"


# ====================================================================
# 대상품목 / 기준정보 / 팔레트
# ====================================================================
def load_targets():
    if os.path.exists(TARGET_FILE):
        try:
            with open(TARGET_FILE, encoding="utf-8") as f:
                items = json.load(f)
            clean = [{"code": int(t["code"]), "name": t.get("name", "")}
                     for t in items]
            return sorted(clean, key=lambda t: t["code"])
        except Exception:
            pass
    save_targets(DEFAULT_TARGETS)
    return sorted([dict(t) for t in DEFAULT_TARGETS], key=lambda t: t["code"])


def save_targets(items):
    clean = [{"code": int(t["code"]), "name": t.get("name", "")}
             for t in items]
    clean.sort(key=lambda t: t["code"])         # 품목코드 순 정렬 저장
    with open(TARGET_FILE, "w", encoding="utf-8") as f:
        json.dump(clean, f, ensure_ascii=False, indent=2)
    return clean


def _find_latest_master():
    """OUTPUT_DIR에서 가장 최신 '통합 지정출고 양식(쿠팡)' xlsx 경로 (없으면 '')."""
    pats = [
        os.path.join(OUTPUT_DIR, "*통합 지정출고 양식(쿠팡)*.xlsx"),
        os.path.join(OUTPUT_DIR, "**", "*통합 지정출고 양식(쿠팡)*.xlsx"),
    ]
    files = []
    for p in pats:
        files += glob.glob(p, recursive=True)
    files = [f for f in files if "~$" not in os.path.basename(f)]
    return max(files, key=os.path.getmtime) if files else ""


_HADAE_CACHE = {}  # (path, mtime) -> {code: 하대}  — 11k행 재파싱 방지(속도)


def _load_hadae_from_item(path):
    """Item_*.xlsx (마스터 캐시) -> {code: 하대(=배면×배단)}. mtime 기준 캐시."""
    if not path or not os.path.isfile(path):
        return {}
    try:
        ckey = (path, os.path.getmtime(path))
    except OSError:
        ckey = None
    if ckey is not None and ckey in _HADAE_CACHE:
        return _HADAE_CACHE[ckey]
    try:
        # ERP export 의 dimension 메타가 깨져있어 read_only 사용 안함
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception:
        return {}
    result = {}
    try:
        ws = wb[wb.sheetnames[0]]
        bm_idx, bd_idx = None, None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                for j, h in enumerate(row):
                    if h == "배면":
                        bm_idx = j
                    elif h == "배단":
                        bd_idx = j
                if bm_idx is None or bd_idx is None:
                    break
                continue
            code = row[0]
            if code is None:
                continue
            try:
                code_i = int(code)
            except Exception:
                continue
            bm = row[bm_idx] if bm_idx < len(row) else None
            bd = row[bd_idx] if bd_idx < len(row) else None
            try:
                if bm is not None and bd is not None and bm != "" and bd != "":
                    h = float(bm) * float(bd)
                    if h > 0:
                        result[code_i] = h
            except (ValueError, TypeError):
                pass
    finally:
        wb.close()
    if ckey is not None:
        _HADAE_CACHE[ckey] = result
    return result


def update_master_cache(src_path):
    """Item_*.xlsx 를 마스터 캐시(MASTER_ITEM_CACHE)로 복사.
    반환: (성공여부, 등록된 품목수)
    """
    import shutil
    if not src_path or not os.path.isfile(src_path):
        return False, 0
    try:
        shutil.copyfile(src_path, MASTER_ITEM_CACHE)
    except Exception:
        return False, 0
    return True, len(_load_hadae_from_item(MASTER_ITEM_CACHE))


def load_item_names():
    """마스터 캐시에서 {code: 품명} 매핑. 없으면 {}."""
    if not os.path.isfile(MASTER_ITEM_CACHE):
        return {}
    try:
        wb = openpyxl.load_workbook(MASTER_ITEM_CACHE, data_only=True)
    except Exception:
        return {}
    result = {}
    try:
        ws = wb[wb.sheetnames[0]]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            code = row[0]
            if code is None or len(row) < 2:
                continue
            try:
                result[int(code)] = str(row[1] or "")
            except Exception:
                pass
    finally:
        wb.close()
    return result


def load_master_data(path=None):
    """하대(배면×배단) + OV5 팔레트 매핑 로드.
    하대 우선순위:
      ① 마스터 캐시 (master_item.xlsx, Item_*.xlsx 형식)
      ② 통합양식 (기준정보 시트의 배면 컬럼) — fallback
    팔레트는 항상 통합양식의 재고정보 시트에서 로드.
    반환: (hadae_map: code->하대, ov5_pallet_map: (code, exp)->[팔레트ID])
    """
    # ① 캐시 우선
    hadae = _load_hadae_from_item(MASTER_ITEM_CACHE)

    # 통합양식: 팔레트 + 하대 fallback
    if path is None:
        path = _find_latest_master()
    pallets = defaultdict(list)
    if path and os.path.isfile(path):
        try:
            wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        except Exception:
            wb = None
        if wb is not None:
            try:
                if not hadae and "기준정보" in wb.sheetnames:
                    ws = wb["기준정보"]
                    bm_idx = None
                    for i, row in enumerate(ws.iter_rows(values_only=True)):
                        if i == 0:
                            for j, h in enumerate(row):
                                if h and "배면" in str(h):
                                    bm_idx = j
                                    break
                            if bm_idx is None:
                                break
                            continue
                        code = row[0]
                        if code is None:
                            continue
                        try:
                            code_i = int(code)
                        except Exception:
                            continue
                        val = row[bm_idx] if bm_idx < len(row) else None
                        if val is not None and val != "":
                            hadae[code_i] = val
                if "재고정보" in wb.sheetnames:
                    ws = wb["재고정보"]
                    for i, row in enumerate(ws.iter_rows(values_only=True)):
                        if i == 0:
                            continue
                        if row[0] != "O" or row[5] != "OV5":
                            continue
                        try:
                            code = int(row[2])
                        except Exception:
                            continue
                        exp = str(row[15]) if row[15] is not None else ""
                        pallet = str(row[18]) if row[18] is not None else ""
                        if pallet:
                            pallets[(code, exp)].append(pallet)
            finally:
                wb.close()
    return hadae, dict(pallets)


# ====================================================================
# 입력 파일 읽기
# ====================================================================
def find_latest(pattern, required=True):
    files = glob.glob(os.path.join(DOWNLOADS, pattern))
    if not files:
        if required:
            sys.exit(f"[오류] Downloads 폴더에 '{pattern}' 파일이 없습니다.")
        return ""
    return max(files, key=os.path.getmtime)


def extract_date(filename):
    m = re.search(r"(\d{8})", os.path.basename(filename))
    return m.group(1) if m else datetime.now().strftime("%Y%m%d")


def _open_fast(path):
    """가능하면 read_only로 빠르게 연다 (대용량 iter_rows 속도↑).
    read_only가 실패하거나 빈 결과면 일반 모드로 폴백."""
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        # ERP export는 dimension 메타가 깨져 read_only가 행을 일찍 끊는다 →
        # reset_dimensions()로 메타 무시하고 데이터 끝까지 읽게 함.
        ws.reset_dimensions()
        return wb, ws, True
    except Exception:
        try:
            wb.close()
        except Exception:
            pass
        wb = openpyxl.load_workbook(path, data_only=True)
        return wb, wb[wb.sheetnames[0]], False


def load_inventory(path):
    wb, ws, _ro = _open_fast(path)
    ov5 = defaultdict(list)
    ov5_name = {}
    avail = defaultdict(lambda: defaultdict(lambda: [0.0, set()]))
    avail_name = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        code = row[INV_CODE]
        if code is None:
            continue
        code = int(code)
        if row[INV_LOC] == "OV5" and row[INV_LOCK_BOX]:
            ov5_name[code] = row[INV_NAME]
            ov5[code].append({"exp": row[INV_EXP], "box": row[INV_LOCK_BOX],
                              "fixloc": row[INV_FIXLOC]})
        av = row[INV_AVAIL_BOX]
        if av:
            avail_name[code] = row[INV_NAME]
            avail[code][row[INV_EXP]][0] += av
            avail[code][row[INV_EXP]][1].add(str(row[INV_LOC]))
    wb.close()
    return ov5, ov5_name, avail, avail_name


def load_kupang_orders(path):
    wb, ws, _ro = _open_fast(path)
    kupang = defaultdict(list)
    kupang_name = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        cust = row[OUT_CUST]
        if not cust or "쿠팡" not in str(cust):
            continue
        code = row[OUT_CODE]
        if code is None:
            continue
        code = int(code)
        kupang_name[code] = row[OUT_NAME]
        kupang[code].append({"cust": cust, "box": row[OUT_BOX] or 0,
                             "ship_no": row[OUT_SHIPNO], "deliv": row[OUT_DELIV]})
    wb.close()
    return kupang, kupang_name


# ====================================================================
# FEFO 그리디 자동 배분
# ====================================================================
def _fefo_allocate(lots, cust_orders):
    """lots: [(exp, locs, lot_box), ...]  유통기한 빠른 순으로 정렬되어 있어야 함.
    cust_orders: {cust: order_box}
    반환: chunks=[(exp, locs, lot_box, cust, qty, lot_idx)], shortages={cust: 부족박스}
    """
    if not cust_orders:
        # 쿠팡 주문이 없으면 배정할 거래처가 없으므로 qty=0 (lot 박스는 lot_box로 별도 표시)
        chunks = [(e, l, b, "", 0, i) for i, (e, l, b) in enumerate(lots)]
        return chunks, {}
    remaining = dict(cust_orders)
    chunks = []
    # 정렬: ① (RC) 포함 센터는 후순위 → ② 잔여주문 큰 순 → ③ 알파벳
    def _prio(c):
        return (1 if "(RC)" in c else 0, -remaining[c], c)
    for lot_idx, (exp, locs, lot_box) in enumerate(lots):
        lot_left = lot_box
        for cust in sorted(remaining, key=_prio):
            if remaining[cust] <= 0 or lot_left <= 0:
                continue
            take = min(lot_left, remaining[cust])
            chunks.append((exp, locs, lot_box, cust, take, lot_idx))
            remaining[cust] -= take
            lot_left -= take
        if lot_left > 0:
            chunks.append((exp, locs, lot_box, "", lot_left, lot_idx))
    shortages = {c: r for c, r in remaining.items() if r > 0}
    return chunks, shortages


# ====================================================================
# 분석
# ====================================================================
def analyze(inv_path, out_path, targets=None, master_path=None,
            avail_override=None):
    """매칭 분석.

    avail_override: {(code, exp_str): new_lot_box} — 대상품목 출고가능 lot박스
        값을 사용자가 화면에서 수정한 경우 그 값으로 덮어씀.
        0/음수면 해당 lot을 제거.
    """
    ov5, ov5_name, avail, avail_name = load_inventory(inv_path)
    kupang, kupang_name = load_kupang_orders(out_path)
    if targets is None:
        targets = load_targets()
    target_map = {int(t["code"]): t.get("name", "") for t in targets}

    baemyeon, ov5_pallet = load_master_data(master_path)

    # 사용자가 화면에서 수정한 lot박스 적용
    if avail_override:
        for (code, exp_s), new_box in avail_override.items():
            try:
                code_i = int(code); exp_s = str(exp_s)
            except Exception:
                continue
            if code_i not in avail:
                continue
            # avail key 는 원본 타입 (str/int) 다양 — 문자열 비교로 찾기
            for orig_exp in list(avail[code_i].keys()):
                if str(orig_exp) == exp_s:
                    if not new_box or new_box <= 0:
                        del avail[code_i][orig_exp]
                    else:
                        avail[code_i][orig_exp][0] = float(new_box)
                    break

    date_str = extract_date(inv_path)
    short = date_str[4:8] if len(date_str) == 8 else date_str

    # ---------- 기능1: OV5 락재고 ----------
    f1 = []
    f1_shortages = {}    # code -> {cust: 부족박스}
    for code in sorted(ov5):
        # lot 집계: (exp, fixloc) -> box
        lot_agg = defaultdict(float)
        for l in ov5[code]:
            lot_agg[(str(l["exp"]), str(l["fixloc"]))] += l["box"]
        lots = sorted([(e, f, b) for (e, f), b in lot_agg.items()])
        orders = kupang.get(code, [])
        cust_orders = defaultdict(float)
        for o in orders:
            cust_orders[o["cust"]] += o["box"]
        cust_orders = dict(cust_orders)
        chunks, shortages = _fefo_allocate(lots, cust_orders)
        if shortages:
            f1_shortages[code] = shortages
        bm = baemyeon.get(code, "")
        for i, (exp, fixloc, lot_box, cust, qty, lot_idx) in enumerate(chunks):
            pallets = ov5_pallet.get((code, exp), [])
            f1.append({
                "rowid": f"F1#{code}#{i}",
                "fn": 1, "code": code, "name": ov5_name[code],
                "exp": exp, "fixloc": fixloc,
                "lot_box": lot_box, "qty": qty,
                "cust": cust, "auto_cust": cust,
                "cust_orders": cust_orders,
                "baemyeon": bm,
                "pallets": ", ".join(pallets),
                "status": ("쿠팡주문 없음" if not orders
                           else "잔여(미배정)" if not cust else "자동 배정"),
            })

    # ---------- 기능2: 대상 품목 출고가능 ----------
    f2 = []
    f2_shortages = {}
    for code in sorted(target_map):
        name = (target_map[code] or kupang_name.get(code)
                or avail_name.get(code) or "(이름 미확인)")
        orders = kupang.get(code, [])
        cust_orders = defaultdict(float)
        for o in orders:
            cust_orders[o["cust"]] += o["box"]
        cust_orders = dict(cust_orders)
        exps = avail.get(code, {})
        lots = [(str(e), ", ".join(sorted(v[1])), v[0])
                for e, v in sorted(exps.items())]
        bm = baemyeon.get(code, "")
        if not lots:
            f2.append({
                "rowid": f"F2#{code}#none",
                "fn": 2, "code": code, "name": name,
                "exp": "-", "locs_str": "",
                "lot_box": 0, "qty": 0,
                "cust": "", "auto_cust": "",
                "cust_orders": cust_orders,
                "baemyeon": bm, "pallets": "",
                "status": "출고가능 없음" if orders else "쿠팡주문 없음",
            })
            if orders and cust_orders:
                f2_shortages[code] = dict(cust_orders)
            continue
        chunks, shortages = _fefo_allocate(lots, cust_orders)
        if shortages:
            f2_shortages[code] = shortages
        for i, (exp, locs_str, lot_box, cust, qty, lot_idx) in enumerate(chunks):
            f2.append({
                "rowid": f"F2#{code}#{i}",
                "fn": 2, "code": code, "name": name,
                "exp": exp, "locs_str": locs_str,
                "lot_box": lot_box, "qty": qty,
                "cust": cust, "auto_cust": cust,
                "cust_orders": cust_orders,
                "baemyeon": bm, "pallets": "",
                "status": ("쿠팡주문 없음" if not orders
                           else "잔여(미배정)" if not cust else "자동 배정"),
            })

    stats = {
        "ov5_items": len({r["code"] for r in f1}),
        "f1_rows": len(f1),
        "f1_shortage_items": len(f1_shortages),
        "target_items": len(target_map),
        "f2_rows": len(f2),
        "f2_shortage_items": len(f2_shortages),
    }

    return {
        "date_short": short,
        "inv_path": inv_path,
        "out_path": out_path,
        "master_path": master_path or _find_latest_master(),
        "default_result_path": os.path.join(
            OUTPUT_DIR, f"{short} 쿠팡 지정출고_자동.xlsx"),
        "targets": targets,
        "f1": f1,
        "f2": f2,
        "f1_shortages": f1_shortages,
        "f2_shortages": f2_shortages,
        "stats": stats,
    }


# ====================================================================
# 엑셀 저장
# ====================================================================
def _style_header(ws, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = BORDER


def _autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _row_fill(status):
    if status == "자동 배정":
        return OK_FILL
    if status == "잔여(미배정)":
        return WARN_FILL
    return GRAY_FILL


def _write_f1_sheet(wb, analysis, sel1):
    ws = wb.create_sheet("기능1_OV5매칭")
    h = ["품목코드", "품명", "유통기한", "lot박스", "고정로케이션",
         "팔레트", "하대(배면×배단)", "거래처", "배정 박스", "상태"]
    ws.append(h)
    _style_header(ws, len(h))
    prev = None
    for rec in analysis["f1"]:
        chosen = sel1.get(rec["rowid"], rec["cust"])
        if not chosen:        # 잔여/미배정 행은 출력 제외
            continue
        cust_disp = chosen
        # 같은 (code, exp, fixloc) 반복 시 좌측 셀들 비우기 (가독성)
        key = (rec["code"], rec["exp"], rec["fixloc"])
        rep = (key == prev)
        prev = key
        ws.append([
            "" if rep else rec["code"],
            "" if rep else rec["name"],
            "" if rep else rec["exp"],
            "" if rep else f"{rec['lot_box']:g}",
            "" if rep else rec["fixloc"],
            "" if rep else rec["pallets"],
            "" if rep else rec["baemyeon"],
            cust_disp, f"{rec['qty']:g}" if rec["qty"] else "", rec["status"],
        ])
        r = ws.max_row
        fill = _row_fill(rec["status"])
        for c in range(1, len(h) + 1):
            ws.cell(row=r, column=c).fill = fill
            ws.cell(row=r, column=c).border = BORDER
        ws.cell(row=r, column=8).fill = SEL_FILL
    _autosize(ws, [11, 30, 12, 9, 14, 22, 10, 28, 10, 14])
    ws.freeze_panes = "A2"


def _write_f2_sheet(wb, analysis, sel2):
    ws = wb.create_sheet("기능2_출고가능매칭")
    h = ["품목코드", "품명", "유통기한", "lot박스", "로케이션",
         "하대(배면×배단)", "거래처", "배정 박스", "상태"]
    ws.append(h)
    _style_header(ws, len(h))

    # 가장 빠른 유통기한 lot 하나로 총 주문을 처리할 수 있는 품목은 제외
    # (FEFO 기준: 선입 lot의 박스 >= 총 주문 박스 → 한 lot으로 충분 → 검토 불필요)
    by_code = defaultdict(list)
    for rec in analysis["f2"]:
        by_code[rec["code"]].append(rec)
    single_lot_ok = set()
    for code, recs in by_code.items():
        total_order = sum(recs[0]["cust_orders"].values())
        if total_order <= 0:
            continue
        exps_with_lot = [(r["exp"], r["lot_box"]) for r in recs if r["exp"] != "-"]
        if not exps_with_lot:
            continue
        earliest_exp = min(e for e, _ in exps_with_lot)
        earliest_lot = next(b for e, b in exps_with_lot if e == earliest_exp)
        if earliest_lot >= total_order:
            single_lot_ok.add(code)

    prev = None
    for rec in analysis["f2"]:
        chosen = sel2.get(rec["rowid"], rec["cust"])
        if not chosen:        # 잔여/미배정 행은 출력 제외
            continue
        if rec["code"] in single_lot_ok:   # 단일 lot으로 처리 가능 → 출력 제외
            continue
        cust_disp = chosen
        key = (rec["code"], rec["exp"])
        rep = (key == prev)
        prev = key
        ws.append([
            "" if rep else rec["code"],
            "" if rep else rec["name"],
            "" if rep else rec["exp"],
            "" if rep else (f"{rec['lot_box']:g}" if rec["lot_box"] else ""),
            "" if rep else rec["locs_str"],
            "" if rep else rec["baemyeon"],
            cust_disp,
            f"{rec['qty']:g}" if rec["qty"] else "",
            rec["status"],
        ])
        r = ws.max_row
        fill = _row_fill(rec["status"])
        for c in range(1, len(h) + 1):
            ws.cell(row=r, column=c).fill = fill
            ws.cell(row=r, column=c).border = BORDER
        ws.cell(row=r, column=7).fill = SEL_FILL
    _autosize(ws, [11, 30, 12, 9, 36, 10, 28, 10, 14])
    ws.freeze_panes = "A2"


def _write_share_sheet(wb, analysis, sel1, sel2):
    """공유 시트 — 기능1(자동 배정)+기능2(기본/2PT) 합본, 거래처별 피킹 공유용.
    · 기능1=파란 밴드 / 기능2=초록 밴드로 구분
    · 거래처 셀 노란 강조
    · 기능2는 유통기한 2개+ 분할 품목만(선입=기본 / 차선=2PT), 품목코드·품명 병합
    """
    BAND1 = PatternFill("solid", fgColor="2E75B6")
    BAND2 = PatternFill("solid", fgColor="538135")
    F1F = PatternFill("solid", fgColor="DDEBF7")
    F2F = OK_FILL
    WB = Font(bold=True, color="FFFFFF", size=11)
    CEN = Alignment(horizontal="center", vertical="center")
    LFT = Alignment(horizontal="left", vertical="center")
    HDR = ["품목코드", "품명", "유통기한", "고정로케이션", "거래처", "배정 박스", "상태"]
    NC = len(HDR)
    ws = wb.create_sheet("공유")
    state = {"r": 1}

    def put_header():
        r = state["r"]
        for c, name in enumerate(HDR, start=1):
            cell = ws.cell(row=r, column=c, value=name)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = CEN
            cell.border = BORDER
        state["r"] += 1

    def put_band(text, fill):
        r = state["r"]
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
        for c in range(1, NC + 1):
            ws.cell(row=r, column=c).fill = fill
            ws.cell(row=r, column=c).border = BORDER
        cell = ws.cell(row=r, column=1, value=text)
        cell.font = WB
        cell.alignment = LFT
        state["r"] += 1

    def write_block(rows, fill):
        """rows: list of [code, name, exp, loc, cust, qty, status]. 같은 code 묶어 1·2열 병합."""
        from itertools import groupby
        for _code, grp in groupby(rows, key=lambda x: x[0]):
            grp = list(grp)
            start = state["r"]
            for vals in grp:
                r = state["r"]
                for c, v in enumerate(vals, start=1):
                    cell = ws.cell(row=r, column=c, value=v)
                    cell.fill = fill
                    cell.border = BORDER
                    cell.alignment = LFT if c == 2 else CEN
                ws.cell(row=r, column=5).fill = SEL_FILL  # 거래처 강조
                state["r"] += 1
            if len(grp) > 1:
                for col in (1, 2):
                    ws.merge_cells(start_row=start, start_column=col,
                                   end_row=state["r"] - 1, end_column=col)
                    ws.cell(row=start, column=col).alignment = (LFT if col == 2 else CEN)

    # ----- 블록1: 기능1 (배정된 것만) -----
    put_band("■ 기능1 · OV5 락재고 — 자동 배정", BAND1)
    put_header()
    f1rows = []
    for rec in analysis["f1"]:
        chosen = sel1.get(rec["rowid"], rec["cust"])
        if not chosen:
            continue
        f1rows.append([rec["code"], rec["name"], rec["exp"], rec["fixloc"],
                       chosen, rec["qty"] if rec["qty"] else "", "자동 배정"])
    if f1rows:
        write_block(f1rows, F1F)
    else:
        ws.cell(row=state["r"], column=1, value="(배정 없음)")
        state["r"] += 1

    state["r"] += 1  # 빈 줄

    # ----- 블록2: 기능2 (유통기한 2개+ 분할만, 기본/2PT) -----
    put_band("■ 기능2 · 출고가능 분할 피킹 — 기본(선입)/2PT(차선)", BAND2)
    put_header()
    by_code = defaultdict(list)
    for rec in analysis["f2"]:
        by_code[rec["code"]].append(rec)
    single_lot_ok = set()
    for code, recs in by_code.items():
        total_order = sum(recs[0]["cust_orders"].values())
        if total_order <= 0:
            continue
        exps_with_lot = [(r["exp"], r["lot_box"]) for r in recs if r["exp"] != "-"]
        if not exps_with_lot:
            continue
        earliest_exp = min(e for e, _ in exps_with_lot)
        earliest_lot = next(b for e, b in exps_with_lot if e == earliest_exp)
        if earliest_lot >= total_order:
            single_lot_ok.add(code)

    f2rows = []
    for code in sorted(by_code):
        if code in single_lot_ok:
            continue
        recs = by_code[code]
        chosen_recs = []
        for rec in recs:
            ch = sel2.get(rec["rowid"], rec["cust"])
            if ch and rec["exp"] != "-":
                chosen_recs.append((rec, ch))
        if not chosen_recs:
            continue
        min_exp = min(rec["exp"] for rec, _ in chosen_recs)
        for rec, ch in sorted(chosen_recs, key=lambda x: x[0]["exp"]):
            status = "기본" if rec["exp"] == min_exp else "2PT"
            f2rows.append([code, rec["name"], rec["exp"], "",
                           ch, rec["qty"] if rec["qty"] else "", status])
    if f2rows:
        write_block(f2rows, F2F)
    else:
        ws.cell(row=state["r"], column=1, value="(분할 품목 없음)")
        state["r"] += 1

    _autosize(ws, [11, 32, 12, 14, 28, 10, 12])
    ws.freeze_panes = "A2"
    note_r = state["r"] + 1
    ws.cell(row=note_r, column=1,
            value="※ 기능1=완제품 자동 배정 / 기능2=유통기한 2개 이상 분할(기본=선입, "
                  "2PT=차선). 거래처(노랑) 확인 후 피킹.").font = Font(italic=True, color="808080")


def _write_shortage_sheet(wb, analysis):
    ws = wb.create_sheet("부족내역")
    h = ["구분", "품목코드", "거래처", "부족 박스"]
    ws.append(h)
    _style_header(ws, len(h))
    for code, sh in sorted(analysis["f1_shortages"].items()):
        for cust, qty in sorted(sh.items()):
            ws.append(["기능1 (OV5)", code, cust, f"{qty:g}"])
            for c in range(1, len(h)+1):
                ws.cell(row=ws.max_row, column=c).fill = WARN_FILL
                ws.cell(row=ws.max_row, column=c).border = BORDER
    for code, sh in sorted(analysis["f2_shortages"].items()):
        for cust, qty in sorted(sh.items()):
            ws.append(["기능2 (출고가능)", code, cust, f"{qty:g}"])
            for c in range(1, len(h)+1):
                ws.cell(row=ws.max_row, column=c).fill = WARN_FILL
                ws.cell(row=ws.max_row, column=c).border = BORDER
    _autosize(ws, [16, 12, 30, 12])
    ws.freeze_panes = "A2"


def _write_summary_sheet(wb, analysis):
    st = analysis["stats"]
    ws = wb.create_sheet("요약", 0)
    info = [
        ["쿠팡 지정출고 자동매칭 결과", ""],
        ["", ""],
        ["기준일자", analysis["date_short"]],
        ["생성시각", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["입력: 재고조회", os.path.basename(analysis["inv_path"])],
        ["입력: 출고진행현황", os.path.basename(analysis["out_path"])],
        ["기준정보/팔레트 소스",
         os.path.basename(analysis["master_path"]) if analysis["master_path"]
         else "(없음 — 하대/팔레트 표시 안 됨)"],
        ["", ""],
        ["[기능1] OV5 락재고 (유통기한별 라인 분할 + FEFO 자동배분)", ""],
        ["  OV5 락재고 품목수", st["ov5_items"]],
        ["  생성된 행 수 (lot×거래처 청크)", st["f1_rows"]],
        ["  거래처 주문 부족 품목수", st["f1_shortage_items"]],
        ["", ""],
        ["[기능2] 대상품목 출고가능 (유통기한별 라인 분할 + FEFO 자동배분)", ""],
        ["  대상 품목수", st["target_items"]],
        ["  생성된 행 수", st["f2_rows"]],
        ["  거래처 주문 부족 품목수", st["f2_shortage_items"]],
        ["", ""],
        ["설명",
         "FEFO(유통기한 빠른 순)로 자동 배분, 행별 거래처는 드롭다운으로 수정 가능."],
        ["",
         "다 채워진 거래처는 다음 행 드롭다운에서 자동 제외됩니다."],
    ]
    for r in info:
        ws.append(r)
    ws["A1"].font = Font(bold=True, size=14)
    for kr in (9, 14):
        ws.cell(row=kr, column=1).font = Font(bold=True, size=11)
    _autosize(ws, [44, 56])


def export(analysis, sel1, sel2, result_path=None):
    if result_path is None:
        result_path = analysis["default_result_path"]
    sel1 = sel1 or {}
    sel2 = sel2 or {}
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _write_summary_sheet(wb, analysis)
    _write_f1_sheet(wb, analysis, sel1)
    _write_f2_sheet(wb, analysis, sel2)
    _write_share_sheet(wb, analysis, sel1, sel2)
    _write_shortage_sheet(wb, analysis)
    wb.save(result_path)
    return result_path


# ====================================================================
# CLI
# ====================================================================
def build(inv_path, out_path):
    a = analyze(inv_path, out_path)
    return export(a, {}, {}), a["stats"]


def main():
    inv = find_latest("로케이션별 재고조회_*.xlsx")
    out = find_latest("출고진행현황_*.xlsx")
    print("[입력] 재고조회      :", os.path.basename(inv))
    print("[입력] 출고진행현황  :", os.path.basename(out))
    a = analyze(inv, out)
    rp = export(a, {}, {})
    st = a["stats"]
    print()
    print(f"[기능1] OV5 {st['ov5_items']}품목, 행 {st['f1_rows']}, "
          f"부족품목 {st['f1_shortage_items']}")
    print(f"[기능2] 대상 {st['target_items']}품목, 행 {st['f2_rows']}, "
          f"부족품목 {st['f2_shortage_items']}")
    print("[완료]", rp)


if __name__ == "__main__":
    main()
