# -*- coding: utf-8 -*-
"""초도물량 차량 배차 (Streamlit 페이지).

초도입고 주문(품목코드·요청박스)을 넣으면:
  1) 로케이션별 재고조회로 출고 창고를 우선순위(예 IC930→IC100→IC920)로 배정
  2) 파레트 혼적 규칙(소형/중형/대형)으로 파레트 구성 — 고정로케이션 피킹 동선 순
  3) 출고 창고별로 차량 분할
하대·품명·고정로케이션은 '공용 기준정보 관리'의 중앙 마스터 사용.
BNF 외 다른 센터 초도물량에도 재사용 가능(우선순위·규칙 조정).
"""
import io
import sys
from datetime import datetime, timezone, timedelta
from pathlib import Path

import openpyxl
import pandas as pd
import streamlit as st

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from chodo import core as C          # noqa: E402
from chodo.excel_io import build_result_bytes    # noqa: E402
from master_hub import store         # noqa: E402

KST = timezone(timedelta(hours=9))
MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

st.title("🚛 초도물량 차량 배차")

hadae_map = store.hadae_map()
name_map = store.name_map()
loc_map = store.loc_map()
_, item_meta = store.load_item()
if not hadae_map:
    st.error("공용 Item 마스터(하대)가 없습니다. 먼저 **'공용 기준정보 관리'** 페이지에서 "
             "Item_*.xlsx 를 업로드하세요.")
    st.stop()
st.caption(f"공용 기준정보: 하대 {len(hadae_map)}품목 · 고정로케이션 {len(loc_map)}품목 "
           f"· Item 갱신 {item_meta.get('갱신','—')}")


# ---------------- 업로드 파서 ----------------
def parse_order(file_bytes: bytes) -> list[dict]:
    """초도입고 파일 → [{제품코드, 요청박스}]. A열 숫자면 데이터로 인식(헤더/빈행 자동 스킵)."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        if row is None or len(row) < 2 or row[0] is None:
            continue
        try:
            code = int(float(row[0]))
        except (TypeError, ValueError):
            continue
        try:
            qty = int(round(float(row[1])))
        except (TypeError, ValueError):
            qty = 0
        if qty > 0:
            rows.append({"제품코드": code, "요청박스": qty})
    wb.close()
    return rows


def parse_inventory(file_bytes: bytes) -> tuple[dict, dict]:
    """로케이션별 재고조회 → (inv_avail, name_fallback).

    inv_avail = {코드: {Inventory: {'box': 출고가능합, 'loc': 고정로케이션}}}
      loc는 공용 고정로케이션 우선, 없으면 재고조회 C열(고정로케이션).
    열: A(0) Inventory, C(2) 고정로케이션, E(4) 제품코드, F(5) 제품명, S(18) 출고가능(Box).
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    avail: dict[int, dict] = {}
    name_fb: dict[int, str] = {}
    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue
        if not row or len(row) < 19:
            continue
        inv = row[0]
        code = row[4]
        if inv is None or code is None:
            continue
        try:
            code = int(float(code))
        except (TypeError, ValueError):
            continue
        try:
            box = float(row[18] or 0)
        except (TypeError, ValueError):
            box = 0.0
        if row[5] and code not in name_fb:
            name_fb[code] = row[5]
        if box <= 0:
            continue
        cfix = row[2]
        d = avail.setdefault(code, {}).setdefault(inv, {"box": 0.0, "loc": None})
        d["box"] += box
        if d["loc"] is None and cfix is not None:
            d["loc"] = str(cfix).strip()
    wb.close()
    # 공용 고정로케이션으로 loc 덮어쓰기(있으면 우선)
    for code, invd in avail.items():
        home = loc_map.get(code)
        if home:
            for inv in invd:
                invd[inv]["loc"] = home
    return avail, name_fb


# ---------------- 입력 ----------------
st.subheader("1️⃣ 입력")
up_order = st.file_uploader("초도입고 예정 품목&수량 (A=품목코드, B=요청박스)",
                            type=["xlsx"], key="c_order")
up_inv = st.file_uploader("로케이션별 재고조회 (출고 창고 배정 + 고정로케이션)",
                          type=["xlsx"], key="c_inv")

with st.expander("⚙️ 옵션 (센터/규칙 조정)", expanded=False):
    prio_txt = st.text_input("출고 우선순위 (쉼표, 앞이 우선)", value="IC930, IC100, IC920")
    o1, o2, o3 = st.columns(3)
    cap = o1.number_input("용차 파레트 한도", 1, 40, 16)
    t1box = o1.number_input("소형 기준(≤박스)", 1, 99, 5)
    t1max = o1.number_input("소형 최대품목", 1, 99, 10)
    t2box = o2.number_input("중형 기준(≤박스)", 1, 99, 24)
    t2max = o2.number_input("중형 최대품목", 1, 99, 3)
    big_pairing = o2.checkbox("대형(25박스+) 페어링 적용", value=True,
                              help="켬: 적재율합 한도 내 최대 2품목 묶음. "
                                   "끔: 25박스 이상은 무조건 1품목=1파레트(별도).")
    t3max = o3.number_input("대형 최대품목", 1, 99, 2, disabled=not big_pairing)
    t3ratio = o3.number_input("대형 적재율합 한도", 0.1, 1.0, 0.8, step=0.05,
                              disabled=not big_pairing)
    pcap = o3.number_input("개수묶음 적재율합 한도", 0.5, 2.0, 1.2, step=0.1)
    gmode_label = o3.selectbox("소형·중형 묶는 기준", ["창고 로케이션 순", "수량(적재율) 순"])
gmode = "location" if gmode_label.startswith("창고") else "quantity"
priority = [s.strip() for s in prio_txt.split(",") if s.strip()]

# ---------------- 산출 ----------------
if st.button("🚛 배차 산출", type="primary", disabled=(up_order is None or up_inv is None)):
    try:
        order = parse_order(up_order.getvalue())
        inv_avail, name_fb = parse_inventory(up_inv.getvalue())
        master = {r["제품코드"]: {
            "제품명": name_map.get(r["제품코드"]) or name_fb.get(r["제품코드"], ""),
            "하대": hadae_map.get(r["제품코드"]),
        } for r in order}
        trucks, pallets, warns, gmeta, shortages = C.dispatch_by_location(
            order, master, inv_avail, priority=priority, voncha_capacity=int(cap),
            tier1_box=int(t1box), tier1_max=int(t1max),
            tier2_box=int(t2box), tier2_max=int(t2max),
            tier3_max=int(t3max), tier3_ratio=float(t3ratio),
            pallet_cap=float(pcap), group_mode=gmode, big_pairing=big_pairing,
        )
        summ = C.summarize(order, pallets, trucks, int(cap), gmeta)
        st.session_state["chodo_res"] = {
            "trucks": C.trucks_to_export(trucks), "summ": summ,
            "warns": warns, "shortages": shortages, "gmeta": gmeta,
            "xlsx": build_result_bytes(C.trucks_to_export(trucks), summ, warns, shortages),
        }
    except Exception as e:
        st.error(f"산출 실패: {e}")

# ---------------- 결과 ----------------
res = st.session_state.get("chodo_res")
if res:
    s = res["summ"]
    m = st.columns(5)
    m[0].metric("총 품목", f"{s['품목수']:,}")
    m[1].metric("총 박스", f"{s['총박스']:,}")
    m[2].metric("총 파레트", f"{s['총파레트']}")
    m[3].metric("차량", f"{s['차량수']}")
    m[4].metric("재고부족", f"{len(res['shortages'])}")

    gcols = st.columns(max(1, len(res["gmeta"])))
    for i, (loc, g) in enumerate(res["gmeta"].items()):
        gcols[i].info(f"**{loc}**\n\n{g['품목수']}품목 · {g['파레트']}파레트 · {g['차량수']}대")

    # 재고부족 별도 알림
    if res["shortages"]:
        with st.expander(f"⚠️ 재고부족 {len(res['shortages'])}품목 — 우선순위 창고에 재고 없음 (별도 시트로 기록)",
                         expanded=True):
            sh = pd.DataFrame([{
                "제품코드": x["제품코드"], "제품명": x["제품명"], "요청박스": x["요청박스"],
                **{f"{k}재고": v for k, v in x["우선순위재고"].items()},
                "기타재고": x["기타재고"],
            } for x in res["shortages"]])
            st.dataframe(sh, width="stretch", hide_index=True)

    # 기타 경고 (하대 없음/기준정보 누락 — 배차에서 제외된 품목)
    other_warns = [w for w in res["warns"] if not w.startswith("재고부족")]
    if other_warns:
        with st.expander(f"❗ 제외된 품목 {len(other_warns)}건 (하대 없음/기준정보 누락) — 공용 마스터 확인 필요"):
            for w in other_warns:
                st.write("· " + w)

    # 배차 결과 표 (평탄화)
    kind_kr = {"full": "꽉참", "big": "대형혼적", "mid": "중형혼적", "small": "소형혼적"}
    flat, prev_loc = [], None
    for t in res["trucks"]:
        for pidx, pal in enumerate(t["pallets"], start=1):
            for j, it in enumerate(pal["items"]):
                flat.append({
                    "출고지": t["loc"] if (t["loc"] != prev_loc and pidx == 1 and j == 0) else "",
                    "차량": f"{t['label']} ({len(t['pallets'])}/{t['capacity']})" if (pidx == 1 and j == 0) else "",
                    "파레트": pidx if j == 0 else "",
                    "유형": kind_kr.get(pal["kind"], pal["kind"]) if j == 0 else "",
                    "제품코드": it["제품코드"], "제품명": it["제품명"],
                    "로케이션": it.get("로케이션", ""), "박스": it["박스"],
                    "하대": it["하대"], "적재율": f"{it['적재율']*100:.0f}%",
                })
        prev_loc = t["loc"]
    st.dataframe(pd.DataFrame(flat), width="stretch", height=460, hide_index=True)

    st.download_button("📥 배차결과 다운로드 (xlsx)", res["xlsx"],
                       file_name=f"초도배차결과_{datetime.now(KST):%y%m%d_%H%M}.xlsx",
                       mime=MIME, width="stretch")
    st.caption("출고지=우선순위 창고 중 주문 전량 커버하는 첫 창고. 창고별로 차량 분리. "
               "혼적=소형(≤5박스,10품목)/중형(6~24,3품목)/대형(25+,적재율합≤0.8,2품목), "
               "고정로케이션 피킹 동선 순. 재고부족은 별도 시트.")
