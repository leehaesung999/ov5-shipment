# -*- coding: utf-8 -*-
"""Excel 입출력 — BNF 초도입고 파일 / 기준정보 읽기, 배차결과 쓰기."""

from __future__ import annotations
import shutil
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter


def read_request(path: str) -> list[dict]:
    """BNF 초도입고 파일 → [{제품코드, 요청박스}].

    형식: 시트 아무거나(active), 헤더에 '품목코드'/'이동요청박스'.
    실제로는 앞쪽 빈 행/헤더를 건너뛰고, A열이 숫자면 데이터로 인식.
    """
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        if row is None or len(row) < 2:
            continue
        code_raw, qty_raw = row[0], row[1]
        if code_raw is None:
            continue
        # 헤더/텍스트 행 스킵: A열이 정수로 변환 가능해야 함
        try:
            code = int(code_raw)
        except (TypeError, ValueError):
            continue
        try:
            qty = int(round(float(qty_raw)))
        except (TypeError, ValueError):
            qty = 0
        if qty <= 0:
            continue
        rows.append({"제품코드": code, "요청박스": qty})
    wb.close()
    return rows


def read_inventory(path: str) -> dict[int, dict[str, dict]]:
    """로케이션별 재고조회 xlsx → {제품코드: {Inventory: {'box': 합, 'loc': 고정로케이션}}}.

    컬럼: A Inventory, B Inventory명, C 고정로케이션(idx2), D Location(idx3),
          E 제품코드, F 제품명, ..., S 출고가능(Box)=idx18.
    같은 (코드, Inventory)의 여러 로케이션/로트는 box 합산.
    'loc'(피킹 동선 정렬용)은 **고정로케이션(C, 홈 위치)** — 품목당 고정값.
    (D Location은 실물 재고가 흩어진 빈이라 동선 기준으로 부적합.)
    """
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    avail: dict[int, dict[str, dict]] = {}
    header_skipped = False
    for row in ws.iter_rows(values_only=True):
        if not header_skipped:
            header_skipped = True
            continue
        inv = row[0]
        fixed_loc = row[2] if len(row) > 2 else None   # 고정로케이션
        code = row[4] if len(row) > 4 else None
        if inv is None or code is None:
            continue
        try:
            code = int(float(code))
        except (TypeError, ValueError):
            continue
        box = row[18] if len(row) > 18 else 0
        try:
            box = float(box or 0)
        except (TypeError, ValueError):
            box = 0.0
        if box <= 0:
            continue
        d = avail.setdefault(code, {}).setdefault(inv, {"box": 0.0, "loc": None})
        d["box"] += box
        if d["loc"] is None and fixed_loc is not None:
            d["loc"] = fixed_loc      # 고정로케이션은 품목당 일정 — 최초값 사용
    wb.close()
    return avail


def read_master(path: str) -> dict[int, dict]:
    """물품정보 xlsx → {제품코드: {제품명, 하대}}.

    이 기준정보는 하대(AF, idx31)가 비어있고 배면(AD,29)×배단(AE,30)=하대.
    AF가 채워져 있으면 AF 우선.
    """
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    master: dict[int, dict] = {}
    header_skipped = False
    for row in ws.iter_rows(values_only=True):
        if not header_skipped:
            header_skipped = True
            continue
        code = row[0]
        if code is None:
            continue
        try:
            code = int(code)
        except (TypeError, ValueError):
            continue
        af = row[31] if len(row) > 31 else None
        bm = row[29] if len(row) > 29 else None  # 배면
        bd = row[30] if len(row) > 30 else None  # 배단
        if af:
            hadae = int(af)
        elif bm and bd:
            hadae = int(round(float(bm) * float(bd)))
        else:
            hadae = None
        master[code] = {
            "제품명": row[1],
            "배면": bm,
            "배단": bd,
            "하대": hadae,
        }
    wb.close()
    return master


def save_master_copy(src: str, dst: str) -> None:
    Path(dst).parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src, dst)


# ---------------- 결과 쓰기 ----------------

_THIN = Side(style="thin", color="888888")
BORDER_ALL = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
FULL_FILL = PatternFill("solid", fgColor="E2EFDA")   # 꽉참 연녹
MIX_FILL = PatternFill("solid", fgColor="FFF2CC")    # 혼적 연노랑
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

KIND_LABEL = {
    "full": "꽉참",
    "big": "대형혼적",
    "mid": "중형혼적",
    "small": "소형혼적",
}


def build_result_bytes(trucks: list, summary: dict, warnings: list[str],
                       shortages: list[dict] | None = None) -> bytes:
    """배차 결과 xlsx를 메모리(BytesIO)로 만들어 bytes 반환 (Streamlit 다운로드용)."""
    import io
    buf = io.BytesIO()
    _write_workbook(buf, trucks, summary, warnings, shortages or [])
    return buf.getvalue()


def write_result(path: str, trucks: list, summary: dict, warnings: list[str],
                 shortages: list[dict] | None = None) -> None:
    """배차 결과를 Excel 파일로 저장 (데스크톱용)."""
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    _write_workbook(path, trucks, summary, warnings, shortages or [])


def _write_workbook(target, trucks: list, summary: dict, warnings: list[str],
                    shortages: list[dict]) -> None:
    """target: 파일경로 또는 file-like. 워크북을 구성해 저장.

    trucks: [{label, capacity, loc, pallets:[{kind, items:[{제품코드,제품명,로케이션,박스,하대,적재율}]}]}]
    shortages: 재고부족 품목 상세 (있으면 '⚠재고부족' 시트로 별도 기록).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "배차결과"

    headers = ["출고지", "차량", "파레트", "유형", "제품코드", "제품명", "로케이션", "박스", "하대", "적재율"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
        cell.alignment = CENTER
        cell.fill = HEADER_FILL
        cell.border = BORDER_ALL

    r = 2
    loc_spans: list[tuple[str, int, int]] = []   # (loc, start_row, end_row)
    cur_loc = None
    loc_start = 2
    for truck in trucks:
        pallets = truck.get("pallets", [])
        if not pallets:
            continue
        loc = truck.get("loc", "")
        if cur_loc is None:
            cur_loc = loc
            loc_start = r
        elif loc != cur_loc:
            loc_spans.append((cur_loc, loc_start, r - 1))
            cur_loc = loc
            loc_start = r

        truck_start = r
        for pidx, pal in enumerate(pallets, start=1):
            items = pal.get("items", [])
            if not items:
                continue
            pal_start = r
            fill = FULL_FILL if pal["kind"] == "full" else MIX_FILL
            for it in items:
                ws.cell(row=r, column=5, value=it["제품코드"]).alignment = CENTER
                ws.cell(row=r, column=6, value=it["제품명"]).alignment = LEFT
                ws.cell(row=r, column=7, value=it.get("로케이션", "")).alignment = CENTER
                ws.cell(row=r, column=8, value=it["박스"]).alignment = CENTER
                ws.cell(row=r, column=9, value=it["하대"]).alignment = CENTER
                rc = ws.cell(row=r, column=10, value=round(it["적재율"], 2))
                rc.alignment = CENTER
                rc.number_format = "0%"
                for col in range(1, 11):
                    ws.cell(row=r, column=col).border = BORDER_ALL
                r += 1
            pal_end = r - 1
            # 파레트 번호(3) + 유형(4) 머지
            ws.cell(row=pal_start, column=3, value=pidx)
            ws.cell(row=pal_start, column=4, value=KIND_LABEL.get(pal["kind"], pal["kind"]))
            for col in (3, 4):
                if pal_end > pal_start:
                    ws.merge_cells(start_row=pal_start, start_column=col,
                                   end_row=pal_end, end_column=col)
                cell = ws.cell(row=pal_start, column=col)
                cell.alignment = CENTER
                cell.fill = fill
                cell.border = BORDER_ALL
        truck_end = r - 1
        # 차량 라벨(2) 머지
        used = len(pallets)
        cap = truck.get("capacity", 0)
        ws.cell(row=truck_start, column=2, value=f"{truck['label']} ({used}/{cap})")
        if truck_end > truck_start:
            ws.merge_cells(start_row=truck_start, start_column=2,
                           end_row=truck_end, end_column=2)
        tc = ws.cell(row=truck_start, column=2)
        tc.alignment = CENTER
        tc.font = Font(bold=True)
        tc.border = BORDER_ALL
    if cur_loc is not None:
        loc_spans.append((cur_loc, loc_start, r - 1))

    # 출고지(1) 머지
    for loc, s0, e0 in loc_spans:
        ws.cell(row=s0, column=1, value=loc)
        if e0 > s0:
            ws.merge_cells(start_row=s0, start_column=1, end_row=e0, end_column=1)
        lc = ws.cell(row=s0, column=1)
        lc.alignment = CENTER
        lc.font = Font(bold=True, size=12)
        lc.border = BORDER_ALL

    widths = {1: 12, 2: 14, 3: 8, 4: 12, 5: 12, 6: 38, 7: 12, 8: 8, 9: 8, 10: 8}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"

    # ---- 요약 시트 ----
    ws2 = wb.create_sheet("요약")
    lines = [
        ("총 품목수", summary.get("품목수")),
        ("총 박스", summary.get("총박스")),
        ("총 파레트", summary.get("총파레트")),
        ("  · 꽉찬 파레트", summary.get("full")),
        ("  · 소형혼적(≤5박스)", summary.get("small")),
        ("  · 중형혼적(6~24박스)", summary.get("mid")),
        ("  · 대형혼적(25박스+)", summary.get("big")),
        ("차량 대수", summary.get("차량수")),
        ("차량 파레트 한도", summary.get("용차한도")),
    ]
    ws2.cell(row=1, column=1, value="항목").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="값").font = Font(bold=True)
    for i, (k, v) in enumerate(lines, start=2):
        ws2.cell(row=i, column=1, value=k)
        ws2.cell(row=i, column=2, value=v).alignment = CENTER
    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 12

    # 출고지별 표
    groups = summary.get("groups") or {}
    gbase = len(lines) + 4
    if groups:
        ws2.cell(row=gbase, column=1, value="출고지").font = Font(bold=True)
        for j, h in enumerate(["출고지", "품목", "박스", "파레트", "차량"]):
            gc = ws2.cell(row=gbase, column=1 + j, value=h)
            gc.font = Font(bold=True)
            gc.fill = HEADER_FILL
        for i, (loc, m) in enumerate(groups.items(), start=gbase + 1):
            ws2.cell(row=i, column=1, value=loc)
            ws2.cell(row=i, column=2, value=m.get("품목수"))
            ws2.cell(row=i, column=3, value=m.get("박스"))
            ws2.cell(row=i, column=4, value=m.get("파레트"))
            ws2.cell(row=i, column=5, value=m.get("차량수"))
        for col in ("C", "D", "E"):
            ws2.column_dimensions[col].width = 10

    if warnings:
        base = gbase + len(groups) + 3
        ws2.cell(row=base, column=1, value="경고 (기준정보 누락 등)").font = Font(bold=True, color="AA0000")
        for i, w in enumerate(warnings, start=base + 1):
            ws2.cell(row=i, column=1, value=w)

    # ---- 재고부족 전용 시트 ----
    if shortages:
        ws3 = wb.create_sheet("⚠재고부족")
        shortfill = PatternFill("solid", fgColor="FCE4E4")
        # 우선순위 로케이션 표시명 (첫 항목 기준)
        prio_keys = list(shortages[0]["우선순위재고"].keys())
        heads = ["제품코드", "제품명", "요청박스"] + [f"{k}재고" for k in prio_keys] + ["기타재고합", "비고"]
        for c, h in enumerate(heads, start=1):
            cell = ws3.cell(row=1, column=c, value=h)
            cell.font = Font(bold=True)
            cell.alignment = CENTER
            cell.fill = HEADER_FILL
            cell.border = BORDER_ALL
        for i, sh in enumerate(shortages, start=2):
            vals = [sh["제품코드"], sh["제품명"], sh["요청박스"]]
            vals += [sh["우선순위재고"].get(k, 0) for k in prio_keys]
            note = "다른 창고에 재고 있음(확인)" if sh.get("기타재고", 0) > 0 else "전 창고 재고 없음"
            vals += [sh.get("기타재고", 0), note]
            for c, v in enumerate(vals, start=1):
                cell = ws3.cell(row=i, column=c, value=v)
                cell.alignment = LEFT if c == 2 else CENTER
                cell.border = BORDER_ALL
                cell.fill = shortfill
        widths3 = [12, 40, 10] + [9] * len(prio_keys) + [11, 22]
        for c, w in enumerate(widths3, start=1):
            ws3.column_dimensions[get_column_letter(c)].width = w
        ws3.freeze_panes = "A2"

    wb.save(target)
